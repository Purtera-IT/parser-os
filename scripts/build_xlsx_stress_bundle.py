"""Adversarial XLSX stress bundle.

Every shape a real managed-services spreadsheet throws at a parser:

  xa_bom_simple.xlsx            Single-sheet BOM, clean columns
  xb_bom_multi_sheet.xlsx       Multiple sheets (BOM / Pricing / Sites)
  xc_hidden_sheets.xlsx         One visible sheet + one hidden (must catch hidden)
  xd_formula_cells.xlsx         Formulas everywhere (=A1+B1 etc) — extract VALUES
  xe_merged_cells.xlsx          Merged header cells, merged data cells
  xf_named_ranges.xlsx          Uses defined names (BOM!Items, Sites!Codes)
  xg_chart_only.xlsx            Sheet that's mostly a chart with title cell
  xh_pivot_table.xlsx           Pivot-style summary table
  xi_data_validation.xlsx       Dropdowns / data validation lists
  xj_huge_table_500_rows.xlsx   500-row BOM for stress
  xk_mixed_currency.xlsx        $/€/£ in same workbook
  xl_site_roster.xlsx           Authoritative site_roster as a sheet
  xm_blank_first_row.xlsx       Headers on row 3 (blank rows above)
  xn_multi_section_in_sheet.xlsx  Two tables in one sheet (BOM + Notes)

Run:
  python scripts/build_xlsx_stress_bundle.py <out_dir>
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Alignment


def _save(wb: Workbook, path: Path) -> Path:
    wb.save(path)
    return path


def _style_header_row(ws, n_cols: int):
    fill = PatternFill("solid", fgColor="DDE6F0")
    bold = Font(bold=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = bold


def xa_bom_simple(out: Path) -> Path:
    p = out / "xa_bom_simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    headers = ["Site ID", "Part Number", "Description", "Qty", "Unit Price", "Total"]
    rows = [
        ["ATL-HQ-01",   "C9300-48P-A",        "48-port PoE+ switch",     5,  3500, 17500],
        ["ATL-HQ-01",   "WAP-9180AX-K9",      "Wi-Fi 6E AP",            50, 1200, 60000],
        ["ATL-WEST-02", "C9300-48P-A",        "48-port PoE+ switch",     3,  3500, 10500],
        ["ATL-WEST-02", "WAP-9180AX-K9",      "Wi-Fi 6E AP",            30, 1200, 36000],
        ["ATL-AIR-03",  "SFP-10G-LR-S=",      "10GBASE-LR module",      12,  420,  5040],
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    _style_header_row(ws, len(headers))
    return _save(wb, p)


def xb_bom_multi_sheet(out: Path) -> Path:
    p = out / "xb_bom_multi_sheet.xlsx"
    wb = Workbook()
    wb.active.title = "BOM"
    ws = wb["BOM"]
    ws.append(["Part", "Qty", "Unit Price"])
    for r in [["C9300-48P-A", 5, 3500], ["WAP-9180AX-K9", 50, 1200]]:
        ws.append(r)

    ws_pricing = wb.create_sheet("Pricing")
    ws_pricing.append(["Tier", "Monthly Fee USD", "Response P1"])
    for r in [["Bronze", 2500, "4 hr"], ["Silver", 5000, "2 hr"], ["Gold", 8500, "1 hr"]]:
        ws_pricing.append(r)

    ws_sites = wb.create_sheet("Sites")
    ws_sites.append(["Site ID", "Facility", "Address"])
    for r in [
        ["ATL-HQ-01",   "Atlanta HQ",       "1200 Peachtree St NE, Atlanta GA"],
        ["ATL-WEST-02", "West Campus",      "3100 Interstate N Pkwy, Atlanta GA"],
        ["ATL-AIR-03",  "Airport Logistics", "6000 N Terminal Pkwy, Atlanta GA"],
    ]:
        ws_sites.append(r)

    for ws in wb.worksheets:
        _style_header_row(ws, ws.max_column)
    return _save(wb, p)


def xc_hidden_sheets(out: Path) -> Path:
    p = out / "xc_hidden_sheets.xlsx"
    wb = Workbook()
    wb.active.title = "Visible BOM"
    wb["Visible BOM"].append(["Part", "Qty"])
    wb["Visible BOM"].append(["WAP-9180AX-K9", 50])

    hidden = wb.create_sheet("HiddenSites")
    hidden.append(["Site ID", "Facility"])
    hidden.append(["ATL-HQ-01", "Atlanta HQ"])
    hidden.append(["ATL-WEST-02", "West Campus"])
    hidden.sheet_state = "hidden"
    return _save(wb, p)


def xd_formula_cells(out: Path) -> Path:
    p = out / "xd_formula_cells.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws.append(["Item", "Qty", "Unit Price", "Total"])
    ws.append(["WAP", 50, 1200, "=B2*C2"])
    ws.append(["Switch", 5, 3500, "=B3*C3"])
    ws.append(["TOTAL", "=SUM(B2:B3)", None, "=SUM(D2:D3)"])
    return _save(wb, p)


def xe_merged_cells(out: Path) -> Path:
    p = out / "xe_merged_cells.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged BOM"
    ws.append(["Atlanta Refresh BOM", None, None, None])
    ws.merge_cells("A1:D1")
    ws.append(["Site ID", "Part Number", "Description", "Qty"])
    ws.append(["ATL-HQ-01", "WAP-9180AX-K9", "Wi-Fi 6E AP", 50])
    ws.append(["ATL-HQ-01", "C9300-48P-A", "48-port switch", 5])
    return _save(wb, p)


def xf_named_ranges(out: Path) -> Path:
    p = out / "xf_named_ranges.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(["Part", "Qty"])
    ws.append(["WAP-9180AX-K9", 50])
    ws.append(["C9300-48P-A", 5])
    # Define a name pointing at the BOM range
    dn = DefinedName("BomItems", attr_text="BOM!$A$1:$B$3")
    wb.defined_names["BomItems"] = dn
    return _save(wb, p)


def xg_chart_only(out: Path) -> Path:
    p = out / "xg_chart_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarterly Spend"
    ws.append(["Quarter", "Spend USD"])
    ws.append(["Q1 2026", 125000])
    ws.append(["Q2 2026", 245000])
    ws.append(["Q3 2026", 180000])
    ws.append(["Q4 2026", 90000])
    chart = BarChart()
    chart.title = "FY26 Spend by Quarter"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=5))
    ws.add_chart(chart, "D2")
    return _save(wb, p)


def xh_pivot_table(out: Path) -> Path:
    p = out / "xh_pivot_table.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Site x Device Pivot"
    ws.append(["", "Switches", "Access Points", "Cameras", "Total"])
    ws.append(["ATL-HQ-01",   5,  50, 12, 67])
    ws.append(["ATL-WEST-02", 3,  30,  8, 41])
    ws.append(["ATL-AIR-03",  2,  20, 15, 37])
    ws.append(["TOTAL",       10, 100, 35, 145])
    return _save(wb, p)


def xi_data_validation(out: Path) -> Path:
    p = out / "xi_data_validation.xlsx"
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = Workbook()
    ws = wb.active
    ws.title = "Site Decisions"
    ws.append(["Site ID", "Decision", "Approved By"])
    ws.append(["ATL-HQ-01", "Approved", "Facilities"])
    ws.append(["ATL-WEST-02", "Approved", "Facilities"])
    ws.append(["ATL-AIR-03", "Pending",  "Security"])
    dv = DataValidation(type="list", formula1='"Approved,Pending,Rejected,Deferred"')
    dv.add("B2:B100")
    ws.add_data_validation(dv)
    return _save(wb, p)


def xj_huge_table_500_rows(out: Path) -> Path:
    p = out / "xj_huge_table_500_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BigBOM"
    ws.append(["Row", "Part Number", "Qty"])
    for i in range(500):
        ws.append([i + 1, f"PART-2026-{i+1:04d}", (i + 1) * 2])
    return _save(wb, p)


def xk_mixed_currency(out: Path) -> Path:
    p = out / "xk_mixed_currency.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Region Spend"
    ws.append(["Region", "Item", "Currency", "Amount"])
    ws.append(["US",   "Wi-Fi APs",   "USD", 60000])
    ws.append(["EMEA", "Wi-Fi APs",   "EUR", 55000])
    ws.append(["UK",   "Wi-Fi APs",   "GBP", 47000])
    ws.append(["JPY",  "Wi-Fi APs",   "JPY", 8400000])
    return _save(wb, p)


def xl_site_roster(out: Path) -> Path:
    p = out / "xl_site_roster.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Site Roster"
    ws.append(["Site ID", "Facility name", "Street address", "MDF / IDF", "Access window", "Escort owner"])
    for r in [
        ["ATL-HQ-01",   "OPTBOT Atlanta HQ",        "1200 Peachtree St NE, Atlanta GA 30309",     "MDF-3A / IDF 2-7",  "Mon-Fri 07:00-18:00", "OPTBOT Facilities"],
        ["ATL-WEST-02", "OPTBOT West Campus",       "3100 Interstate N Pkwy, Atlanta GA 30339",   "MDF-W1 / IDF W2-3", "Mon-Fri 07:00-18:00", "OPTBOT Facilities"],
        ["ATL-AIR-03",  "OPTBOT Airport Logistics", "6000 N Terminal Pkwy, Atlanta GA 30320",     "MDF-A / IDF A1",     "Mon-Sat 06:00-22:00", "OPTBOT Security"],
        ["ATL-047-04",  "OPTBOT Brady Training",    "047 Brady Ave NW, Atlanta GA 30318",         "MDF-B / IDF B1-2",   "Mon-Fri 08:00-17:00", "OPTBOT Facilities"],
        ["ATL-CP-05",   "OPTBOT College Park",      "1850 Sullivan Rd, College Park GA 30337",    "MDF-CP / staging",   "Mon-Fri 07:00-15:00", "OPTBOT Logistics"],
    ]:
        ws.append(r)
    _style_header_row(ws, 6)
    return _save(wb, p)


def xm_blank_first_row(out: Path) -> Path:
    p = out / "xm_blank_first_row.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Late Headers"
    ws.append([])
    ws.append([])
    ws.append(["BOM v2", None, None])
    ws.append(["Part", "Qty", "Price"])
    ws.append(["WAP-9180AX-K9", 50, 1200])
    ws.append(["C9300-48P-A", 5, 3500])
    return _save(wb, p)


def xn_multi_section_in_sheet(out: Path) -> Path:
    p = out / "xn_multi_section_in_sheet.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"
    ws.append(["Hardware BOM"])
    ws.append(["Part", "Qty"])
    ws.append(["WAP-9180AX-K9", 50])
    ws.append(["C9300-48P-A", 5])
    ws.append([])
    ws.append(["Notes / Constraints"])
    ws.append(["Item", "Note"])
    ws.append(["Cutover", "Saturdays only at ATL-HQ-01"])
    ws.append(["Escort", "OPTBOT Facilities for HQ; OPTBOT Security for AIR"])
    return _save(wb, p)


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python scripts/build_xlsx_stress_bundle.py <out_dir>", file=sys.stderr)
        return 2
    out = Path(sys.argv[1]).resolve() / "artifacts"
    out.mkdir(parents=True, exist_ok=True)
    builders = [
        xa_bom_simple, xb_bom_multi_sheet, xc_hidden_sheets,
        xd_formula_cells, xe_merged_cells, xf_named_ranges,
        xg_chart_only, xh_pivot_table, xi_data_validation,
        xj_huge_table_500_rows, xk_mixed_currency, xl_site_roster,
        xm_blank_first_row, xn_multi_section_in_sheet,
    ]
    for b in builders:
        p = b(out)
        print(f"  -> {p.name}")
    print(f"\n{len(builders)} XLSX adversarial files in {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
