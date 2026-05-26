"""100-package stress generator across 20 real-shaped managed-services domains.

Each domain creates a realistic multi-artifact bundle:
  - SOW (PDF)
  - BOM (XLSX or PDF)
  - Site roster (PDF / XLSX)
  - Pricing acceptance (PDF)
  - Schedule (PDF / XLSX)

Domains:
  k12_school_district      Refresh of network across 8 schools
  hospital_lan_refresh     Critical-care LAN modernization
  retail_chain_pos         POS rollout across 142 stores
  hospitality_hotel_wifi   Hotel chain Wi-Fi 6E refresh
  fed_gov_datacenter       Federal datacenter consolidation
  university_campus_av     University AV/AVoIP across 6 buildings
  museum_security          Museum CCTV + access control
  airport_terminal_fiber   Airport terminal fiber backbone
  stadium_paging           Stadium paging + PA
  oil_gas_remote           Oil&gas remote-site connectivity
  utilities_scada          Utility SCADA + IoT sensors
  banking_branch_refresh   Bank branch network refresh
  manufacturing_floor      Manufacturing-floor IoT + cameras
  telco_macro_site         Telco macro-site DAS + small cells
  retail_loss_prevention   Retail LP cameras + AI analytics
  k12_classroom_av         K-12 classroom AV refresh
  warehouse_wms            Warehouse WMS network + scanners
  campus_access_control    Campus access control + badging
  data_center_imac         Data center IMAC + decom
  cruise_ship_av           Cruise-ship AV + crew network
"""
from __future__ import annotations

import sys
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


_S = getSampleStyleSheet()
_T = ParagraphStyle("t", parent=_S["Title"], fontSize=14, spaceAfter=10)
_H2 = ParagraphStyle("h2", parent=_S["Heading2"], fontSize=11, spaceBefore=8, spaceAfter=4)
_BODY = ParagraphStyle("b", parent=_S["Normal"], fontSize=10, leading=13, spaceAfter=6)
_CELL = ParagraphStyle("c", parent=_S["Normal"], fontSize=9)


def _tbl(headers, rows, widths_in):
    data = [[Paragraph(h, _CELL) for h in headers]] + [
        [Paragraph(str(c or ""), _CELL) for c in r] for r in rows
    ]
    t = Table(data, colWidths=[w * inch for w in widths_in], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dde6f0")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return t


def _pdf(path: Path, story, landscape_mode=False):
    pagesize = landscape(LETTER) if landscape_mode else LETTER
    doc = SimpleDocTemplate(
        str(path), pagesize=pagesize,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
    )
    doc.build(story)


def _xlsx_bom(path: Path, rows: list[list], headers: list[str]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    fill = PatternFill("solid", fgColor="DDE6F0")
    bold = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = bold
    wb.save(path)
    return path


def _site_roster_pdf(path: Path, deal_title: str, sites: list[list[str]]) -> Path:
    story = [
        Paragraph(deal_title, _T),
        Paragraph("Authoritative Site Roster", _H2),
        Paragraph("kind=physical_site for all rows below.", _BODY),
        _tbl(
            ["Site ID", "Facility name", "Street address", "MDF / IDF", "Access window", "Escort owner"],
            sites,
            [0.85, 1.55, 2.65, 1.55, 1.5, 1.55],
        ),
    ]
    _pdf(path, story, landscape_mode=True)
    return path


def _sow_pdf(path: Path, deal_title: str, scope_lines: list[str], total_usd: int, milestones: list[tuple[str, str]]) -> Path:
    story = [Paragraph(deal_title, _T), Paragraph("Statement of Work", _H2)]
    for line in scope_lines:
        story.append(Paragraph(line, _BODY))
    story.append(Paragraph("Milestones", _H2))
    for m, d in milestones:
        story.append(Paragraph(f"{m}: {d}.", _BODY))
    story.append(Paragraph("Contract Value", _H2))
    story.append(Paragraph(f"Total: USD ${total_usd:,}.00 fixed-price.", _BODY))
    _pdf(path, story)
    return path


def _sla_pdf(path: Path, title: str, sla_lines: list[str]) -> Path:
    story = [Paragraph(title, _T), Paragraph("Service Level Agreement", _H2)]
    for line in sla_lines:
        story.append(Paragraph(line, _BODY))
    _pdf(path, story)
    return path


# Domain builders — each returns the bundle directory it created

def domain_k12_school_district(out: Path) -> Path:
    base = out / "k12_school_district"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["ELM-K-01",  "Elm Elementary",        "100 Oak Ave, Springfield IL 62701",        "MDF-1A / IDF 1-3", "Mon-Fri 07:00-16:00", "Facilities"],
        ["MAP-K-02",  "Maple Middle School",   "200 Pine St, Springfield IL 62702",        "MDF-2A / IDF 2-4", "Mon-Fri 07:00-16:00", "Facilities"],
        ["OAK-K-03",  "Oak Ridge High",        "300 Birch Rd, Springfield IL 62703",       "MDF-3A / IDF 3-6", "Mon-Fri 07:00-17:00", "Security"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Springfield USD-186 Refresh 2026", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Springfield USD-186 Refresh 2026",
        [
            "Install 320 wireless access points across 3 schools.",
            "Install 22 Cisco C9300-48P-A distribution switches.",
            "Provide 12 months of T2 managed service at Silver tier.",
        ],
        total_usd=1_245_000,
        milestones=[("Kickoff", "Mar 15, 2026"), ("ATP", "August 30, 2026")],
    )
    _xlsx_bom(art / "03_bom.xlsx",
        [
            ["ELM-K-01",  "WAP-9180AX-K9", "Wi-Fi 6E AP", 90, 1200],
            ["MAP-K-02",  "WAP-9180AX-K9", "Wi-Fi 6E AP", 110, 1200],
            ["OAK-K-03",  "WAP-9180AX-K9", "Wi-Fi 6E AP", 120, 1200],
            ["ELM-K-01",  "C9300-48P-A",   "48-port PoE+ switch", 6, 3500],
            ["MAP-K-02",  "C9300-48P-A",   "48-port PoE+ switch", 7, 3500],
            ["OAK-K-03",  "C9300-48P-A",   "48-port PoE+ switch", 9, 3500],
        ],
        ["Site ID", "Part Number", "Description", "Qty", "Unit Price"],
    )
    return base


def domain_hospital_lan_refresh(out: Path) -> Path:
    base = out / "hospital_lan_refresh"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["MEM-HQ-01",  "Memorial Main Hospital",   "1500 Medical Dr, Phoenix AZ 85013",  "MDF-A1 / IDF A2-8", "24x7", "Hospital Facilities"],
        ["MEM-WB-02",  "Memorial West Branch",     "8800 W Bell Rd, Glendale AZ 85308",  "MDF-W1 / IDF W2-5", "24x7", "Hospital Facilities"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Memorial Health LAN Refresh 2026", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Memorial Health LAN Refresh 2026",
        [
            "Install 280 wireless access points covering all patient floors.",
            "Install 36 Cisco C9500-40X core switches with HA pairs.",
            "Migrate 14 IDFs from legacy chassis to C9300 stack.",
            "Provide 36 months of Gold-tier managed service post-cutover.",
        ],
        total_usd=3_450_000,
        milestones=[("Kickoff", "Apr 1, 2026"), ("HIPAA cutover review", "May 15, 2026"), ("ATP", "Dec 1, 2026")],
    )
    _sla_pdf(
        art / "03_sla.pdf", "Memorial Health Managed Services SLA",
        [
            "Priority 1 (clinical down): Response within 30 minutes, resolution within 2 hours.",
            "Priority 2 (degraded): Response within 1 hour, resolution within 4 hours.",
            "Network uptime: 99.99% measured monthly.",
            "After-hours pager coverage 24x7 with named on-call engineer.",
        ],
    )
    return base


def domain_retail_chain_pos(out: Path) -> Path:
    base = out / "retail_chain_pos"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["STORE-142", "Cherry Creek",       "3030 E 1st Ave, Denver CO 80206",     "MDF-Backroom", "Mall hours", "Store Mgr"],
        ["STORE-143", "Park Meadows",       "8401 Park Meadows Ctr Dr, Lone Tree CO 80124", "MDF-A", "Mall hours", "Store Mgr"],
        ["STORE-144", "Flatiron Crossing",  "1 Flatiron Crossing Dr, Broomfield CO 80021",  "MDF-A", "Mall hours", "Store Mgr"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "RetailCo POS Rollout - 3 Stores", sites)
    _sow_pdf(
        art / "02_sow.pdf", "RetailCo POS Rollout",
        [
            "Deploy 24 POS-T terminals (8 per store) and 12 POS-P printers (4 per store).",
            "Replace 6 Cisco WAP-9180AX-K9 access points (2 per store).",
            "Configure 6 VLANs per store for POS / Guest / IoT / Inventory / Security / Mgmt.",
        ],
        total_usd=485_000,
        milestones=[("Kickoff", "Feb 1, 2026"), ("Pilot store cutover", "Feb 22, 2026"), ("All stores live", "Mar 31, 2026")],
    )
    _xlsx_bom(art / "03_bom.xlsx",
        [
            ["STORE-142", "POS-T-K9",          "POS terminal Wi-Fi", 8, 850],
            ["STORE-143", "POS-T-K9",          "POS terminal Wi-Fi", 8, 850],
            ["STORE-144", "POS-T-K9",          "POS terminal Wi-Fi", 8, 850],
            ["STORE-142", "POS-P-Receipt",     "POS receipt printer", 4, 220],
            ["STORE-143", "POS-P-Receipt",     "POS receipt printer", 4, 220],
            ["STORE-144", "POS-P-Receipt",     "POS receipt printer", 4, 220],
            ["STORE-142", "WAP-9180AX-K9",     "Wi-Fi 6E AP",        2, 1200],
            ["STORE-143", "WAP-9180AX-K9",     "Wi-Fi 6E AP",        2, 1200],
            ["STORE-144", "WAP-9180AX-K9",     "Wi-Fi 6E AP",        2, 1200],
        ],
        ["Site ID", "Part Number", "Description", "Qty", "Unit Price"],
    )
    return base


def domain_hospitality_hotel_wifi(out: Path) -> Path:
    base = out / "hospitality_hotel_wifi"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["MAR-ATL-01", "Marriott Atlanta",     "265 Peachtree Center Ave, Atlanta GA",   "MDF-1A / IDF per floor", "24x7", "Hotel Engineering"],
        ["MAR-MIA-02", "Marriott Miami Beach", "1 Lincoln Rd, Miami Beach FL",            "MDF-1A / IDF per floor", "24x7", "Hotel Engineering"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Marriott Wi-Fi 6E Refresh", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Marriott Wi-Fi 6E Refresh",
        [
            "Install 480 wireless access points across guest rooms, lobbies, and meeting rooms.",
            "Deploy 24 Cisco C9300-48P-A switches for AP backhaul.",
            "Provide 24 months of Silver-tier managed service.",
        ],
        total_usd=2_180_000,
        milestones=[("Atlanta cutover", "Mar 31, 2026"), ("Miami cutover", "May 31, 2026")],
    )
    return base


def domain_fed_gov_datacenter(out: Path) -> Path:
    base = out / "fed_gov_datacenter"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["FED-DC-01", "GSA Reston Datacenter", "1800 F St NW, Washington DC 20405",   "MDF-Cage7 / IDF Cage7-2", "24x7 escorted", "GSA Security"],
        ["FED-DC-02", "GSA Sterling Backup",   "44425 Atlantic Blvd, Sterling VA",    "MDF-A / IDF B",            "24x7 escorted", "GSA Security"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "GSA Datacenter Consolidation FY26", sites)
    _sow_pdf(
        art / "02_sow.pdf", "GSA Datacenter Consolidation FY26",
        [
            "Install 16 Cisco Nexus 9504 chassis (8 per DC).",
            "Migrate 480 VLANs from legacy Catalyst 6500 to Nexus.",
            "Provide TS/SCI cleared engineers for all on-site work.",
            "FedRAMP High compliance required for all managed-service tooling.",
        ],
        total_usd=12_400_000,
        milestones=[("PoP Authority to Operate", "Mar 15, 2026"), ("Primary DC cutover", "Jun 30, 2026"), ("ATP", "Dec 31, 2026")],
    )
    return base


def domain_university_campus_av(out: Path) -> Path:
    base = out / "university_campus_av"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["UVA-ENG-01",  "School of Engineering",     "351 McCormick Rd, Charlottesville VA", "MDF-Eng1A", "Mon-Fri 06:00-22:00", "UVA Facilities"],
        ["UVA-LAW-02",  "School of Law",              "580 Massie Rd, Charlottesville VA",    "MDF-Law1A", "Mon-Fri 06:00-22:00", "UVA Facilities"],
        ["UVA-DAR-03",  "Darden Business School",     "100 Darden Blvd, Charlottesville VA",  "MDF-Dar1A", "Mon-Fri 06:00-23:00", "UVA Facilities"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "UVA AVoIP Refresh Phase 2", sites)
    _sow_pdf(
        art / "02_sow.pdf", "UVA AVoIP Refresh Phase 2",
        [
            "Install Q-SYS Core 110f processors (6 across 3 schools).",
            "Deploy 120 NV-32-H AVoIP encoders + 180 NV-32-H decoders.",
            "Provide Crestron NVX licensing for 24 lecture halls.",
        ],
        total_usd=1_840_000,
        milestones=[("Mockup classroom approval", "Feb 28, 2026"), ("Phase 2 complete", "Aug 15, 2026")],
    )
    return base


def domain_museum_security(out: Path) -> Path:
    base = out / "museum_security"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["MOMA-NYC-01", "MOMA Manhattan", "11 W 53rd St, New York NY 10019", "MDF-B2", "After-hours", "MOMA Security"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "MOMA Security Modernization", sites)
    _sow_pdf(
        art / "02_sow.pdf", "MOMA Security Modernization",
        [
            "Install 240 mini-dome IP cameras with thermal-sensing variants in storage vaults.",
            "Deploy 80 card readers + 40 biometric readers across staff zones.",
            "Integrate with existing Lenel OnGuard access control.",
        ],
        total_usd=2_750_000,
        milestones=[("Curator coordination meeting", "Mar 1, 2026"), ("Vault cameras live", "Jun 1, 2026")],
    )
    return base


def domain_airport_terminal_fiber(out: Path) -> Path:
    base = out / "airport_terminal_fiber"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["LAX-T1-01", "LAX Terminal 1",  "1 World Way, Los Angeles CA 90045", "MDF-T1-A", "Coord w/ TSA", "LAWA Ops"],
        ["LAX-T2-02", "LAX Terminal 2",  "1 World Way, Los Angeles CA 90045", "MDF-T2-A", "Coord w/ TSA", "LAWA Ops"],
        ["LAX-T3-03", "LAX Terminal 3",  "1 World Way, Los Angeles CA 90045", "MDF-T3-A", "Coord w/ TSA", "LAWA Ops"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "LAX Terminal Fiber Backbone Refresh", sites)
    _sow_pdf(
        art / "02_sow.pdf", "LAX Terminal Fiber Backbone Refresh",
        [
            "Pull 18,000 linear feet of OS2 single-mode fiber (288 strand).",
            "Install 36 MPO-24 cassettes per terminal.",
            "Replace 12 SFP-10G-LR-S= modules per terminal.",
        ],
        total_usd=4_650_000,
        milestones=[("T1 fiber cutover", "Mar 30, 2026"), ("All terminals", "Aug 31, 2026")],
    )
    return base


def domain_stadium_paging(out: Path) -> Path:
    base = out / "stadium_paging"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["MBS-ATL-01", "Mercedes-Benz Stadium", "1 AMB Dr NW, Atlanta GA 30313", "MDF-Field-A", "Event-day coord", "MBS Ops"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Mercedes-Benz Stadium PA Upgrade", sites)
    _sow_pdf(
        art / "02_sow.pdf", "MBS Stadium PA Upgrade",
        [
            "Install 320 Biamp Tesira ceiling speakers across concourses.",
            "Deploy 8 Tesira FORTE X 800 DSP processors.",
            "Provide acoustic modeling and STIPA verification per FIFA spec.",
        ],
        total_usd=1_950_000,
        milestones=[("DSP cutover", "Apr 30, 2026"), ("Final acoustic verification", "Jul 15, 2026")],
    )
    return base


def domain_oil_gas_remote(out: Path) -> Path:
    base = out / "oil_gas_remote"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["EXX-PRMN-01", "Exxon Permian Hub A", "30.5 mi N of Midland TX",  "MDF-Remote-A", "24x7 lone-worker", "EXX HSE"],
        ["EXX-PRMN-02", "Exxon Permian Hub B", "42.2 mi N of Midland TX",  "MDF-Remote-A", "24x7 lone-worker", "EXX HSE"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Exxon Permian Connectivity", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Exxon Permian Connectivity",
        [
            "Deploy Cradlepoint AER3100 routers at 14 remote pads.",
            "Install ruggedized Wi-Fi 6E APs in Class I Div 2 enclosures.",
            "Provide cellular failover via 2 carriers per site.",
        ],
        total_usd=890_000,
        milestones=[("Hub A connectivity live", "Mar 15, 2026"), ("All pads online", "May 31, 2026")],
    )
    return base


def domain_utilities_scada(out: Path) -> Path:
    base = out / "utilities_scada"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["DUKE-CHA-01", "Duke Charlotte Substation A", "401 S Tryon St, Charlotte NC", "Substation MDF", "24x7", "Duke Field Ops"],
        ["DUKE-RAL-02", "Duke Raleigh Substation B",   "100 Fayetteville St, Raleigh NC", "Substation MDF", "24x7", "Duke Field Ops"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Duke Energy SCADA Modernization", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Duke Energy SCADA Modernization",
        [
            "Install 24 hardened Cisco IE-4000 industrial switches across 12 substations.",
            "Deploy 96 IoT sensors for transformer temperature + oil monitoring.",
            "NERC CIP-007 compliant baseline + audit trail required.",
        ],
        total_usd=1_650_000,
        milestones=[("CIP audit", "Mar 31, 2026"), ("ATP", "Nov 30, 2026")],
    )
    return base


def domain_banking_branch_refresh(out: Path) -> Path:
    base = out / "banking_branch_refresh"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["BOA-NYC-01", "BoA Manhattan Branch", "1 Bryant Park, New York NY",  "MDF-Floor4", "Mon-Sat 08:00-18:00", "BoA Security"],
        ["BOA-SFO-02", "BoA SF Embarcadero",   "555 California St, San Francisco CA", "MDF-Floor3", "Mon-Sat 08:00-18:00", "BoA Security"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "BoA Branch Refresh 2026", sites)
    _sow_pdf(
        art / "02_sow.pdf", "BoA Branch Refresh 2026",
        [
            "Refresh 64 branch routers (Cisco ISR4451-X).",
            "Install Zscaler ZIA + ZPA tunnels at every branch.",
            "Provide PCI DSS 4.0 segmentation review.",
        ],
        total_usd=3_800_000,
        milestones=[("PCI scan", "Mar 15, 2026"), ("Branch cutover wave 1", "Jun 1, 2026")],
    )
    return base


def domain_manufacturing_floor(out: Path) -> Path:
    base = out / "manufacturing_floor"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["FORD-DEA-01", "Ford Dearborn Plant", "1 American Rd, Dearborn MI 48126", "MDF-Plant-A / IDF-Line per cell", "Plant shift hours", "Ford Plant Eng"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Ford Plant IoT Refresh", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Ford Plant IoT Refresh",
        [
            "Deploy 320 IoT vibration sensors on production lines 1-12.",
            "Install 48 hardened Aruba 6300 switches in plant cells.",
            "Configure OT-IT segmentation per IEC 62443 SL2.",
        ],
        total_usd=2_240_000,
        milestones=[("Line 1 instrumented", "Mar 31, 2026"), ("All lines live", "Aug 31, 2026")],
    )
    return base


def domain_telco_macro_site(out: Path) -> Path:
    base = out / "telco_macro_site"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["VZ-CELL-001", "Verizon Macro 001", "32.5 mi NE Phoenix AZ", "Cabinet-A", "Carrier-coord", "Tower Tech"],
        ["VZ-CELL-002", "Verizon Macro 002", "18.2 mi W Phoenix AZ",  "Cabinet-A", "Carrier-coord", "Tower Tech"],
        ["VZ-CELL-003", "Verizon Macro 003", "44.1 mi S Phoenix AZ",  "Cabinet-A", "Carrier-coord", "Tower Tech"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Verizon Macro Site C-band Add", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Verizon Macro Site C-band Add",
        [
            "Install 12 Ericsson AIR 6488 radios across 3 sectors.",
            "Add C-band 3.7-3.98 GHz support to existing AAU.",
            "Coordinate carrier-aggregation testing with NOC.",
        ],
        total_usd=720_000,
        milestones=[("Sector 1 RFR", "Mar 1, 2026"), ("All sectors integrated", "Apr 30, 2026")],
    )
    return base


def domain_retail_loss_prevention(out: Path) -> Path:
    base = out / "retail_loss_prevention"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["TGT-NYC-01", "Target Brooklyn",  "139 Flatbush Ave, Brooklyn NY",   "MDF-Backroom", "Store hours", "Store AP"],
        ["TGT-NYC-02", "Target Manhattan", "112 W 34th St, New York NY",      "MDF-Backroom", "Store hours", "Store AP"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Target LP Camera Refresh", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Target LP Camera Refresh",
        [
            "Install 84 4K dome cameras with AI loss-prevention analytics.",
            "Deploy 4 NVR-9XL servers (2 per store) with 90-day retention.",
            "Integrate with Genetec Security Center 5.11.",
        ],
        total_usd=1_120_000,
        milestones=[("Pilot store analytics validation", "Feb 28, 2026"), ("Both stores live", "Apr 30, 2026")],
    )
    return base


def domain_warehouse_wms(out: Path) -> Path:
    base = out / "warehouse_wms"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["AMZN-DAL-01", "Amazon DAL5 Fulfillment Center", "33500 Cabot Dr, Dallas TX", "MDF-Floor-A", "24x7", "Amazon Ops"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Amazon DAL5 WMS Network", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Amazon DAL5 WMS Network",
        [
            "Deploy 240 Zebra TC58 ruggedized handheld scanners.",
            "Install 96 Wi-Fi 6E APs across 1.2M sqft floor.",
            "Configure 18 SSIDs for picker / packer / receiver / mgmt / IoT.",
        ],
        total_usd=1_580_000,
        milestones=[("Peak-season cutover", "Sep 30, 2026")],
    )
    return base


def domain_campus_access_control(out: Path) -> Path:
    base = out / "campus_access_control"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["GOOG-MV-01", "Google Mountain View",  "1600 Amphitheatre Pkwy, Mountain View CA", "MDF-Bldg40", "Badged 24x7", "Google Security"],
        ["GOOG-SVL-02", "Google Sunnyvale",     "1295 Charleston Rd, Sunnyvale CA",          "MDF-Bldg41", "Badged 24x7", "Google Security"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Google Campus Access Refresh", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Google Campus Access Refresh",
        [
            "Install 480 HID Signo 40 readers across 18 buildings.",
            "Deploy 24 Mercury LP4502 controllers.",
            "Migrate from legacy Lenel to Genetec Synergis.",
        ],
        total_usd=4_320_000,
        milestones=[("Mountain View phase 1", "Mar 31, 2026"), ("All buildings live", "Dec 31, 2026")],
    )
    return base


def domain_data_center_imac(out: Path) -> Path:
    base = out / "data_center_imac"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["EQX-DC-01", "Equinix DC10 Ashburn", "21701 Filigree Ct, Ashburn VA", "Cage-217 / Cabinet-217-A1", "24x7 escorted", "Equinix Floor"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Equinix Cage 217 IMAC", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Equinix Cage 217 IMAC",
        [
            "Install 12 Cisco UCS C240 M7 servers per Cabinet A1-A4.",
            "Move 36 existing patch cords to new Bel den panel.",
            "Add cross-connects: 8 to Cabinet B7 via MMR.",
            "Decommission 24 EOL HPE DL360 G9 servers; ship to ITAD vendor.",
        ],
        total_usd=412_000,
        milestones=[("IMAC window", "Mar 15-22, 2026")],
    )
    return base


def domain_cruise_ship_av(out: Path) -> Path:
    base = out / "cruise_ship_av"
    art = base / "artifacts"
    art.mkdir(parents=True, exist_ok=True)
    sites = [
        ["RCL-SHIP-01", "Royal Symphony of the Seas", "Dry dock, Cádiz Spain", "Bridge MDF", "Dry-dock window only", "Ship Engineer"],
    ]
    _site_roster_pdf(art / "01_site_roster.pdf", "Symphony AV Refresh Dry Dock", sites)
    _sow_pdf(
        art / "02_sow.pdf", "Symphony AV Refresh Dry Dock",
        [
            "Replace 64 ceiling speakers in main theatre and Studio B.",
            "Deploy Q-SYS Core 510i + 12 NV-32 endpoints.",
            "Provide IP67-rated outdoor speakers for pool deck.",
        ],
        total_usd=2_460_000,
        milestones=[("Dry-dock entry", "Sep 1, 2026"), ("Ship returns to service", "Oct 22, 2026")],
    )
    return base


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python scripts/build_domain_packages.py <out_dir>", file=sys.stderr)
        return 2
    out = Path(sys.argv[1]).resolve()
    out.mkdir(parents=True, exist_ok=True)
    builders = [
        ("k12_school_district",      domain_k12_school_district),
        ("hospital_lan_refresh",     domain_hospital_lan_refresh),
        ("retail_chain_pos",         domain_retail_chain_pos),
        ("hospitality_hotel_wifi",   domain_hospitality_hotel_wifi),
        ("fed_gov_datacenter",       domain_fed_gov_datacenter),
        ("university_campus_av",     domain_university_campus_av),
        ("museum_security",          domain_museum_security),
        ("airport_terminal_fiber",   domain_airport_terminal_fiber),
        ("stadium_paging",           domain_stadium_paging),
        ("oil_gas_remote",           domain_oil_gas_remote),
        ("utilities_scada",          domain_utilities_scada),
        ("banking_branch_refresh",   domain_banking_branch_refresh),
        ("manufacturing_floor",      domain_manufacturing_floor),
        ("telco_macro_site",         domain_telco_macro_site),
        ("retail_loss_prevention",   domain_retail_loss_prevention),
        ("warehouse_wms",            domain_warehouse_wms),
        ("campus_access_control",    domain_campus_access_control),
        ("data_center_imac",         domain_data_center_imac),
        ("cruise_ship_av",           domain_cruise_ship_av),
    ]
    for name, fn in builders:
        b = fn(out)
        print(f"  -> {b.name}/")
    print(f"\n{len(builders)} domain bundles in {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
