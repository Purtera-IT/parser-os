from __future__ import annotations

import csv
import os
import re
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.ids import stable_id
from app.core.item_identity import merge_parser_value_identity
from app.core.normalizers import normalize_entity_key, normalize_text, parse_quantity
from app.core.segments import ArtifactSegment
from app.core.schemas import (
    ArtifactType,
    AtomType,
    AuthorityClass,
    EvidenceAtom,
    ParserOutput,
    ReviewStatus,
    SourceRef,
    ParserCapability,
    ParserMatch,
)
from app.parsers.base import BaseParser
from app.parsers.segmenters import segment_quote
from app.parsers.sheet_classifier import SheetDestination, classify_sheet
from app.parsers.structured_projection import (
    derived_files_for,
    make_page,
    make_paragraph,
    make_section,
    make_structured_document,
    make_table,
    stamp_section_and_block_ids,
)
from app.domain.schemas import DomainPack

STRUCTURED_SCHEMA_QUOTE = "orbitbrief.quote.structured.v1"

# Canonical column keys -> normalized header substring aliases (lowercase, punctuation-stripped variants added at match time).
HEADER_ALIASES: dict[str, set[str]] = {
    "description": {
        "description",
        "item",
        "line item",
        "product",
        "material",
        "service",
        "scope item",
        "work item",
        "equipment",
        "labor item",
        "cable type",
        "line",
    },
    "quantity": {
        "qty",
        "qty.",
        "quantity",
        "count",
        "units",
        "no",
        "#",
        "quoted qty",
        "bid qty",
        "order qty",
        "est qty",
    },
    "uom": {
        "uom",
        "unit",
        "measure",
        "ea",
        "each",
        "lf",
        "ft",
        "feet",
        "lot",
    },
    "unit_price": {
        "unit price",
        "unit cost",
        "price",
        "cost",
        "rate",
        "labor rate",
    },
    "extended_price": {
        "ext price",
        "extended price",
        "line total",
        "total price",
        "amount",
        "extended cost",
    },
    "part_number": {
        "part",
        "part number",
        "sku",
        "item number",
        "mpn",
        "model",
        "manufacturer part number",
        "catalog #",
        "catalog#",
    },
    "manufacturer": {
        "manufacturer",
        "mfr",
        "brand",
        "make",
    },
    "material_spec": {
        "quoted material / spec",
        "quoted material spec",
        "material spec",
        "category",
        "cable type",
        "cat",
        "plenum",
        "shielded",
        "utp",
        "stp",
        "material",
        "spec",
    },
    "lead_time": {
        "lead time",
        "eta",
        "availability",
        "delivery",
        "ship date",
    },
    "included": {
        "included",
        "included?",
        "in scope?",
        "base bid",
        "alt",
        "alternate",
        "option",
        "excluded",
    },
    "notes": {
        "notes",
        "comments",
        "clarifications",
        "assumptions",
        "exclusions",
    },
    "section": {
        "section",
        "category",
        "phase",
        "area",
        "system",
    },
    # Site / facility identifier — when a BOM row carries a Site ID
    # column, include it in the emitted atom raw_text so the site
    # entity extractor can bind the row to a physical_site.
    "site_id": {
        "site id",
        "site",
        "site code",
        "site key",
        "site number",
        "facility id",
        "facility code",
        "facility",
        "location",
        "location id",
        "location code",
        "store #",
        "store id",
        "store number",
        "building",
        "building id",
        "campus id",
    },
    # Currency code (USD / EUR / GBP / JPY) — included in atom text
    # so multi-currency BOMs don't silently lose the currency tag.
    "currency": {
        "currency",
        "ccy",
        "currency code",
        "iso currency",
    },
    # Amount / total — when the row has its own Amount column
    # (different shape from unit_price/extended_price).
    "amount": {
        "amount",
        "total",
        "value",
        "spend",
        "cost",
        "subtotal",
    },
    # Region (US / EMEA / UK / APAC) — multi-region BOMs use this
    # to group line items. Surface in atom text for cross-doc
    # reasoning.
    "region": {
        "region",
        "geo",
        "geography",
        "country",
        "territory",
        "market",
    },
    "vendor": {
        "vendor",
        "supplier",
        "subcontractor",
        "bidder",
    },
}

RowKind = Literal[
    "real_line_item",
    "subtotal",
    "grand_total",
    "tax",
    "shipping",
    "discount",
    "allowance",
    "alternate",
    "option",
    "excluded",
    "included_no_qty",
    "section_header",
    "blank",
    "malformed",
]


# Service-line classifier (mirrors the one in xlsx_parser for quote
# rows). Routes labor / training / hypercare / governance / consulting
# line items to `service:` instead of `device:`.
_SERVICE_DESCRIPTION_TOKENS: tuple[str, ...] = (
    "labor", "labour", "hours", "hour", "after-hours", "after hours",
    "support", "supports",
    "training", "trainings", "adoption",
    "workshop", "workshops", "discovery",
    "design", "engineering services",
    "professional services", "managed services",
    "consulting", "consultancy", "advisory",
    "hypercare", "warranty", "warrantee",
    "project management", "program management", "pmo",
    "governance", "oversight",
    "implementation", "installation services", "deployment services",
    "commissioning", "decommissioning",
    "migration", "cutover", "go-live", "go live",
    "documentation services", "as-built", "as built",
    "testing services", "validation services", "uat", "acceptance",
    "rfp response", "proposal preparation",
)


def _looks_like_service_description(value: str) -> bool:
    """Return True if a quote line-item description is a service
    rather than a physical device."""
    lower = value.lower()
    return any(token in lower for token in _SERVICE_DESCRIPTION_TOKENS)


def _header_cell_keys(cell: Any) -> set[str]:
    raw = str(cell or "").strip()
    if not raw:
        return set()
    lowered = raw.lower()
    keys: set[str] = {
        normalize_text(raw).strip(".:"),
        normalize_text(raw.replace("/", " ")).strip(".:"),
        normalize_text(raw.replace("/", " ").replace("?", "")).strip(".:"),
        re.sub(r"\s+", " ", lowered).strip(),
    }
    return {k for k in keys if k}


def _merge_header_cells(top: Any, bottom: Any) -> str:
    a = str(top or "").strip()
    b = str(bottom or "").strip()
    if a and b:
        return f"{a} {b}".strip()
    return a or b


def _header_map_from_row(row: list[Any]) -> dict[str, int]:
    """Map canonical header keys → column index. Each column index is
    claimed by at most ONE canonical key (RF8 fix). Without this, a
    column whose header is "Category" would get claimed by both
    ``material_spec`` and ``section`` because both alias sets contain
    "category" — producing duplicate locator entries like
    ``{section: C, material_spec: C}`` in every emitted atom.
    """
    current_map: dict[str, int] = {}
    claimed_cols: set[int] = set()
    for col_idx, cell in enumerate(row):
        cell_keys = _header_cell_keys(cell)
        if not cell_keys:
            continue
        if col_idx in claimed_cols:
            continue
        for canonical, aliases in HEADER_ALIASES.items():
            if canonical in current_map:
                continue
            if cell_keys & aliases:
                current_map[canonical] = col_idx
                claimed_cols.add(col_idx)
                break
    return current_map


def _header_map_from_two_rows(row_top: list[Any], row_bot: list[Any]) -> dict[str, int]:
    width = max(len(row_top), len(row_bot))
    merged: list[str] = []
    for col in range(width):
        top = row_top[col] if col < len(row_top) else None
        bot = row_bot[col] if col < len(row_bot) else None
        merged.append(_merge_header_cells(top, bot))
    return _header_map_from_row(merged)


def _qualifies_quote_header(header_map: dict[str, int], *, strict_minimum_signals: bool = False) -> bool:
    keys = set(header_map)
    if not keys:
        return False
    if keys <= {"notes", "section"}:
        return False
    has_desc = "description" in keys
    has_qty = "quantity" in keys
    has_part = "part_number" in keys
    has_inc = "included" in keys
    has_price = ("unit_price" in keys) or ("extended_price" in keys)
    has_mat = "material_spec" in keys
    has_vendor = "vendor" in keys
    commercial_line = has_price or has_part or has_inc or has_mat or has_vendor
    # Routing/sniff: do not treat description+qty alone as quote-like (site rosters).
    q1 = has_desc and has_qty and (commercial_line if strict_minimum_signals else True)
    q2 = has_desc and has_inc
    q3 = has_desc and has_price
    q4 = has_part and has_qty
    q5 = has_desc and has_mat and (has_price or has_inc or has_qty or has_part)
    return bool(q1 or q2 or q3 or q4 or q5)


def _detect_header_advanced(
    rows: list[list[Any]], scan_limit: int = 40
) -> tuple[int | None, dict[str, int], str, list[str]]:
    """Return (header_row_index_0based, header_map, mode single|pair, diagnostics)."""
    diagnostics: list[str] = []
    best_idx: int | None = None
    best_map: dict[str, int] = {}
    best_score = -1
    best_mode = "single"

    limit = min(scan_limit, len(rows))
    for idx in range(limit):
        row = rows[idx]
        single_map = _header_map_from_row(row)
        if _qualifies_quote_header(single_map):
            score = len(single_map)
            if score > best_score:
                best_score = score
                best_idx = idx
                best_map = single_map
                best_mode = "single"
        if idx + 1 < len(rows):
            pair_map = _header_map_from_two_rows(row, rows[idx + 1])
            if _qualifies_quote_header(pair_map):
                score = len(pair_map) + 0.25  # tie-break: prefer single if equal
                if score > best_score:
                    best_score = score
                    best_idx = idx
                    best_map = pair_map
                    best_mode = "pair"
    if best_idx is not None:
        diagnostics.append(f"header_row={best_idx + 1} mode={best_mode} keys={sorted(best_map.keys())}")
    return best_idx, best_map, best_mode, diagnostics


def _sheet_is_false_positive(
    rows: list[list[Any]], header_idx: int | None, header_map: dict[str, int]
) -> tuple[bool, str]:
    """Reject cover / instructions / terms-only sheets."""
    if header_idx is None or not header_map:
        return True, "no_header"
    sample = " ".join(
        normalize_text(str(c or "")) for row in rows[: min(25, len(rows))] for c in row
    )
    if "terms and conditions" in sample and "quantity" not in header_map and len(header_map) <= 3:
        return True, "terms_like"
    if "instructions to bidders" in sample and not _qualifies_quote_header(header_map):
        return True, "instructions"
    # Sparse grid with no qty column and only prose in first column
    if "quantity" not in header_map and "part_number" not in header_map:
        nonempty = sum(1 for r in rows[header_idx + 1 : header_idx + 12] if any(str(c or "").strip() for c in r))
        if nonempty <= 1 and len(header_map) <= 2:
            return True, "sparse_no_table"
    return False, ""


def _classify_row(
    desc: str,
    part: str,
    quantity_raw: str,
    unit_price: str,
    extended: str,
    notes: str,
) -> RowKind:
    d = normalize_text(desc).strip()
    p = normalize_text(part).strip()
    label = d or p
    if not label and not str(quantity_raw or "").strip() and not str(unit_price or "").strip():
        return "blank"
    if label:
        if re.match(r"^(total|subtotal|grand total)\b", label):
            return "grand_total" if label.startswith("grand") else "subtotal"
        if re.match(r"^(tax|sales tax|vat)\b", label):
            return "tax"
        if re.match(r"^(shipping|freight|delivery fee)\b", label):
            return "shipping"
        if re.match(r"^(discount|credit)\b", label):
            return "discount"
        if "allowance" in label:
            return "allowance"
        if re.match(r"^(alternate|alt\d|option)\b", label):
            return "alternate"
        if label.endswith(":") and len(label) < 48 and not str(quantity_raw or "").strip():
            return "section_header"
    n = normalize_text(notes).lower()
    if "not included" in n or "excluded from" in n or "out of scope" in n:
        if not str(quantity_raw or "").strip():
            return "excluded"
    if str(quantity_raw or "").strip().lower() in {"included", "inc", "n/c", "nc"} and not re.search(r"\d", str(quantity_raw)):
        return "included_no_qty"
    if label and not re.search(r"\d", str(quantity_raw or "")) and not str(unit_price or "").strip():
        if len(label) < 4:
            return "malformed"
    return "real_line_item"


def _extract_quantity_from_description(description: str) -> str | None:
    t = str(description or "")
    patterns = [
        r"\(\s*(\d[\d,]*)\s*\)",
        r"qty\.?\s*[:=]?\s*(\d[\d,]*)",
        r"quantity\s*[:=]?\s*(\d[\d,]*)",
        r"(\d[\d,]*)\s*(ea|each|lot)\b",
    ]
    for pat in patterns:
        m = re.search(pat, t, flags=re.I)
        if m:
            return m.group(1).replace(",", "")
    return None


def parse_quote_quantity(
    description: str,
    quantity_cell: str,
    uom_cell: str,
    notes: str,
) -> dict[str, Any]:
    raw = str(quantity_cell or "").strip()
    if not raw:
        emb = _extract_quantity_from_description(description)
        if emb:
            raw = emb
    combined = normalize_text(f"{raw} {uom_cell or ''}").strip()
    low = combined.lower()
    result: dict[str, Any] = {
        "quantity_raw": raw,
        "uom": str(uom_cell or "").strip() or None,
        "quantity": None,
        "unit": "count",
        "uncertain": True,
        "quantity_status": "missing",
        "quantity_min": None,
        "quantity_max": None,
    }
    if not raw and not _extract_quantity_from_description(description):
        if "included" in low or low in {"inc", "y", "yes"}:
            result["quantity_status"] = "included_no_qty"
            result["uncertain"] = False
            return result
        result["quantity_status"] = "missing"
        return result

    if low in {"n/a", "na", "tbd", "t.b.d.", "pending"}:
        result["quantity_status"] = "tbd" if "tbd" in low or "pending" in low else "not_applicable"
        result["uncertain"] = True
        return result
    if "allowance" in low or low == "lot" or "lot" in low:
        result["quantity_status"] = "allowance" if "allowance" in low else "known"
        m = re.match(r"^\s*(\d[\d,]*)\s+lot", low)
        if m:
            q = int(m.group(1).replace(",", ""))
            result["quantity"] = q
            result["uncertain"] = False
            result["quantity_status"] = "known"
            return result
        raw_digits = str(quantity_cell or "").strip().replace(",", "")
        m0 = re.match(r"^\s*(-?\d[\d,]*(?:\.\d+)?)", raw_digits)
        if m0:
            q = float(m0.group(1))
            if q.is_integer():
                q = int(q)
            result["quantity"] = q
            result["uncertain"] = False
        return result
    if "included" in low and not re.search(r"\d", raw):
        result["quantity_status"] = "included_no_qty"
        result["uncertain"] = False
        return result

    mrange = re.match(r"^\s*(\d[\d,]*)\s*[-–]\s*(\d[\d,]*)\s*$", raw.replace(",", ""))
    if mrange:
        lo = int(mrange.group(1))
        hi = int(mrange.group(2))
        result["quantity_min"] = lo
        result["quantity_max"] = hi
        result["quantity_status"] = "range"
        result["uncertain"] = True
        return result

    mapprox = re.match(r"^\s*(?:~|approx\.?|approximately)\s*(\d[\d,]*)", low)
    if mapprox:
        q = int(mapprox.group(1).replace(",", ""))
        result["quantity"] = q
        result["quantity_status"] = "known"
        result["uncertain"] = True
        return result

    mneg = re.match(r"^\s*\(\s*(\d[\d,]*(?:\.\d+)?)\s*\)\s*$", raw)
    if mneg:
        q = -float(mneg.group(1).replace(",", ""))
        if q.is_integer():
            q = int(q)
        result["quantity"] = q
        result["quantity_status"] = "known"
        result["uncertain"] = False
        return result

    mq = re.match(r"^\s*(-?\d[\d,]*(?:\.\d+)?)\s*([a-z%]{1,8})?\s*$", low.replace(",", ""))
    if mq:
        q = float(mq.group(1))
        if q.is_integer():
            q = int(q)
        result["quantity"] = q
        result["unit"] = mq.group(2) or "count"
        st = "zero" if result["quantity"] == 0 else "known"
        result["quantity_status"] = st
        result["uncertain"] = False
        return result

    legacy = parse_quantity(raw)
    result.update(legacy)
    if legacy.get("quantity") is not None:
        result["quantity_status"] = "zero" if legacy.get("quantity") == 0 else "known"
    else:
        result["quantity_status"] = "malformed"
    result["uncertain"] = bool(legacy.get("uncertain", True))
    return result


def parse_money_cell(text: str, *, side: Literal["unit", "extended"] = "unit") -> dict[str, Any]:
    """Parse one money cell; populate only the fields for `side` (unit vs extended column)."""
    raw = str(text or "").strip()
    out: dict[str, Any] = {
        "unit_price_raw": None,
        "unit_price_amount": None,
        "extended_price_raw": None,
        "extended_price_amount": None,
        "currency": None,
        "price_status": "missing",
    }
    if side == "unit":
        out["unit_price_raw"] = raw
    else:
        out["extended_price_raw"] = raw
    if not raw:
        out["price_status"] = "missing"
        return out
    low = raw.lower()
    if low in {"included", "n/c", "nc", "no charge", "no-charge", "-"}:
        out["price_status"] = "included" if low == "included" else "no_charge"
        return out
    cleaned = re.sub(r"[$€£]", "", raw)
    cleaned = cleaned.replace(",", "").strip()
    neg = False
    if cleaned.startswith("(") and cleaned.endswith(")"):
        neg = True
        cleaned = cleaned[1:-1].strip()
    try:
        val = float(cleaned)
        if neg:
            val = -val
        if side == "unit":
            out["unit_price_amount"] = val
        else:
            out["extended_price_amount"] = val
        out["price_status"] = "known"
        if "$" in raw or "€" in raw or "£" in raw:
            out["currency"] = "USD" if "$" in raw else None
    except ValueError:
        out["price_status"] = "malformed"
    return out


def normalize_inclusion(included_cell: str, notes: str) -> dict[str, Any]:
    cell = normalize_text(included_cell).lower()
    note = normalize_text(notes).lower()
    text = f"{cell} {note}".strip()
    out: dict[str, Any] = {
        "included": None,
        "inclusion_status": "unknown",
        "owner_furnished_signal": False,
    }
    # Note-first exclusions (contractor not supplying / out of base bid).
    exclusion_note = re.search(
        r"\b("
        r"not included|excluded from|excluded|out of scope|by others|not in contract|"
        r"n\.?\s*i\.?\s*c\.?|\bnic\b|"
        r"by owner|owner furnished|(?<![a-z])ofe(?![a-z])|customer provided"
        r")\b",
        note,
        re.I,
    )
    if exclusion_note:
        out["included"] = False
        out["inclusion_status"] = "excluded"
        if re.search(
            r"\b(owner furnished|(?<![a-z])ofe(?![a-z])|customer provided|by owner)\b",
            note,
            re.I,
        ):
            out["owner_furnished_signal"] = True
        return out
    if cell in {"no", "n", "false", "0"}:
        out["included"] = False
        out["inclusion_status"] = "excluded"
    elif re.search(r"\b(yes|y|true|1|included|base bid)\b", cell):
        out["included"] = True
        out["inclusion_status"] = "included"
    if out["inclusion_status"] != "excluded":
        if re.search(
            r"\b(optional|alternate|alt\s*1|option|separate price|alternate price)\b",
            text,
            re.I,
        ):
            out["inclusion_status"] = "optional"
        elif re.search(r"\ballowance only\b|\ballowance\b", text, re.I):
            out["inclusion_status"] = "allowance"
        elif re.search(r"\btbd\b|to be confirmed|pending field verification|for reference only|budgetary only\b", text, re.I):
            out["inclusion_status"] = "tbd"
    if out["inclusion_status"] == "unknown" and re.search(
        r"\b(for reference only|budgetary only|pending field verification)\b",
        note,
        re.I,
    ):
        out["inclusion_status"] = "tbd"
    return out


def _cable_category_from_blob(blob: str) -> str | None:
    """Distinguish Cat6A from Cat6 (never classify 6A as Cat6)."""
    if re.search(
        r"\b(cat6a|cat[\s-]*6[\s-]*a|category[\s-]*6[\s-]*a|category\s+6a)\b",
        blob,
        re.I,
    ):
        return "cat6a"
    if re.search(
        r"\b(cat6(?!a)|cat[\s-]*6(?!\s*a)|category[\s-]*6(?!\s*a))\b",
        blob,
        re.I,
    ):
        return "cat6"
    return None


def _material_heuristics(description: str, material_spec: str, notes: str) -> dict[str, Any]:
    blob = normalize_text(f"{description} {material_spec} {notes}").lower()
    out: dict[str, Any] = {
        "normalized_item": normalize_text(description).strip(),
        "material_family": None,
        "cable_category": _cable_category_from_blob(blob),
        "shielding": None,
        "jacket_rating": None,
        "item_kind": "other",
        "port_count": None,
        "is_scope_pollution_candidate": False,
    }
    if "stp" in blob or "shielded" in blob:
        out["shielding"] = "shielded"
    elif "utp" in blob or "unshielded" in blob:
        out["shielding"] = "unshielded"
    if "plenum" in blob:
        out["jacket_rating"] = "plenum"
    if "patch panel" in blob:
        out["item_kind"] = "patch_panel"
        m_ports = re.search(r"\b(\d{1,3})\s*[-\s]*port\b", blob, re.I)
        if m_ports:
            out["port_count"] = int(m_ports.group(1))
    elif "keystone" in blob and "jack" in blob:
        out["item_kind"] = "keystone_jack"
    elif re.search(r"\bjacks?\b", blob):
        out["item_kind"] = "jack"
    elif "faceplate" in blob or "wall plate" in blob:
        out["item_kind"] = "faceplate"
    elif "rj45" in blob or "termination" in blob:
        out["item_kind"] = "termination"
    elif "patch cord" in blob:
        out["item_kind"] = "patch_cord"
    elif "raceway" in blob or "conduit" in blob:
        out["item_kind"] = "raceway" if "raceway" in blob else "conduit"
    elif "certif" in blob or "test export" in blob or "fluke" in blob:
        out["item_kind"] = "certification"
    elif "labor" in blob or "programming" in blob:
        out["item_kind"] = "labor"
    elif "drop" in blob or "cable" in blob:
        out["item_kind"] = "cable_drop"
    if re.search(r"\b(power|amp|circuit|voltage|120v|208v)\b", blob) and "patch" not in blob:
        out["is_scope_pollution_candidate"] = True
        if out["item_kind"] == "other":
            out["item_kind"] = "power"
    return out


def _price_math_mismatch(
    qty_obj: dict[str, Any],
    up_money: dict[str, Any],
    ext_money: dict[str, Any],
) -> bool:
    q = qty_obj.get("quantity")
    up = up_money.get("unit_price_amount")
    ext = ext_money.get("extended_price_amount")
    if q is None or up is None or ext is None:
        return False
    if up_money.get("price_status") != "known" or ext_money.get("price_status") != "known":
        return False
    if qty_obj.get("quantity_status") not in {"known", "zero"}:
        return False
    try:
        prod = float(q) * float(up)
        extf = float(ext)
    except (TypeError, ValueError):
        return False
    tol = max(0.02, 1e-4 * max(abs(prod), abs(extf), 1.0))
    return abs(prod - extf) > tol


def _quote_row_needs_review(
    row_kind: RowKind,
    qty_obj: dict[str, Any],
    inclusion_status: str,
    unit_price_raw: str,
    extended_raw: str,
    up_money: dict[str, Any],
    ext_money: dict[str, Any],
    flags: list[str],
) -> bool:
    if row_kind in {"alternate", "option", "allowance", "excluded", "malformed"}:
        return True
    qs = str(qty_obj.get("quantity_status") or "")
    if qs in {"tbd", "range", "allowance", "missing", "malformed"}:
        return True
    if inclusion_status in {"optional", "alternate", "allowance", "tbd", "excluded"}:
        return True
    if up_money.get("price_status") == "malformed" and unit_price_raw.strip():
        return True
    if ext_money.get("price_status") == "malformed" and extended_raw.strip():
        return True
    if "quote_parser:price_math_mismatch" in flags:
        return True
    if qty_obj.get("uncertain"):
        return True
    return False


def _comparison_key(mat: dict[str, Any], normalized_item: str) -> str:
    ik = str(mat.get("item_kind") or "other")
    cat = mat.get("cable_category")
    sh = mat.get("shielding")
    ports = mat.get("port_count")

    def _shield_key() -> str:
        if sh == "unshielded":
            return "utp"
        if sh == "shielded":
            return "stp"
        return "unknown_shield"

    if ik == "cable_drop":
        tier = cat or "unknown_tier"
        return f"cabling:{tier}:{_shield_key()}:drop"
    if ik == "termination":
        return "cabling:rj45:termination"
    if ik == "patch_panel" and ports:
        return f"cabling:patch_panel:{int(ports)}_port"
    if ik == "patch_panel":
        return "cabling:patch_panel:unknown_port"
    if ik == "faceplate":
        return "cabling:faceplate"
    if ik == "jack":
        return "cabling:jack"
    if ik == "keystone_jack":
        return "cabling:keystone_jack"
    if ik == "patch_cord":
        return "cabling:patch_cord"
    if ik in {"raceway", "conduit"}:
        return "pathway:raceway_conduit"
    if ik == "certification":
        return "testing:certification_export"
    if ik == "power":
        return "electrical:power_location"
    if ik == "labor":
        return "labor:generic"
    slug = re.sub(r"[^a-z0-9]+", "_", normalize_text(normalized_item).lower()).strip("_") or "item"
    return f"unknown:{slug[:96]}"


def _commercial_role(
    mat: dict[str, Any],
    inclusion_status: str,
    row_kind: RowKind,
    inc_obj: dict[str, Any],
) -> str:
    ik = str(mat.get("item_kind") or "other")
    if inclusion_status == "allowance" or row_kind == "allowance":
        return "allowance"
    if inclusion_status in {"optional", "alternate"} or row_kind in {"alternate", "option"}:
        return "alternate"
    if inc_obj.get("owner_furnished_signal"):
        return "owner_furnished"
    if ik == "certification":
        return "testing"
    if ik in {"raceway", "conduit"}:
        return "pathway"
    if ik == "power":
        return "electrical"
    if ik == "labor":
        return "labor"
    if ik == "other":
        return "unknown"
    return "material"


def _scope_relevance(
    mat: dict[str, Any],
    included_bool: bool | None,
    inclusion_status: str,
    row_kind: RowKind,
) -> str:
    if mat.get("is_scope_pollution_candidate") and mat.get("item_kind") == "power":
        return "scope_pollution_candidate"
    if inclusion_status == "excluded" or included_bool is False:
        return "excluded_candidate"
    if inclusion_status in {"optional", "alternate"}:
        return "optional_candidate"
    if inclusion_status == "allowance":
        return "allowance_candidate"
    if included_bool is True and inclusion_status == "included":
        return "in_scope_candidate"
    if included_bool is True:
        return "in_scope_candidate"
    return "unknown"


def _confidence_dimensions(
    header_map: dict[str, int],
    qty_obj: dict[str, Any],
    mat: dict[str, Any],
    inc_obj: dict[str, Any],
    column_count: int,
) -> dict[str, str]:
    hm = len(header_map)
    header_mapping = "high" if hm >= 5 else ("medium" if hm >= 3 else "low")
    qs = str(qty_obj.get("quantity_status") or "")
    if qs in {"known", "zero"} and not qty_obj.get("uncertain"):
        quantity_parse = "high"
    elif qs in {"known", "zero", "included_no_qty", "not_applicable"}:
        quantity_parse = "medium"
    else:
        quantity_parse = "low"
    item_normalization = "high" if mat.get("item_kind") not in (None, "other") else "low"
    inc = str(inc_obj.get("inclusion_status") or "unknown")
    inclusion_parse = "high" if inc != "unknown" else "low"
    source_ref = "high" if column_count >= 2 else ("medium" if column_count == 1 else "low")
    return {
        "header_mapping": header_mapping,
        "quantity_parse": quantity_parse,
        "item_normalization": item_normalization,
        "inclusion_parse": inclusion_parse,
        "source_ref": source_ref,
    }


def _parser_explanation(
    header_map: dict[str, int],
    row_kind: RowKind,
    qty_obj: dict[str, Any],
    inclusion_status: str,
    item_kind: str | None,
    comparison_key: str,
) -> list[str]:
    keys = ",".join(sorted(header_map.keys()))
    return [
        f"header_keys:{keys}",
        f"row_kind:{row_kind}",
        f"quantity_status:{qty_obj.get('quantity_status')}",
        f"inclusion_status:{inclusion_status}",
        f"item_kind:{item_kind or 'other'}",
        f"comparison_key:{comparison_key}",
    ]


def _source_row_key(filename: str, sheet_name: str, row_number: int) -> str:
    return f"{filename}:{sheet_name}:row_{row_number}"


def _vendor_line_universal_fields(
    header_map: dict[str, int],
    row_kind: RowKind,
    qty_obj: dict[str, Any],
    inc_obj: dict[str, Any],
    mat: dict[str, Any],
    inclusion_status: str,
    included_bool: bool | None,
    column_count: int,
    filename: str,
    sheet_name: str,
    row_number: int,
) -> dict[str, Any]:
    ck = _comparison_key(mat, mat.get("normalized_item") or "")
    cr = _commercial_role(mat, inclusion_status, row_kind, inc_obj)
    sr = _scope_relevance(mat, included_bool, inclusion_status, row_kind)
    cd = _confidence_dimensions(header_map, qty_obj, mat, inc_obj, column_count)
    expl = _parser_explanation(
        header_map,
        row_kind,
        qty_obj,
        inclusion_status,
        mat.get("item_kind"),
        ck,
    )
    return {
        "comparison_key": ck,
        "commercial_role": cr,
        "scope_relevance": sr,
        "authority_boundary": "vendor_quote_can_conflict_but_not_define_scope",
        "confidence_dimensions": cd,
        "parser_explanation": expl,
        "source_row_key": _source_row_key(filename, sheet_name, row_number),
    }


def _scan_quote_metadata(rows: list[list[Any]], max_rows: int = 18) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    blob_lines: list[str] = []
    for row in rows[:max_rows]:
        line = " ".join(str(c or "").strip() for c in row if str(c or "").strip())
        if line:
            blob_lines.append(normalize_text(line))
    blob = " ".join(blob_lines).lower()
    m = re.search(r"\bquote\s*#\s*([A-Z0-9-]+)\b", blob, re.I)
    if m:
        meta["quote_number"] = m.group(1)
    m = re.search(r"\b(?:vendor|from)\s*:\s*([^|]+?)(?:\||$)", blob, re.I)
    if m:
        meta["vendor_name"] = m.group(1).strip()[:120]
    m = re.search(r"\b(?:project|customer)\s*:\s*([^|]+?)(?:\||$)", blob, re.I)
    if m:
        meta["project_name"] = m.group(1).strip()[:120]
    m = re.search(r"\bexpiration\s*[:#]?\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})\b", blob, re.I)
    if m:
        meta["expiration_date"] = m.group(1)
    m = re.search(r"\bvalid (?:through|until)\s*[:#]?\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})\b", blob, re.I)
    if m:
        meta["quote_valid_through"] = m.group(1)
    m = re.search(r"\brevision\s*[:#]?\s*([0-9]+|[A-Z]-?[0-9]+)\b", blob, re.I)
    if m:
        meta["revision"] = m.group(1).strip()
    return meta


class QuoteParser(BaseParser):
    parser_name = "quote"
    parser_version = "quote_parser_v1_4_1"
    capability = ParserCapability(
        parser_name=parser_name,
        parser_version=parser_version,
        supported_extensions=[".xlsx", ".csv", ".txt"],
        supported_artifact_types=[ArtifactType.vendor_quote, ArtifactType.xlsx, ArtifactType.csv, ArtifactType.txt],
        emitted_atom_types=[AtomType.vendor_line_item, AtomType.quantity, AtomType.constraint],
        supported_domain_packs=["*"],
        requires_binary=False,
        supports_source_replay=True,
    )

    def match(self, path: Path, sample_text: str | None, domain_pack: DomainPack | None) -> ParserMatch:
        del domain_pack
        suffix = path.suffix.lower()
        confidence = 0.0
        reasons: list[str] = []
        if suffix not in {".xlsx", ".csv", ".txt"}:
            return ParserMatch(
                parser_name=self.parser_name,
                confidence=0.0,
                reasons=[],
                artifact_type=ArtifactType.vendor_quote,
            )

        # RF1 — explicitly cede known structured-row CSVs to the
        # XlsxParser typed-row profiler. asset_inventory.csv,
        # site_list.csv, risk_register.csv, license_support_matrix.csv
        # have their own ``asset_record`` / ``site_roster`` / ``risk``
        # / ``support_entitlement`` AtomTypes (PR2). Letting the quote
        # parser claim them collapses the rich structured fields into
        # vendor_line_item + quantity atoms only.
        _STRUCTURED_FILENAMES = {
            "asset_inventory", "site_list", "risk_register",
            "license_support_matrix", "support_matrix", "lifecycle",
        }
        stem = path.stem.lower().replace("-", "_")
        if any(s in stem for s in _STRUCTURED_FILENAMES):
            return ParserMatch(
                parser_name=self.parser_name,
                confidence=0.0,
                reasons=["ceded_to_xlsx_typed_row_profiler"],
                artifact_type=ArtifactType.vendor_quote,
            )

        # PR1 (post-v3 review) — multi-sheet operations workbooks
        # contain a BOM/quote sheet AS WELL AS asset / site / port /
        # circuit / risk / cutover / alert sheets. Don't let
        # QuoteParser swallow the whole artifact in that case.
        if suffix == ".xlsx":
            from app.parsers.spreadsheet_route_signals import (
                sniff_operations_workbook_strength,
            )

            ops_score, ops_reasons = sniff_operations_workbook_strength(path)
            if ops_score >= 0.55:
                return ParserMatch(
                    parser_name=self.parser_name,
                    confidence=0.0,
                    reasons=["ceded_to_xlsx_operations_workbook", *ops_reasons],
                    artifact_type=ArtifactType.vendor_quote,
                )

        from app.parsers.spreadsheet_route_signals import path_quote_filename_hint

        if path_quote_filename_hint(path):
            confidence = 0.95
            reasons.append("filename_quote_hint")
        elif self.looks_like_quote_artifact(path):
            confidence = 0.86
            reasons.append("header_quote_hint")
        elif sample_text and "part number" in normalize_text(sample_text):
            confidence = 0.8
            reasons.append("text_part_number_hint")
        return ParserMatch(
            parser_name=self.parser_name,
            confidence=confidence,
            reasons=reasons,
            artifact_type=ArtifactType.vendor_quote,
        )

    def parse(self, artifact_path: Path) -> list[EvidenceAtom]:
        artifact_id = stable_id("art", str(artifact_path))
        return self.parse_artifact("unknown_project", artifact_id, artifact_path)

    def segment_artifact(self, project_id: str, artifact_id: str, path: Path) -> list[ArtifactSegment]:
        return segment_quote(project_id=project_id, artifact_id=artifact_id, path=path, parser_version=self.parser_version)

    def parse_artifact(
        self,
        project_id: str,
        artifact_id: str,
        path: Path,
        domain_pack: DomainPack | None = None,
    ) -> list[EvidenceAtom]:
        return self.parse_artifact_full(
            project_id=project_id,
            artifact_id=artifact_id,
            path=path,
            domain_pack=domain_pack,
        ).atoms

    def parse_artifact_full(
        self,
        project_id: str,
        artifact_id: str,
        path: Path,
        domain_pack: DomainPack | None = None,
    ) -> ParserOutput:
        del domain_pack
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            atoms = self._parse_xlsx(project_id=project_id, artifact_id=artifact_id, path=path)
            sheets = self._collect_sheet_rows_xlsx(path)
            artifact_type = ArtifactType.xlsx
        elif suffix == ".csv":
            atoms = self._parse_csv(project_id=project_id, artifact_id=artifact_id, path=path)
            sheets = self._collect_sheet_rows_csv(path)
            artifact_type = ArtifactType.csv
        elif suffix == ".txt":
            atoms = self._parse_txt(project_id=project_id, artifact_id=artifact_id, path=path)
            sheets = self._collect_sheet_rows_txt(path)
            artifact_type = ArtifactType.txt
        else:
            return ParserOutput(atoms=[])

        structured_doc = self._build_structured_doc(
            filename=path.name,
            artifact_type=artifact_type,
            sheets=sheets,
        )
        stamp_section_and_block_ids(structured_doc, artifact_seed=artifact_id)
        return ParserOutput(
            atoms=atoms,
            derived_files=derived_files_for(artifact_path=path, structured_doc=structured_doc),
        )

    def _collect_sheet_rows_xlsx(self, path: Path) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return []
        return [
            {"name": sheet.title, "rows": [list(row) for row in sheet.iter_rows(values_only=True)]}
            for sheet in workbook.worksheets
        ]

    def _collect_sheet_rows_csv(self, path: Path) -> list[dict[str, Any]]:
        try:
            text_head = path.read_text(encoding="utf-8", errors="ignore")[:8192]
            first_line = text_head.splitlines()[0] if text_head else ""
            delimiter = ","
            if first_line.count("|") > first_line.count(",") and "|" in first_line:
                delimiter = "|"
            elif "\t" in first_line and first_line.count("\t") > first_line.count(","):
                delimiter = "\t"
            with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                reader = csv.reader(handle, delimiter=delimiter)
                rows = [list(row) for row in reader]
            return [{"name": "csv", "rows": rows}]
        except Exception:
            return []

    def _collect_sheet_rows_txt(self, path: Path) -> list[dict[str, Any]]:
        try:
            content = path.read_text(encoding="utf-8", errors="ignore")
            lines = [line for line in content.splitlines() if line.strip()]
            rows = [re.split(r"[,\t|]", line) for line in lines]
            return [{"name": "text", "rows": rows}]
        except Exception:
            return []

    def _build_structured_doc(
        self,
        *,
        filename: str,
        artifact_type: ArtifactType,
        sheets: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Render a vendor quote as one page per sheet/section: each
        page hosts the priced line-item table so OrbitBrief can show
        the LLM exactly which lines were quoted, in source order.
        """
        pages: list[dict[str, Any]] = []
        for index, sheet in enumerate(sheets):
            name = sheet.get("name") or f"sheet_{index}"
            rows = sheet.get("rows") or []
            section_blocks: list[dict[str, Any]] = []
            idx, hmap, _, _ = _detect_header_advanced(rows, scan_limit=40) if rows else (None, {}, None, None)
            if idx is not None and rows:
                header_row = rows[idx] or []
                columns = [
                    (str(c).strip() if c is not None else "") or f"col_{i + 1}"
                    for i, c in enumerate(header_row)
                ]
                table_rows: list[dict[str, Any]] = []
                for row_index in range(idx + 1, len(rows)):
                    row = rows[row_index] or []
                    if all(str(c or "").strip() == "" for c in row):
                        continue
                    cells: dict[str, Any] = {}
                    for col_index, col_name in enumerate(columns):
                        value = row[col_index] if col_index < len(row) else ""
                        cells[col_name] = "" if value is None else str(value).strip()
                    table_rows.append(cells)
                if table_rows:
                    section_blocks.append(make_table(columns=columns, rows=table_rows))
            if not section_blocks:
                snippet: list[str] = []
                for row in rows[:25]:
                    line = " | ".join(str(c).strip() for c in (row or []) if c is not None and str(c).strip())
                    if line:
                        snippet.append(line)
                if snippet:
                    section_blocks.append(make_paragraph("\n".join(snippet)))
            pages.append(
                make_page(
                    page=index,
                    title=name,
                    sections=[
                        make_section(heading=name, level=2, blocks=section_blocks)
                    ],
                )
            )
        if not pages:
            pages.append(
                make_page(
                    page=0,
                    title=filename,
                    sections=[
                        make_section(
                            heading=filename,
                            level=2,
                            blocks=[make_paragraph("(empty quote)")],
                        )
                    ],
                )
            )
        return make_structured_document(
            schema_version=STRUCTURED_SCHEMA_QUOTE,
            filename=filename,
            artifact_type=artifact_type.value,
            title=filename,
            metadata=[f"sheet: {s.get('name')}" for s in sheets if s.get("name")],
            pages=pages,
        )

    @classmethod
    def looks_like_quote_artifact(cls, path: Path) -> bool:
        from app.parsers.spreadsheet_route_signals import (
            likely_site_roster_header_row,
            path_quote_filename_hint,
        )

        if path_quote_filename_hint(path):
            return True

        suffix = path.suffix.lower()
        try:
            if suffix == ".xlsx":
                workbook = load_workbook(path, read_only=True, data_only=True)
                for sheet in workbook.worksheets:
                    rows = [list(row) for _, row in zip(range(45), sheet.iter_rows(values_only=True))]
                    idx, hmap, _, _ = _detect_header_advanced(rows, scan_limit=40)
                    if idx is not None and _qualifies_quote_header(hmap, strict_minimum_signals=True):
                        if likely_site_roster_header_row(rows[idx], hmap):
                            continue
                        bad, _ = _sheet_is_false_positive(rows, idx, hmap)
                        if not bad:
                            return True
            elif suffix in {".csv", ".txt"}:
                content = path.read_text(encoding="utf-8", errors="ignore")
                sample_rows = [re.split(r"[,\t|]", line) for line in content.splitlines()[:45] if line.strip()]
                idx, hmap, _, _ = _detect_header_advanced(sample_rows, scan_limit=40)
                if idx is not None and _qualifies_quote_header(hmap, strict_minimum_signals=True):
                    if likely_site_roster_header_row(sample_rows[idx], hmap):
                        return False
                    bad, _ = _sheet_is_false_positive(sample_rows, idx, hmap)
                    if not bad:
                        return True
        except Exception:
            return False
        return False

    @staticmethod
    def _is_quote_header_map(header_map: dict[str, int]) -> bool:
        return _qualifies_quote_header(header_map)

    def _parse_xlsx(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        # Evaluate uncalculated formula cells (subtotals, grand totals, extended
        # cost = qty*price) so a workbook Excel never recalculated doesn't ship
        # blank "Amount" cells. Keyed (sheet, col, row) -> value; overlaid onto
        # None cells below. Best-effort: unparseable formulas stay blank.
        try:
            from app.parsers.xlsx_formula_eval import evaluate_workbook_formulas
            _formula_vals = evaluate_workbook_formulas(str(path))
        except Exception:
            _formula_vals = {}
        atoms: list[EvidenceAtom] = []
        # Build a fallback XlsxParser once so per-sheet "no quote
        # shape" sheets (Pricing tables, Site lists, Notes tabs) can
        # still emit row atoms via the generic xlsx path.
        fallback_parser = None
        try:
            from app.parsers.xlsx_parser import XlsxParser
            fallback_parser = XlsxParser()
        except Exception:  # pragma: no cover
            fallback_parser = None
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if _formula_vals:
                for _ri, _row in enumerate(rows, 1):
                    for _ci in range(len(_row)):
                        if _row[_ci] is None:
                            _fv = _formula_vals.get((sheet.title, _ci + 1, _ri))
                            if _fv is not None:
                                _row[_ci] = int(_fv) if float(_fv).is_integer() else _fv
            sheet_atoms = self._parse_sheet(
                project_id=project_id,
                artifact_id=artifact_id,
                filename=path.name,
                sheet_name=sheet.title,
                artifact_type=ArtifactType.xlsx,
                rows=rows,
            )
            if sheet_atoms:
                atoms.extend(sheet_atoms)
                continue
            # Sheet produced 0 atoms via the quote path. Try the
            # generic xlsx row emitter (catches Pricing / Site Roster /
            # Notes tabs that don't fit the quote schema).
            if fallback_parser is not None:
                try:
                    fb = fallback_parser._parse_sheet_rows(
                        project_id=project_id,
                        artifact_id=artifact_id,
                        filename=path.name,
                        artifact_type=ArtifactType.xlsx,
                        sheet_name=sheet.title,
                        rows=rows,
                    )
                except Exception:  # pragma: no cover
                    fb = []
                if fb:
                    atoms.extend(fb)
        return atoms

    def _parse_csv(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        text_head = path.read_text(encoding="utf-8", errors="ignore")[:8192]
        first_line = text_head.splitlines()[0] if text_head else ""
        delimiter = ","
        if first_line.count("|") > first_line.count(",") and "|" in first_line:
            delimiter = "|"
        elif "\t" in first_line and first_line.count("\t") > first_line.count(","):
            delimiter = "\t"
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            rows = [list(row) for row in reader]
        return self._parse_sheet(
            project_id=project_id,
            artifact_id=artifact_id,
            filename=path.name,
            sheet_name="csv",
            artifact_type=ArtifactType.csv,
            rows=rows,
        )

    def _parse_txt(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        content = path.read_text(encoding="utf-8", errors="ignore")
        lines = [line for line in content.splitlines() if line.strip()]
        rows = [re.split(r"[,\t|]", line) for line in lines]
        return self._parse_sheet(
            project_id=project_id,
            artifact_id=artifact_id,
            filename=path.name,
            sheet_name="txt",
            artifact_type=ArtifactType.txt,
            rows=rows,
        )

    def _parse_sheet(
        self,
        project_id: str,
        artifact_id: str,
        filename: str,
        sheet_name: str,
        artifact_type: ArtifactType,
        rows: list[list[Any]],
    ) -> list[EvidenceAtom]:
        if not rows:
            return []
        # Sheet-role gate (shared with xlsx_parser). Only DROP-destination
        # sheets (empty / cover / instruction / lookup-helper noise) are
        # skipped here. COMMERCIAL-destination sheets (rate cards, price
        # catalogs, deal-financials) are exactly what a quote parser
        # *should* mine — they fall through to normal line-item parsing so
        # their pricing is captured rather than discarded.
        if classify_sheet(sheet_name, rows).destination is SheetDestination.DROP:
            return []
        header_idx, header_map, header_mode, diag = _detect_header_advanced(rows, scan_limit=40)
        if header_idx is None or not header_map:
            return []
        bad, reason = _sheet_is_false_positive(rows, header_idx, header_map)
        if bad:
            return []

        data_start = header_idx + (2 if header_mode == "pair" else 1)
        meta = _scan_quote_metadata(rows[:data_start])
        atoms: list[EvidenceAtom] = []
        if meta:
            atoms.append(
                self._constraint_atom(
                    project_id,
                    artifact_id,
                    filename,
                    sheet_name,
                    artifact_type,
                    header_idx + 1,
                    header_map,
                    {"quote_parser_metadata": meta, "diagnostics": diag + ["quote_metadata_scanned"]},
                    spreadsheet_cell_locator=False,
                )
            )

        # v49.2: emit raw_table_row atoms for centralized classification
        # in _enrich_table_atoms(). Parsers no longer call the schema
        # registry directly — they just preserve {_columns, _row} on
        # each row and one central function does all schema work.
        _raw_headers: list[str] = []
        if 0 <= header_idx < len(rows):
            _raw_headers = [str(c or "").strip() for c in rows[header_idx]]

        for row_idx in range(data_start, len(rows)):
            row = rows[row_idx]
            values = self._extract_row_values(row, header_map)
            if all(not str(v).strip() for v in values.values()):
                continue
            desc = values.get("description", "")
            part = values.get("part_number", "")
            qty_raw = values.get("quantity", "")
            notes = values.get("notes", "")
            up = values.get("unit_price", "")
            extp = values.get("extended_price", "")
            rk = _classify_row(desc, part, qty_raw, up, extp, notes)
            if rk in {"blank", "subtotal", "grand_total", "tax", "shipping", "discount", "section_header"}:
                continue
            if rk in {"malformed"} and not (desc or part):
                continue
            atoms.extend(
                self._row_to_atoms(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    filename=filename,
                    sheet_name=sheet_name,
                    artifact_type=artifact_type,
                    row_number=row_idx + 1,
                    header_map=header_map,
                    values=values,
                    row_kind=rk,
                    diagnostics=diag,
                    raw_columns=_raw_headers or None,
                    raw_row=[str(c or "").strip() for c in row] if _raw_headers else None,
                )
            )
            # v49.2: emit raw_table_row for centralized post-parse
            # classification. Replaces v49.0's per-parser schema call.
            if _raw_headers:
                _row_cells = [str(c or "").strip() for c in row]
                _row_text = " | ".join(c for c in _row_cells if c)[:4000]
                _rtr_id = stable_id("atm", artifact_id, "raw_table_row", sheet_name, row_idx)
                _rtr_src = SourceRef(
                    id=stable_id("src", _rtr_id),
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    locator={"sheet": sheet_name, "row": row_idx + 1, "extraction": "raw_table_row_v49_2"},
                    extraction_method="raw_table_row_v49_2",
                    parser_version=self.parser_version,
                )
                atoms.append(EvidenceAtom(
                    id=_rtr_id,
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=AtomType.raw_table_row,
                    raw_text=_row_text,
                    normalized_text=_row_text.lower(),
                    value={
                        "_columns": list(_raw_headers),
                        "_row": _row_cells,
                        "_table_idx": 0,
                        # 1-based so the typed atom (bom_line/service_line) shares
                        # the (sheet,row) cell key with the vendor_line_item /
                        # quantity emitted for this same row (which use
                        # row_idx+1) — lets the line-item-double suppressor and
                        # cross_type_dedup line them up.
                        "_row_idx": row_idx + 1,
                        "_filename": filename,
                        "_sheet": sheet_name,
                        "_artifact_type": "xlsx",
                    },
                    entity_keys=[],
                    source_refs=[_rtr_src],
                    receipts=[],
                    authority_class=AuthorityClass.vendor_quote,
                    confidence=0.80,
                    confidence_raw=0.80,
                    calibrated_confidence=0.80,
                    review_status=ReviewStatus.auto_accepted,
                    review_flags=[],
                    parser_version=self.parser_version,
                ))
        return atoms

    def _extract_row_values(self, row: list[Any], header_map: dict[str, int]) -> dict[str, str]:
        extracted: dict[str, str] = {}
        for key, idx in header_map.items():
            value = row[idx] if idx < len(row) else ""
            extracted[key] = str(value).strip() if value is not None else ""
        return extracted

    def _constraint_atom(
        self,
        project_id: str,
        artifact_id: str,
        filename: str,
        sheet_name: str,
        artifact_type: ArtifactType,
        row_number: int,
        header_map: dict[str, int],
        value: dict[str, Any],
        *,
        spreadsheet_cell_locator: bool = True,
    ) -> EvidenceAtom:
        if spreadsheet_cell_locator:
            columns = {key: get_column_letter(index + 1) for key, index in header_map.items()}
            locator: dict[str, Any] = {"sheet": sheet_name, "row": row_number, "columns": columns}
        else:
            locator = {"sheet": sheet_name, "row": row_number}
        source_ref = SourceRef(
            id=stable_id("src", artifact_id, sheet_name, row_number, "meta"),
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            locator=locator,
            extraction_method="quote_parser_metadata",
            parser_version=self.parser_version,
        )
        return EvidenceAtom(
            id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, "quote_meta"),
            project_id=project_id,
            artifact_id=artifact_id,
            atom_type=AtomType.constraint,
            raw_text="Quote parser metadata",
            normalized_text="quote parser metadata",
            value=value,
            entity_keys=[],
            source_refs=[source_ref],
            authority_class=AuthorityClass.vendor_quote,
            confidence=0.75,
            review_status=ReviewStatus.auto_accepted,
            review_flags=["quote_parser:metadata"],
            parser_version=self.parser_version,
        )

    def _row_to_atoms(
        self,
        project_id: str,
        artifact_id: str,
        filename: str,
        sheet_name: str,
        artifact_type: ArtifactType,
        row_number: int,
        header_map: dict[str, int],
        values: dict[str, str],
        row_kind: RowKind,
        diagnostics: list[str],
        raw_columns: list[str] | None = None,
        raw_row: list[str] | None = None,
    ) -> list[EvidenceAtom]:
        part_number = values.get("part_number", "")
        description = values.get("description", "")
        quantity_raw = values.get("quantity", "")
        unit_price_raw = values.get("unit_price", "")
        extended_raw = values.get("extended_price", "")
        lead_time = values.get("lead_time", "")
        material_spec = values.get("material_spec", "")
        included_raw = values.get("included", "")
        notes = values.get("notes", "")
        uom_cell = values.get("uom", "")
        section = values.get("section", "")
        # Site / location context — when the row has a Site ID
        # column, include it in the atom raw_text so downstream
        # entity_extraction picks up the site:* key. Without this,
        # BOM rows like "ATL-HQ-01 | C9300-48P-A | 5" lose the
        # site context after row-splitting.
        site_id_value = (
            values.get("site_id", "")
            or values.get("site", "")
            or values.get("location", "")
            or values.get("location_id", "")
            or ""
        ).strip()
        # Currency / amount context — when the row has explicit
        # Currency / Amount columns (multi-currency BOMs especially),
        # include them in the raw_text so money entity extraction
        # picks up USD/EUR/GBP/JPY codes and the amount values.
        currency_value = (values.get("currency", "") or "").strip()
        amount_value = (
            values.get("amount", "")
            or values.get("total", "")
            or values.get("value", "")
            # extended_price is the canonical bucket for "Amount" /
            # "Total" / "Line total" columns (HEADER_ALIASES routes
            # the keyword "amount" here).
            or values.get("extended_price", "")
            or ""
        )
        if not isinstance(amount_value, str):
            amount_value = str(amount_value)
        amount_value = amount_value.strip()
        # Region tag (US / EMEA / UK / APAC) — useful for cross-doc
        # reasoning even though it's not a canonical entity.
        region_value = (values.get("region", "") or "").strip()

        inc_obj = normalize_inclusion(included_raw, notes)
        qty_obj = parse_quote_quantity(description, quantity_raw, uom_cell, notes)
        up_money = parse_money_cell(unit_price_raw, side="unit")
        ext_money = parse_money_cell(extended_raw, side="extended")
        mat = _material_heuristics(description, material_spec, notes)

        included_bool = inc_obj.get("included")
        inclusion_status = inc_obj.get("inclusion_status") or "unknown"

        entity_keys: list[str] = []
        if description:
            # Service-vs-device classification — labor / support / training
            # / hypercare / governance / consulting line items are routed
            # to `service:` instead of `device:` so the device namespace
            # contains only physical hardware.
            etype = "service" if _looks_like_service_description(description) else "device"
            entity_keys.append(normalize_entity_key(etype, description))
        if part_number:
            entity_keys.append(normalize_entity_key("part", part_number))

        used_keys = {k for k, v in values.items() if str(v or "").strip()}
        columns = {k: get_column_letter(header_map[k] + 1) for k in used_keys if k in header_map}
        column_count = len(columns)
        universals = _vendor_line_universal_fields(
            header_map=header_map,
            row_kind=row_kind,
            qty_obj=qty_obj,
            inc_obj=inc_obj,
            mat=mat,
            inclusion_status=inclusion_status,
            included_bool=included_bool,
            column_count=column_count,
            filename=filename,
            sheet_name=sheet_name,
            row_number=row_number,
        )

        source_ref = SourceRef(
            id=stable_id("src", artifact_id, sheet_name, row_number),
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            locator={"sheet": sheet_name, "row": row_number, "columns": columns},
            extraction_method="quote_header_mapping_v1_4_1",
            parser_version=self.parser_version,
        )

        atoms: list[EvidenceAtom] = []

        def append_atom(
            atom_type: AtomType,
            raw_text: str,
            value: dict[str, Any],
            confidence: float,
            flags: list[str],
            review_status: ReviewStatus,
        ) -> None:
            atoms.append(
                EvidenceAtom(
                    id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, atom_type.value, raw_text),
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=raw_text,
                    normalized_text=normalize_text(raw_text),
                    value=value,
                    entity_keys=entity_keys,
                    source_refs=[source_ref],
                    authority_class=AuthorityClass.vendor_quote,
                    confidence=confidence,
                    review_status=review_status,
                    review_flags=flags,
                    parser_version=self.parser_version,
                )
            )

        vli_value: dict[str, Any] = {
            "part_number": part_number,
            "description": description,
            "quantity": quantity_raw,
            "quantity_parsed": qty_obj,
            "unit_price_raw": unit_price_raw,
            "unit_price_parsed": up_money,
            "extended_price_raw": extended_raw,
            "extended_price_parsed": ext_money,
            "lead_time": lead_time,
            "material_spec": material_spec,
            "section": section,
            "notes": notes,
            "included": included_bool,
            "inclusion_status": inclusion_status,
            "row_kind": row_kind,
            "normalized_item": mat["normalized_item"],
            "material_family": mat.get("material_family"),
            "cable_category": mat.get("cable_category"),
            "shielding": mat.get("shielding"),
            "jacket_rating": mat.get("jacket_rating"),
            "item_kind": mat.get("item_kind"),
            "port_count": mat.get("port_count"),
            "is_scope_pollution_candidate": mat.get("is_scope_pollution_candidate"),
            "parser_diagnostics": diagnostics[:12],
            **universals,
        }
        vli_value = merge_parser_value_identity(
            vli_value,
            raw_text=f"{part_number} {description} {material_spec} {notes}".strip(),
        )
        # Bound (header: value) view so the classifier sees a RICH line instead
        # of the thin "NetWave AP-9700" raw_text. Bind the FULL raw row — EVERY
        # column the sheet has (Item ID, Unit, Purpose, Requires Serial Capture,
        # ...), not a hardcoded canonical subset that silently drops the rest.
        # _atom_bound_text renders these at decide-time; raw_text is left
        # untouched so the vendor-mismatch graph / commercial packet logic is
        # unaffected. Fall back to the canonical fields only when the raw row
        # wasn't threaded in (csv/txt paths).
        if raw_columns and raw_row:
            vli_value["_columns"] = list(raw_columns)
            vli_value["_row"] = list(raw_row)
        else:
            vli_value["cells"] = {
                k: str(v).strip()
                for k, v in (
                    ("SKU/Part", part_number),
                    ("Description", description),
                    ("Material", material_spec),
                    ("Qty", qty_obj.get("quantity_raw") or quantity_raw),
                    ("Unit Price", unit_price_raw),
                    ("Extended", extended_raw),
                    ("Lead Time", lead_time),
                    ("Notes", notes),
                )
                if str(v or "").strip()
            }

        has_line = bool(
            part_number
            or description
            or (quantity_raw != "" and quantity_raw is not None)
            or unit_price_raw
            or extended_raw
            or material_spec
            or str(included_raw or "").strip()
            or notes
        )
        flags: list[str] = []
        if qty_obj.get("uncertain"):
            flags.append("quote_parser:ambiguous_quantity")
        if up_money.get("price_status") == "malformed" and unit_price_raw.strip():
            flags.append("quote_parser:malformed_money")
        if ext_money.get("price_status") == "malformed" and extended_raw.strip():
            flags.append("quote_parser:malformed_extended_money")
        if _price_math_mismatch(qty_obj, up_money, ext_money):
            flags.append("quote_parser:price_math_mismatch")

        review_status = (
            ReviewStatus.needs_review
            if _quote_row_needs_review(
                row_kind,
                qty_obj,
                inclusion_status,
                unit_price_raw,
                extended_raw,
                up_money,
                ext_money,
                flags,
            )
            else ReviewStatus.auto_accepted
        )

        if has_line and row_kind in {"real_line_item", "allowance", "alternate", "option", "excluded", "included_no_qty", "malformed"}:
            append_atom(
                AtomType.vendor_line_item,
                (f"Line item {part_number} {description}".strip()
                 + (f" at {site_id_value}" if site_id_value else "")
                 + (f" [{region_value}]" if region_value else "")
                 + (f" {currency_value} {amount_value}".strip() if (currency_value and amount_value) else "")).strip(),
                vli_value,
                0.88 if not flags else 0.72,
                flags,
                review_status,
            )

        q_emit = qty_obj.get("quantity") is not None or qty_obj.get("quantity_status") in {
            "zero",
            "included_no_qty",
            "tbd",
            "not_applicable",
            "range",
            "allowance",
        }
        if q_emit or (quantity_raw != "" and quantity_raw is not None):
            qval = dict(qty_obj)
            qval["legacy"] = parse_quantity(quantity_raw) if quantity_raw else parse_quantity(str(qty_obj.get("quantity") or ""))
            qval.update(
                {
                    "comparison_key": universals["comparison_key"],
                    "commercial_role": universals["commercial_role"],
                    "scope_relevance": universals["scope_relevance"],
                    "authority_boundary": universals["authority_boundary"],
                    "source_row_key": universals["source_row_key"],
                    "comparison_basis": "vendor_proposed_quantity",
                    "included": included_bool,
                    "inclusion_status": inclusion_status,
                    "item_kind": mat.get("item_kind"),
                    "cable_category": mat.get("cable_category"),
                    "shielding": mat.get("shielding"),
                    "jacket_rating": mat.get("jacket_rating"),
                    "normalized_item": mat.get("normalized_item"),
                }
            )
            qval = merge_parser_value_identity(
                qval,
                raw_text=f"{description} {material_spec} {notes} {quantity_raw}".strip(),
            )
            # Context view so the head sees WHY this isolated quantity exists:
            # it's the vendor-proposed quantity for THIS item, the unit the
            # cross-doc quantity-contradiction check reconciles against the
            # scoped / approved-roster quantity. Renders as
            # "Item: ... | Vendor Qty: 94 | Reconcile against: scoped quantity".
            qval["cells"] = {
                k: str(v).strip()
                for k, v in (
                    ("Item", part_number or description or mat.get("normalized_item")),
                    ("Vendor Qty", qty_obj.get("quantity_raw") or quantity_raw),
                    ("Site", site_id_value),
                    ("Reconcile against", "scoped / approved-roster quantity for this item"),
                )
                if str(v or "").strip()
            }
            # The vendor_line_item now carries the quantity (qty is a FIELD on
            # the line, and graph_builder's _is_qty_node treats the line as the
            # reconciliation node). The standalone `quantity` atom is therefore a
            # redundant comparison node — reconciliation is a deal-level EDGE
            # between lines, not a parse-time fact. We still EMIT it (its full
            # pipeline removal — parser contract + quantity_claim packet across
            # every deal — is the Phase-2 conflicts/verify refactor), but it is
            # suppressed from the head's view as a reconciliation-only input.
            # Set SOWSMITH_DROP_QUANTITY_ATOM=1 to stop emitting entirely.
            if os.environ.get("SOWSMITH_DROP_QUANTITY_ATOM") != "1":
                append_atom(
                    AtomType.quantity,
                    (f"Quantity {qty_obj.get('quantity_raw') or quantity_raw}"
                     + (f" {part_number}" if part_number else "")
                     + (f" at {site_id_value}" if site_id_value else "")).strip(),
                    qval,
                    0.88 if not qty_obj.get("uncertain") else 0.7,
                    list(flags),
                    review_status,
                )

        if lead_time:
            append_atom(
                AtomType.constraint,
                f"Lead time {lead_time}",
                {"lead_time": lead_time},
                0.85,
                [],
                ReviewStatus.auto_accepted,
            )
        return atoms
