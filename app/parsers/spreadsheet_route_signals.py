"""Shared signals for routing .xlsx/.csv between quote and xlsx parsers."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.normalizers import normalize_text

# Tokens matched as path segments / filename stems (not naive substrings — e.g. "po" must not match "repo").
QUOTE_PATH_TOKENS = frozenset(
    {
        "quote",
        "vendor",
        "po",
        "purchase_order",
        "bom",
        "pricing",
        "estimate",
        "proposal",
        "material_quote",
    }
)
_SHORT_QUOTE_PATH_TOKENS = frozenset({"po", "bom"})
ROSTER_SCHEDULE_PATH_TOKENS = frozenset(
    {
        "site_list",
        "roster",
        "schedule",
        "drop_schedule",
        "cable_schedule",
        "room_schedule",
        "device_schedule",
        "access_tracker",
        "scope_matrix",
    }
)

WIDE_SCHEDULE_MATERIAL_MARKERS = (
    "rj45",
    "cat6 utp",
    "cat6 stp",
    "cat6a",
    "cat5",
    "fiber",
    "smf",
    "mmf",
    "om3",
    "om4",
)


def _segment_matches_quote_token(seg: str, tok: str) -> bool:
    s = seg.lower()
    base = s.split(".", 1)[0]
    if tok not in s and tok not in base:
        return False
    if tok not in _SHORT_QUOTE_PATH_TOKENS:
        return tok in base or tok in s
    chunks = [c for c in re.split(r"[-_.\s]+", base) if c]
    chunks += [c for c in re.split(r"[-_.\s]+", s) if c]
    return tok in chunks or base == tok


def _path_segments_for_routing_tokens(path: Path, *, max_levels: int = 5) -> list[str]:
    """
    Walk up from the file toward the root, collecting at most ``max_levels`` names.

    Using the full absolute path string matched pytest tmp directory names (e.g.
    ``test_ip_camera_vendor_mismatch0``) against routing tokens like \"vendor\".
    """
    parts: list[str] = []
    cur: Path = path
    for _ in range(max_levels):
        parts.append(cur.name)
        parent = cur.parent
        if parent == cur:
            break
        cur = parent
    return list(reversed(parts))


def path_quote_filename_hint(path: Path) -> bool:
    for seg in _path_segments_for_routing_tokens(path):
        seg_l = seg.lower()
        if not seg_l:
            continue
        for tok in QUOTE_PATH_TOKENS:
            if _segment_matches_quote_token(seg_l, tok):
                return True
    return False


def path_roster_schedule_hint(path: Path) -> bool:
    tail = "/".join(_path_segments_for_routing_tokens(path)).lower()
    tail_compact = tail.replace(" ", "").replace("-", "_")
    if "sitelist" in tail_compact:
        return True
    return any(tok in tail for tok in ROSTER_SCHEDULE_PATH_TOKENS)


def _header_blob_from_row(row: list[Any]) -> str:
    parts: list[str] = []
    for cell in row:
        t = normalize_text(str(cell or "")).lower()
        if t:
            parts.append(t)
    return " ".join(parts)


def likely_site_roster_header_row(header_row: list[Any], hmap: dict[str, int]) -> bool:
    """Site roster: site/room/plate + device/hostname + qty, without commercial columns."""
    keys = set(hmap.keys())
    blob = _header_blob_from_row(header_row)
    has_siteish = bool(
        re.search(
            r"\b(site|building|campus|facility|room|floor|plate|drop|hostname|location|area)\b",
            blob,
            re.I,
        )
    )
    has_device_col = bool(re.search(r"\b(device|hostname|equipment)\b", blob, re.I))
    has_qty = "quantity" in keys
    commercial = bool(
        keys
        & {
            "unit_price",
            "extended_price",
            "part_number",
            "vendor",
            "included",
            "material_spec",
        }
    )
    return has_siteish and has_device_col and has_qty and not commercial


def _first_sheet_rows_xlsx(path: Path, max_rows: int = 80) -> list[list[Any]] | None:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return None
    try:
        sheet = wb.worksheets[0]
        return [list(row) for _, row in zip(range(max_rows), sheet.iter_rows(values_only=True))]
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass


def _first_sheet_rows_csv(path: Path, max_rows: int = 80) -> list[list[Any]] | None:
    try:
        text_head = path.read_text(encoding="utf-8", errors="ignore")[:65536]
        first_line = text_head.splitlines()[0] if text_head else ""
        delimiter = ","
        if first_line.count("|") > first_line.count(",") and "|" in first_line:
            delimiter = "|"
        elif "\t" in first_line and first_line.count("\t") > first_line.count(","):
            delimiter = "\t"
        rows: list[list[Any]] = []
        for line in text_head.splitlines()[:max_rows]:
            if not line.strip():
                continue
            rows.append(list(next(csv.reader([line], delimiter=delimiter))))
        return rows or None
    except Exception:  # noqa: BLE001
        return None


def peek_first_table(path: Path) -> tuple[list[list[Any]] | None, list[str]]:
    """Load first tabular preview for path; reasons describe peek outcome."""
    reasons: list[str] = []
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = _first_sheet_rows_xlsx(path)
        if rows is None:
            reasons.append("peek:xlsx_load_failed")
        else:
            reasons.append(f"peek:xlsx_rows={len(rows)}")
        return rows, reasons
    if suffix == ".csv":
        rows = _first_sheet_rows_csv(path)
        if rows is None:
            reasons.append("peek:csv_load_failed")
        else:
            reasons.append(f"peek:csv_rows={len(rows)}")
        return rows, reasons
    reasons.append("peek:unsupported_suffix")
    return None, reasons


def sniff_quote_header_ok(path: Path) -> tuple[bool, list[str]]:
    """True if workbook qualifies as quote-like (not a bare roster Device+Qty layout)."""
    from app.parsers.quote_parser import _detect_header_advanced, _sheet_is_false_positive, _qualifies_quote_header

    reasons: list[str] = []
    rows, pr = peek_first_table(path)
    reasons.extend(pr)
    if not rows:
        reasons.append("quote_sniff:no_rows")
        return False, reasons
    idx, hmap, _, diag = _detect_header_advanced(rows, scan_limit=40)
    reasons.extend(diag)
    if idx is None:
        reasons.append("quote_sniff:no_header_row")
        return False, reasons
    header_row = rows[idx]
    if likely_site_roster_header_row(header_row, hmap):
        reasons.append("quote_sniff:rejected_site_roster_device_qty")
        return False, reasons
    bad, why = _sheet_is_false_positive(rows, idx, hmap)
    if bad:
        reasons.append(f"quote_sniff:false_positive:{why}")
        return False, reasons
    ok = _qualifies_quote_header(hmap, strict_minimum_signals=True)
    reasons.append(f"quote_sniff:header_qualifies={ok}")
    return ok, reasons


def quote_commercial_header_score(hmap: dict[str, int]) -> float:
    """How strongly headers look like commercial / BOM (0..1)."""
    keys = set(hmap.keys())
    score = 0.0
    if "unit_price" in keys or "extended_price" in keys:
        score += 0.45
    if "part_number" in keys:
        score += 0.25
    if "included" in keys:
        score += 0.2
    if "vendor" in keys:
        score += 0.15
    if "material_spec" in keys:
        score += 0.1
    return min(1.0, score)


def _wide_schedule_material_count(row: list[Any]) -> int:
    n = 0
    for cell in row:
        t = normalize_text(str(cell or "")).lower()
        if not t:
            continue
        if any(m in t for m in WIDE_SCHEDULE_MATERIAL_MARKERS):
            n += 1
    return n


def sniff_xlsx_roster_schedule_strength(path: Path) -> tuple[float, list[str]]:
    """0..1 strength that this is roster/schedule rather than quote."""
    from app.parsers.quote_parser import _detect_header_advanced

    reasons: list[str] = []
    score = 0.0
    if path_roster_schedule_hint(path):
        score += 0.45
        reasons.append("xlsx_sniff:path_roster_schedule_token")
    rows, pr = peek_first_table(path)
    reasons.extend(pr)
    if not rows:
        reasons.append("xlsx_sniff:no_rows")
        return score, reasons
    idx, hmap, _, _ = _detect_header_advanced(rows, scan_limit=40)
    if idx is None:
        reasons.append("xlsx_sniff:no_header_row")
        return score, reasons
    header_row = rows[idx]
    blob = _header_blob_from_row(header_row)
    if re.search(r"\b(plate|drop|rj45|room|site|building|hostname|device)\b", blob, re.I):
        score += 0.2
        reasons.append("xlsx_sniff:schedule_roster_header_tokens")
    if _wide_schedule_material_count(header_row) >= 2:
        score += 0.25
        reasons.append("xlsx_sniff:wide_material_columns")
    if "quantity" in hmap:
        score += 0.05
        reasons.append("xlsx_sniff:has_qty_column")
    commercial_keys = hmap.keys() & {"unit_price", "extended_price", "part_number", "vendor", "included"}
    if commercial_keys:
        score = max(0.0, score - 0.35)
        reasons.append(f"xlsx_sniff:penalize_commercial_columns:{sorted(commercial_keys)}")
    score = min(1.0, score)
    reasons.append(f"xlsx_sniff:strength={score:.2f}")
    return score, reasons


def resolve_quote_vs_xlsx_tie(path: Path) -> tuple[str | None, list[str]]:
    """
    When quote and xlsx are both top matches, pick parser or None for legacy tie-break.

    Returns (parser_name or None, accumulated reasons).
    """
    from app.parsers.quote_parser import _detect_header_advanced

    reasons: list[str] = []
    qpath = path_quote_filename_hint(path)
    rpath = path_roster_schedule_hint(path)
    quote_ok, qr = sniff_quote_header_ok(path)
    reasons.extend(qr)
    xscore, xr = sniff_xlsx_roster_schedule_strength(path)
    reasons.extend(xr)

    rows, _ = peek_first_table(path)
    comm_score = 0.0
    if rows:
        idx, hmap, _, _ = _detect_header_advanced(rows, scan_limit=40)
        if idx is not None:
            comm_score = quote_commercial_header_score(hmap)
            reasons.append(f"tie:commercial_header_score={comm_score:.2f}")

    if qpath and not rpath:
        reasons.append("tie_resolve:quote_path_tokens_only")
        return "quote", reasons
    if rpath and not qpath:
        reasons.append("tie_resolve:roster_schedule_path_tokens_only")
        return "xlsx", reasons

    if qpath and rpath:
        reasons.append("tie:path_has_both_quote_and_schedule_tokens")

    if quote_ok and xscore < 0.35:
        reasons.append("tie_resolve:quote_headers_clear_low_schedule_signal")
        return "quote", reasons
    if not quote_ok and (rpath or xscore >= 0.45):
        reasons.append("tie_resolve:not_quote_headers_or_strong_schedule")
        return "xlsx", reasons

    if quote_ok and (rpath or xscore >= 0.45):
        if comm_score >= 0.25:
            reasons.append("tie_resolve:ambiguous_prefer_quote_commercial_headers")
            return "quote", reasons
        reasons.append("tie_resolve:ambiguous_prefer_xlsx_schedule_roster_signals")
        return "xlsx", reasons

    if not quote_ok and not rpath and qpath and xscore < 0.35:
        reasons.append("tie_resolve:quote_path_weak_schedule_headers_not_quote_like")
        return "quote", reasons

    if not quote_ok:
        reasons.append("tie_resolve:default_xlsx_no_quote_header_match")
        return "xlsx", reasons

    reasons.append("tie_resolve:quote_headers_default")
    return "quote", reasons
