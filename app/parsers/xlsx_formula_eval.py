"""Lightweight Excel formula evaluator for uncalculated workbooks.

openpyxl(data_only=True) returns ``None`` for formula cells when the file was
saved without Excel recalculating — so subtotals / grand totals / extended-cost
columns come through blank even though the value "is there" in Excel. No formula
engine is bundled, and a full one is overkill, so this evaluates the patterns
that actually show up in pricing/BOM sheets:

  • numeric constants
  • cell arithmetic            =F2*G2,  =A1+B1-C1,  =D2/2
  • SUM over a range/column    =SUM('Hardware BOM'!H:H),  =SUM(B2:B6)

Dependencies (extended cost -> subtotal -> grand total) are resolved by
iterating to a fixed point. Anything it can't evaluate (IF, VLOOKUP, ...) is
left alone for the caller to fall back on. Pure, no I/O beyond the initial load.
"""
from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import column_index_from_string

# A cell value map keyed by (sheet_title, col_index, row_index) -> float.
_CellMap = dict[tuple[str, int, int], float]

_REF_RE = re.compile(
    r"(?:'(?P<q>[^']+)'|(?P<s>[A-Za-z_][A-Za-z0-9_.]*))?!"  # optional sheet!
    r"?\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)"
)


def evaluate_workbook_formulas(path: str) -> _CellMap:
    """Return {(sheet, col, row): value} for formula cells openpyxl left blank."""
    from openpyxl import load_workbook

    wb_v = load_workbook(path, data_only=True)
    wb_f = load_workbook(path, data_only=False)

    vals: _CellMap = {}
    forms: dict[tuple[str, int, int], str] = {}
    for ws in wb_v.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                    vals[(ws.title, c.column, c.row)] = float(c.value)
    for ws in wb_f.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    key = (ws.title, c.column, c.row)
                    if key not in vals:
                        forms[key] = c.value[1:].strip()

    # Fixed-point: re-evaluate unresolved formulas until nothing new resolves.
    for _ in range(25):
        changed = False
        for key, formula in forms.items():
            if key in vals:
                continue
            v = _eval(formula, key[0], vals, forms)
            if v is not None:
                vals[key] = v
                changed = True
        if not changed:
            break
    return vals


def _eval(formula: str, sheet: str, vals: _CellMap, forms: dict) -> float | None:
    f = formula.strip()
    m = re.fullmatch(r"(?i)\s*sum\((?P<arg>.+)\)\s*", f)
    if m:
        return _eval_sum(m.group("arg"), sheet, vals, forms)
    # Arithmetic: substitute every cell ref with its value, then eval the math.
    expr_parts: list[str] = []
    last = 0
    for ref in _REF_RE.finditer(f):
        ref_sheet = ref.group("q") or ref.group("s") or sheet
        col = column_index_from_string(ref.group("col"))
        row = int(ref.group("row"))
        key = (ref_sheet, col, row)
        if key in vals:
            v = vals[key]
        elif key in forms:
            return None  # dependency not ready — retry next pass
        else:
            v = 0.0  # empty cell = 0 in Excel arithmetic
        expr_parts.append(f[last:ref.start()])
        expr_parts.append(repr(v))
        last = ref.end()
    expr_parts.append(f[last:])
    expr = "".join(expr_parts)
    if not re.fullmatch(r"[-+*/(). 0-9eE]+", expr.strip()):
        return None
    try:
        return float(eval(expr, {"__builtins__": {}}, {}))  # noqa: S307 (sanitized)
    except Exception:
        return None


def _eval_sum(arg: str, cur_sheet: str, vals: _CellMap, forms: dict) -> float | None:
    arg = arg.strip()
    sheet = cur_sheet
    if "!" in arg:
        sh, rng = arg.split("!", 1)
        sheet = sh.strip().strip("'")
    else:
        rng = arg
    parts = rng.strip().split(":")
    if len(parts) == 1:
        parts = [parts[0], parts[0]]
    a = re.match(r"\$?([A-Z]{1,3})\$?(\d*)", parts[0].strip())
    b = re.match(r"\$?([A-Z]{1,3})\$?(\d*)", parts[1].strip())
    if not a or not b:
        return None
    c1, c2 = column_index_from_string(a.group(1)), column_index_from_string(b.group(1))
    r1 = int(a.group(2)) if a.group(2) else None
    r2 = int(b.group(2)) if b.group(2) else None
    lo_c, hi_c = min(c1, c2), max(c1, c2)

    def _in_range(col: int, row: int) -> bool:
        if not (lo_c <= col <= hi_c):
            return False
        if r1 is None:  # whole-column SUM
            return True
        return min(r1, r2) <= row <= max(r1, r2)

    # If any cell in the range is still an unresolved formula, retry next pass.
    for (s, col, row) in forms:
        if s == sheet and (s, col, row) not in vals and _in_range(col, row):
            return None
    total = 0.0
    for (s, col, row), v in vals.items():
        if s == sheet and _in_range(col, row):
            total += v
    return total


__all__ = ["evaluate_workbook_formulas"]
