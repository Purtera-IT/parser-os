from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.ids import stable_id
from app.core.item_identity import merge_parser_value_identity
from app.core.normalizers import normalize_entity_key, normalize_text, parse_quantity
from app.core.segments import ArtifactSegment
from app.core.schemas import (
    ArtifactType,
    AtomType,
    AuthorityClass,
    EvidenceAtom,
    ParserOutput,
    ReviewStatus,
    SourceRef,
    ParserCapability,
    ParserMatch,
)
from app.parsers.base import BaseParser
from app.parsers.binary_markers import emit_zip_binary_markers
from app.parsers.segmenters import segment_xlsx
from app.parsers.sheet_classifier import (
    SheetDestination,
    SheetRole,
    classify_sheet,
)
from app.parsers.structured_projection import (
    derived_files_for,
    make_page,
    make_section,
    make_structured_document,
    make_table,
    stamp_section_and_block_ids,
)
from app.domain.schemas import DomainPack

parser_name = "xlsx"
parser_version = "xlsx_parser_v2_1"

# Consumed by parser-os-service `_attachments_status` after compile.
ARTIFACT_PARSE_ERROR_PREFIX = "artifact_parse_error:"
STRUCTURED_SCHEMA_XLSX = "orbitbrief.xlsx.structured.v1"
STRUCTURED_SCHEMA_CSV = "orbitbrief.csv.structured.v1"

# Cell-fact patterns (PR2). When a generic row's individual cell text
# matches one of these patterns we ALSO emit a sub-atom typed as
# ``exclusion`` or ``risk``, anchored to the parent row by
# ``value.parent_row_atom_id``. This stops "EOL", "not included",
# "single point of failure", … from being buried inside generic
# row text where the packetizer can't find them.
_EXCLUSION_CELL_RE = re.compile(
    # PR3 (post-v3 review) — dropped "unless noted". That phrase is
    # almost always a SUPPORT-TIER conditional ("8x5 vendor support
    # unless noted"), not a contractual exclusion. The packetizer
    # exclusion gate now also checks the canonical_field so a cell
    # in a support_level / coverage / support_entitlement column
    # never participates in a scope_exclusion packet.
    r"\b(exclud(?:e|ed|es|ing)|not covered|not included|unsupported|"
    r"no sla|outside coverage)\b",
    re.I,
)
_RISK_CELL_RE = re.compile(
    r"\b(eol|end of life|unsupported|critical|high risk|"
    r"single point of failure|mismatch|missing)\b",
    re.I,
)
# PR3 — conditional support boundary phrasing. "8x5 vendor support
# unless noted" / "no SLA if expired" / "coverage exclusion if expired".
# These are SUPPORT TIER descriptions, not contractual exclusions.
_CONDITIONAL_SUPPORT_RE = re.compile(
    r"\b(unless\s+noted|sla\s+exclusion\s+if\s+expired|"
    r"no\s+sla\s+if\s+expired|coverage\s+exclusion\s+if\s+expired|"
    r"unless\s+otherwise\s+noted|where\s+applicable)\b",
    re.I,
)
_SUPPORT_LEVEL_FIELDS: frozenset[str] = frozenset(
    {
        "support_level",
        "coverage",
        "support_entitlement",
        "support_tier",
    }
)


# PR2 (post-v3 review) — per-sheet profile for operational workbooks.
# Used by ``XlsxParser._emit_operational_sheet_rows`` to dispatch
# every row of every named sheet to its proper AtomType + value.kind
# + authority class.
@dataclass(frozen=True)
class _OperationalSheetProfile:
    atom_type: AtomType
    kind: str
    authority_class: AuthorityClass
    confidence: float


_OPERATIONAL_SHEET_PROFILES: dict[str, _OperationalSheetProfile] = {
    "readme": _OperationalSheetProfile(
        AtomType.project_metadata,
        "project_metadata_row",
        AuthorityClass.customer_current_authored,
        0.90,
    ),
    "dashboard": _OperationalSheetProfile(
        AtomType.project_metadata,
        "dashboard_metric_row",
        AuthorityClass.customer_current_authored,
        0.90,
    ),
    "asset inventory": _OperationalSheetProfile(
        AtomType.asset_record,
        "asset_inventory_row",
        AuthorityClass.approved_site_roster,
        0.94,
    ),
    "site survey raw": _OperationalSheetProfile(
        AtomType.site_survey_row,
        "site_survey_row",
        AuthorityClass.customer_current_authored,
        0.92,
    ),
    "port map & vlans": _OperationalSheetProfile(
        AtomType.port_vlan_assignment,
        "port_vlan_assignment_row",
        AuthorityClass.approved_site_roster,
        0.94,
    ),
    "port map vlans": _OperationalSheetProfile(
        AtomType.port_vlan_assignment,
        "port_vlan_assignment_row",
        AuthorityClass.approved_site_roster,
        0.94,
    ),
    "circuit inventory": _OperationalSheetProfile(
        AtomType.circuit_inventory,
        "circuit_inventory_row",
        AuthorityClass.approved_site_roster,
        0.92,
    ),
    "license support": _OperationalSheetProfile(
        AtomType.support_entitlement,
        "support_entitlement_row",
        AuthorityClass.vendor_quote,
        0.92,
    ),
    "noc alert matrix": _OperationalSheetProfile(
        AtomType.alert_route,
        "alert_route_row",
        AuthorityClass.customer_current_authored,
        0.92,
    ),
    "risk register": _OperationalSheetProfile(
        AtomType.risk,
        "risk_register_row",
        AuthorityClass.customer_current_authored,
        0.93,
    ),
    "cutover validation": _OperationalSheetProfile(
        AtomType.cutover_validation,
        "cutover_validation_row",
        AuthorityClass.customer_current_authored,
        0.93,
    ),
    "source refs": _OperationalSheetProfile(
        AtomType.project_metadata,
        "source_reference_row",
        AuthorityClass.machine_extractor,
        0.75,
    ),
}


# Service-line classifier — when a BOM line item description matches
# one of these tokens (case-insensitive substring match), the line is
# routed to `service:` instead of `device:`. Real device names rarely
# include these tokens; service line items almost always do.
_SERVICE_LINE_TOKENS: tuple[str, ...] = (
    "labor", "labour", "hours", "hour", "after-hours", "after hours",
    "support", "supports", "supported",
    "training", "trainings", "adoption",
    "workshop", "workshops", "discovery",
    "design", "designs", "engineering services",
    "professional services", "managed services",
    "consulting", "consultancy", "advisory",
    "hypercare", "warranty", "warrantee",
    "project management", "program management", "pmo",
    "governance", "oversight",
    "implementation", "installation services", "deployment",
    "commissioning", "decommissioning",
    "migration", "cutover", "go-live",
    "documentation services", "as-built", "as built",
    "testing services", "validation", "uat", "acceptance",
    "rfp response", "proposal preparation",
)


def _looks_like_service_line(value: str) -> bool:
    """Return True if a BOM line-item description is a service rather
    than a physical device.

    Routes labor / support / training / hypercare / governance /
    consulting / project-management line items to `service:<slug>`
    instead of `device:<slug>` so the device namespace stays clean.
    """
    lower = value.lower()
    return any(token in lower for token in _SERVICE_LINE_TOKENS)


def _norm_op_sheet(sheet_name: str) -> str:
    return normalize_text(sheet_name).replace("_", " ").strip()


# ── Commercial-sheet money extraction ───────────────────────────────
# Pricing in estimating workbooks is stored as bare numeric cells, not
# "$"-formatted text, so the symbol-based money extractor in
# entity_extraction misses it. These helpers recover money two ways:
#   • money columns  — a header cell naming a money concept (cost / sell /
#     price / rate / total / revenue / margin …) over numeric data cells
#     (master catalogs, rate cards).
#   • label→value    — a money-concept label cell paired with the nearest
#     numeric cell to its right on the same row (deal-financials summaries
#     like "Total Deal Revenue | 21560").
_MONEY_CONCEPT_RE = re.compile(
    r"revenue|cost|margin|price|\bsell\b|\bfee\b|budget|subtotal|amount"
    r"|\brate\b|charge|expense|extended|unit\s*price|total",
    re.I,
)

# SEMANTIC augmentation of the money-header judgment: a column header names a
# money concept by MEANING, not keyword — "MSRP", "Burdened Rate", "Line Extended"
# all mean money but dodge the keyword list. Union-only (can ADD a money column
# the regex missed, never remove one), so it cannot regress existing extraction;
# falls back to the regex when the embedder is offline.
_MONEY_HEADER_RULE = None


def _money_header_rule():
    global _MONEY_HEADER_RULE
    if _MONEY_HEADER_RULE is None:
        from app.core.semantic_rules import SemanticRule

        _MONEY_HEADER_RULE = SemanticRule(
            name="xlsx_money_header",
            positives=[
                "Unit Price", "Sell Price", "Unit Cost", "Extended Price",
                "Line Total", "Burdened Rate", "MSRP", "List Price",
                "Total Amount", "Labor Cost", "As-Sold Revenue", "Gross Margin",
                "Monthly Recurring Charge", "Discount", "Sell Rate",
            ],
            negatives=[
                "Site", "Quantity", "Description", "Job Description", "Notes",
                "City", "Building", "Contact", "SKU", "Part Number", "Country",
                "Unit", "Status", "Date", "Owner",
            ],
            threshold=0.60,
            lexical_fallback=lambda t: bool(_MONEY_CONCEPT_RE.search(t)),
        )
    return _MONEY_HEADER_RULE


def _is_money_header(text: str) -> bool:
    """Money-concept header: keyword fast-path (also the offline fallback), then a
    SEMANTIC second chance for short header-like labels the keyword list misses."""
    t = str(text or "").strip()
    if not t or "%" in t:
        return False
    if _MONEY_CONCEPT_RE.search(t):
        return True
    # bound the embedding to short, non-numeric, header-shaped cells
    if any(ch.isdigit() for ch in t) or not (0 < len(t.split()) <= 5):
        return False
    return _money_header_rule().fires(t)
# Floor that rejects tax multipliers (1.09 / 1.34), margin ratios (0.27),
# and other sub-dollar line noise while keeping genuine prices/totals.
_MIN_MONEY_VALUE = 5.0


def _is_money_number(value: Any) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and abs(value) >= _MIN_MONEY_VALUE
    )


def _rate_card_value_columns(rows: list[list[Any]]) -> set[int]:
    """Numeric rate columns on a per-country labor rate table.

    Rate-card sheets label columns ``Networking L1 Technician 2 hr. min`` —
    not ``$ Cost`` — so :func:`_money_columns` returns empty and the
    commercial emitter would emit nothing, triggering the coverage backstop
    that re-floods scope. This keys on the universal table shape: a
    ``Country`` header + technician/hour/min columns with numeric cells.
    """
    hdr_i = _first_nonblank_header_row(rows)
    if hdr_i is None:
        return set()
    hdr = [str(c or "").strip().lower() for c in rows[hdr_i]]
    if not any(h == "country" or h.startswith("country") for h in hdr):
        return set()
    cols: set[int] = set()
    for idx, h in enumerate(hdr):
        if not h:
            continue
        if re.search(
            r"technician|hr\.?\s*min|hour\s+minimum|networking\s+l[12]|^request$",
            h,
            re.I,
        ):
            cols.add(idx)
    if cols:
        return cols
    # Fallback: columns after Country with mostly-numeric data rows.
    data = [
        r for r in rows[hdr_i + 1 :]
        if any(str(c or "").strip() for c in r)
    ]
    if len(data) < 2:
        return set()
    ncols = max(len(r) for r in data)
    for ci in range(1, ncols):
        nums = sum(
            1 for r in data
            if ci < len(r) and _is_money_number(r[ci])
        )
        if nums >= max(2, int(0.6 * len(data))):
            cols.add(ci)
    return cols


def _money_columns(rows: list[list[Any]]) -> set[int]:
    """Column indices whose header names a money concept (% columns excluded)."""
    cols: set[int] = set()
    for row in rows[:20]:
        for idx, c in enumerate(row):
            t = str(c or "").strip()
            if t and "%" not in t and _MONEY_CONCEPT_RE.search(t):
                cols.add(idx)
    # Semantic second chance — catch money headers the keyword list misses
    # ("MSRP", "Burdened Rate", "Extended"), scoped to the header row only so the
    # embedding cost is bounded (~10-20 cells, cached). Union-only: never removes.
    hdr = _first_nonblank_header_row(rows)
    if hdr is not None and 0 <= hdr < len(rows):
        for idx, c in enumerate(rows[hdr]):
            if idx not in cols and _is_money_header(str(c or "").strip()):
                cols.add(idx)
    return cols


def _row_money_values(row: list[Any], money_cols: set[int]) -> list[float]:
    """Money amounts on a row, via money columns and label→value pairs."""
    vals: list[float] = []
    for idx in money_cols:
        if idx < len(row) and _is_money_number(row[idx]):
            vals.append(float(row[idx]))
    for idx, c in enumerate(row):
        t = str(c or "").strip()
        if not t or "%" in t or not _MONEY_CONCEPT_RE.search(t):
            continue
        for j in range(idx + 1, len(row)):
            if _is_money_number(row[j]):
                vals.append(float(row[j]))
                break
    return vals


def _is_side_label_value(cells: list[str]) -> bool:
    """A short ``label -> number`` fact sitting in a side calc block beside the
    main priced table (e.g. a travel breakdown's "Team | 4", "Weeks per tech |
    3"). It carries no money of its own, so the money gate would drop it — but
    it's real content, so keep it when it pairs a text label with a number."""
    nonblank = [c for c in cells if c]
    if not (2 <= len(nonblank) <= 6):
        return False

    def _numeric(c: str) -> bool:
        return c.replace(",", "").replace(".", "").replace("$", "").lstrip("-").isdigit()

    return any(not _numeric(c) for c in nonblank) and any(_numeric(c) for c in nonblank)


# ── Financial-summary (Deal Kit / P&L) structured extraction ─────────
# Deal-kit / estimating workbooks lay the deal economics out as a 2-D
# label→value grid, not a row table, so the generic row emitter mashes
# unrelated cells together. These helpers read the grid the way a human
# does: find a known label, take the value to its right. The vocabulary
# is the standard estimating-workbook field set (revenue/cost/margin by
# category + deal header fields) — universal across the org's deals, no
# customer-specific terms.

# P&L line metric: "<Category> Revenue|Cost|Margin" (with optional
# leading "Total"). Captures the category and which metric.
_PL_METRIC_RE = re.compile(
    # keep a leading "Total" INSIDE the category so the display label stays
    # faithful to the sheet ("Total Deal", "Total Labor"); _norm_pl_category
    # drops it for the grouping KEY only.
    r"^(?P<cat>(?:total\s+)?.+?)\s+(?P<metric>revenue|cost|margin)$", re.I
)
# Margin-percent line: "Margin % on <Category>".
_PL_MARGIN_PCT_RE = re.compile(r"^margin\s*%\s*on\s+(?P<cat>.+)$", re.I)

# Deal-header fields → normalized key. Matched against the lowercased,
# whitespace-collapsed label cell.
_DEAL_HEADER_LABELS: dict[str, str] = {
    "oppty #": "opportunity_id",
    "oppty#": "opportunity_id",
    "opportunity #": "opportunity_id",
    "opportunity#": "opportunity_id",
    "sales rep": "sales_rep",
    "customer": "customer",
    "end user": "end_user",
    "quoted w/ partner": "quoted_with_partner",
    "qty of sites": "site_count",
    "# of sites": "site_count",
    "division": "division",
    "project duration (months)": "project_duration",
    "project duration": "project_duration",
    "billing type": "billing_type",
    "region": "region",
    "channel/direct": "channel",
    "enterprise/ technical": "segment",
    "enterprise/technical": "segment",
    "date": "deal_date",
}

# Don't treat these as values when they sit to the right of a label —
# they are themselves column labels in the adjacent block of the grid.
_PL_METRIC_WORDS = frozenset({"revenue", "cost", "margin"})


# A margin-percent label doesn't always read "Margin % on <Category>". Deal kits
# phrase it many ways ("Net Margin (%)", "Gross Margin %", bare "Margin %") and
# sometimes carry a garbled template label ("Lift/Rental on Miscellaneous" in the
# margin-% slot). Recognise the MEANING by embedding instead of one rigid regex,
# so the recogniser generalises across every estimating workbook. Offline, the
# lexical fallback fires on the universal cue (a "margin" label with a percent
# marker); the embedder (dev/prod) additionally catches phrasings with neither.
_PL_MARGIN_PCT_RULE = None
_PCT_CUE_RE = re.compile(r"%|percent|\bpct\b", re.I)


def _margin_pct_lexical(text: str) -> bool:
    t = (text or "").lower()
    return "margin" in t and bool(_PCT_CUE_RE.search(t))


def _pl_margin_pct_rule():
    global _PL_MARGIN_PCT_RULE
    if _PL_MARGIN_PCT_RULE is None:
        from app.core.semantic_rules import SemanticRule
        _PL_MARGIN_PCT_RULE = SemanticRule(
            name="pl_margin_pct_label",
            positives=[
                "Margin % on Total Deal", "Margin % on Labor", "Margin % on PMO",
                "Margin % on Materials", "Net Margin (%)", "Net Margin %",
                "Gross Margin %", "Gross Margin (%)", "Margin %", "GM %",
                "Profit Margin %", "Margin Percent", "% Margin",
            ],
            negatives=[
                "Total Deal Revenue", "Total Labor Cost", "Total Deal Margin",
                "Net Profit ($)", "Net Profit", "Margin $", "Margin Dollars",
                "Customer", "PO Number", "Sales Rep", "Total PMO Revenue",
                # other percent-bearing labels that are NOT a margin %
                "% Complete", "% of Total", "Tax %", "Discount %",
                "Markup %", "% Allocation",
            ],
            threshold=0.60,
            lexical_fallback=_margin_pct_lexical,
        )
    return _PL_MARGIN_PCT_RULE


def _is_margin_pct_label(orig: str) -> tuple[bool, str | None]:
    """(is_margin_pct, explicit_category_or_None). Tries the strict "Margin % on
    <cat>" regex first (which also yields the category), then the semantic rule
    for the many category-less phrasings ("Net Margin (%)", "Margin %")."""
    m = _PL_MARGIN_PCT_RE.match(orig)
    if m:
        return True, m.group("cat")
    low = orig.lower()
    # structural prefilter: a margin-% label always carries a PERCENT cue. This is
    # what separates it from the bare "<Category> Margin" dollar row (which has no
    # %, matches the metric regex below, and must NOT be read as a percentage). It
    # also bounds the embedder to plausible candidates. The semantic rule then
    # disambiguates among percent-bearing labels ("Net Margin (%)" → yes,
    # "% Complete" / "Tax %" → no).
    if not _PCT_CUE_RE.search(low):
        return False, None
    try:
        if _pl_margin_pct_rule().fires(orig):
            return True, None
    except Exception:
        pass
    return False, None


def _looks_like_pct_or_error(value: Any) -> bool:
    """A cell value that reads as a margin-percent result: a fraction/percent, or
    a failed formula (#DIV/0!) — the value shape that sits in a margin-% slot."""
    if _excel_error_str(value) is not None:
        return True
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return -1.5 <= float(value) <= 1.5      # a fraction (0.2737), not a $ figure
    s = str(value).strip()
    return bool(s) and s.endswith("%")


def _norm_label(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


_DIGIT_RE = re.compile(r"\d")


def _is_bare_numeric(s: str) -> bool:
    """True if the cell is just a number (with optional $/%/,/./()/sign/space) —
    no letters. Such a string is never a real field LABEL, and a label-less bare
    number is a stray table cell, not a fact. Used to drop keyval junk like
    '70: 0.3684' (an unlabeled cost/sell/margin totals row paired as a fake
    'label: value') and lone '1.15' / '69.0' cells."""
    t = (s or "").strip()
    return bool(t) and re.fullmatch(r"[-+]?[\d.,$%()\s]+", t) is not None


def _is_substantive_annotation(text: str) -> bool:
    """A titleless free-text cell block worth keeping as a loose annotation, vs a
    decorative one-word caption. Substantive = carries a quantity ("920 hours",
    "Best Buy quoted 950 hours") OR reads as a real phrase (>=3 words / >=24
    chars). Must contain a letter, so a stray number alone (already captured by
    the numeric extractors) doesn't double-emit here."""
    t = (text or "").strip()
    if not t or not any(c.isalpha() for c in t):
        return False
    words = t.split()
    return bool(_DIGIT_RE.search(t)) or len(words) >= 3 or len(t) >= 24


# Offline fallback for the effort/hours-metric judgment. The embedding rule below
# generalises beyond these surface words ("Man-Days", "Crew Hours", "Build Time",
# "FTE-weeks", "Engineering Effort"); this net is what fires when the embedder is
# unreachable.
_HOURS_LABEL_RE = re.compile(r"\b(hour|hrs?\b|total|labor|effort|man[- ]?day|fte)", re.I)
_NUMLIKE_RE = re.compile(r"^[-$(]?\s?[\d,]+(?:\.\d+)?\s?%?\)?$")

_HOURS_METRIC_RULE = None


def _hours_metric_lexical(text: str) -> bool:
    return bool(_HOURS_LABEL_RE.search(text or ""))


def _hours_metric_rule():
    """SemanticRule: does this column header name a LABOR-EFFORT / HOURS metric
    (the thing a side estimate benchmarks against)? Embedding generalises past the
    fixed vocabulary — "Man-Days", "Crew Hours", "Engineering Effort", "Build
    Time", "FTE-weeks" all fire — with the regex as the offline-safe fallback."""
    global _HOURS_METRIC_RULE
    if _HOURS_METRIC_RULE is None:
        from app.core.semantic_rules import SemanticRule
        _HOURS_METRIC_RULE = SemanticRule(
            name="hours_effort_metric_label",
            positives=[
                "Lead Tech Hrs", "LV Tech Hrs", "Helper Hrs", "PM Hrs",
                "Total Base Hrs", "Labor Hours", "Total Hours", "Crew Hours",
                "Man-Days", "Man Hours", "Engineering Effort", "Build Time",
                "FTE-weeks", "Field Labor Hours", "Install Hours", "Total Effort",
            ],
            negatives=[
                "Quote Line Item", "Category", "Unit", "Qty", "Drops", "Price",
                "Total Price", "Material Cost", "Margin %", "Sell Rate",
                "Per Drop", "Availability", "Product Description",
            ],
            threshold=0.58,
            lexical_fallback=_hours_metric_lexical,
        )
    return _HOURS_METRIC_RULE


def _is_hours_metric_label(text: str) -> bool:
    """True when a (non-numeric, non-rate) cell label names an hours/effort metric.
    Structural prefilter (short, not a 'per <unit>' rate) gates the embedder so it
    only judges plausible header cells; the semantic rule then decides by meaning."""
    t = (text or "").strip()
    if not t or _NUMLIKE_RE.match(t) or len(t) > 40 or "per " in t.lower():
        return False
    try:
        return _hours_metric_rule().fires(t)
    except Exception:
        return _hours_metric_lexical(t)


def _sheet_hours_context(rows: list[list[Any]], limit: int = 6) -> str:
    """A compact digest of the sheet's own hour totals, e.g.
    'Lead Tech Hrs: 212.75, LV Tech Hrs: 941.06, Helper Hrs: 364.75, total 1148.81'.
    A side annotation ('Best Buy Quoted 920 hours') only means something against
    the estimate it benchmarks, so this travels with the annotation as context.

    Estimate totals are COLUMN-aligned: the hour labels are column headers and the
    totals sit in a numeric row far below, so pair the totals row to its header row
    by column index (not row adjacency)."""
    grid = [[("" if c is None else str(c).strip()) for c in r] for r in rows]

    def _num(s: str) -> float | None:
        if not _NUMLIKE_RE.match(s):
            return None
        try:
            return float(re.sub(r"[,$%()]", "", s))
        except ValueError:
            return None

    # header row = a row with >=2 hours/effort-metric LABELS (semantic, not a fixed
    # word list — "Man-Days"/"Crew Hours"/"Engineering Effort" all qualify)
    hdr_i = None
    for i, r in enumerate(grid):
        labels = [c for c in r if c and _is_hours_metric_label(c)]
        if len(labels) >= 2:
            hdr_i = i
            break
    if hdr_i is None:
        return ""
    hdr = grid[hdr_i]
    hour_cols = [ci for ci, h in enumerate(hdr) if h and _is_hours_metric_label(h)]
    # The TOTALS row is the column-sum row — same hour columns, largest magnitudes,
    # and (unlike a task row) no leading text label. Across all candidate rows pick
    # the one whose hour-column values sum largest; that's the estimate total.
    best_pairs: list[str] = []
    best_grand: float | None = None
    best_sum = 0.0
    for r in grid[hdr_i + 1:]:
        vals = {ci: _num(r[ci]) for ci in hour_cols if ci < len(r) and _num(r[ci]) is not None}
        if not vals:
            continue
        s = sum(abs(v) for v in vals.values())
        if s <= best_sum:
            continue
        pairs = [f"{hdr[ci].rstrip(':')}: {v:g}" for ci, v in vals.items() if abs(v) >= 1]
        if not pairs:
            continue
        row_nums = [v for v in (_num(c) for c in r) if v is not None]
        best_sum, best_pairs = s, pairs[:limit]
        best_grand = max(row_nums, key=abs) if row_nums else None
    if not best_pairs:
        return ""
    if best_grand is not None and not any(("%g" % best_grand) in p for p in best_pairs):
        best_pairs.append(f"total ~{best_grand:g}")
    return ", ".join(best_pairs)


def _is_label_cell(text: str) -> bool:
    """True when a cell reads like a label (a header/P&L term), so it is
    never mistaken for the *value* of the label to its left."""
    if text in _DEAL_HEADER_LABELS:
        return True
    if _PL_MARGIN_PCT_RE.match(text):
        return True
    m = _PL_METRIC_RE.match(text)
    return bool(m and m.group("metric").lower() in _PL_METRIC_WORDS)


def _coerce_pl_number(value: Any) -> float | None:
    """Parse a financial cell to a float. Returns None for blanks,
    Excel errors (#DIV/0!), and non-numeric text (TBD / No)."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s.startswith("#"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s2 = s.replace("(", "").replace(")", "").replace("$", "").replace(",", "").replace("%", "").strip()
    if not s2:
        return None
    try:
        n = float(s2)
    except ValueError:
        return None
    return -n if neg else n


def _norm_pl_category(raw: str) -> tuple[str, str]:
    """Return (key, display) for a P&L category label.

    ``display`` is the RAW label as written in the sheet (kept faithful — "Total
    Deal", "Total Labor", "Materials"). ``key`` is a normalized slug for grouping
    only: it drops a leading "Total" so the summary line and the detail block
    collapse to one category, but that normalization never touches what's shown."""
    disp = re.sub(r"\s+", " ", raw.strip())   # faithful label, e.g. "Total Deal"
    key = re.sub(r"^total\s+", "", disp.lower(), flags=re.I).strip() or disp.lower()
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    # 'deal' is the grand-total category — normalize a few synonyms (KEY only).
    if key in {"deal", "overall_deal", "project"}:
        key = "deal"
    return key, disp


def _right_neighbor_value(row: list[Any], idx: int) -> Any:
    """The first non-empty cell to the right of ``idx`` in ``row``."""
    for j in range(idx + 1, len(row)):
        v = row[j]
        if v is not None and str(v).strip():
            return v
    return None


def _looks_like_header_label(text: str) -> bool:
    """Structural test for a deal-header *label* cell, independent of the
    canonical vocabulary.

    A header label is a short, mostly-text caption ("PO Number", "Account
    Manager", "Site Contact") sitting to the left of its value. This lets
    the extractor capture deal-kit header fields that aren't in the
    canonical map, so non-standard fields are never silently dropped. P&L
    metric / margin lines are excluded — they belong to the P&L block, not
    the header."""
    t = re.sub(r"\s+", " ", str(text or "").strip()).rstrip(":").strip()
    if not t:
        return False
    if not re.search(r"[A-Za-z]", t):  # need a letter; pure numbers aren't labels
        return False
    if _PL_MARGIN_PCT_RE.match(t):
        return False
    m = _PL_METRIC_RE.match(t)
    if m and m.group("metric").lower() in _PL_METRIC_WORDS:
        return False
    # Labels are captions, not prose: keep them short.
    return len(t.split()) <= 6


# Excel error literals leak into cells as text (#DIV/0!, #REF!, #N/A, …).
# They are never a real field value — a formula failed — so they must
# never be captured as a deal-header field.
_EXCEL_ERROR_RE = re.compile(
    r"^#(?:DIV/0!|REF!|N/A|VALUE!|NAME\?|NULL!|NUM!|SPILL!|CALC!|GETTING_DATA|#+)$",
    re.I,
)


def _excel_error_str(value: Any) -> str | None:
    """The Excel error literal (``#DIV/0!``, ``#REF!``, …) the cell carries, else
    None. A failed formula is still FAITHFUL content — ``Margin % on PMO`` showing
    ``#DIV/0!`` tells the PM the metric is undefined because PMO revenue is $0 — so
    a P&L row whose value is an error is captured (flagged), never silently dropped."""
    if value is None or isinstance(value, bool):
        return None
    s = re.sub(r"\s+", " ", str(value).strip())
    return s if s and _EXCEL_ERROR_RE.match(s) else None


def _coerce_header_value(val: Any) -> str | None:
    """Normalize a header *value* cell to a string, or None when it isn't a
    usable value (blank, an Excel error literal, or itself a label/P&L term)."""
    if val is None:
        return None
    if hasattr(val, "isoformat"):
        try:
            iso = val.isoformat()
            # a date cell read as a midnight datetime -> show the date only
            # ('2026-05-27', not '2026-05-27T00:00:00').
            val = iso[:10] if "T00:00:00" in iso else iso
        except Exception:
            val = str(val)
    sval = re.sub(r"\s+", " ", str(val).strip())
    if not sval:
        return None
    if _EXCEL_ERROR_RE.match(sval):
        return None
    if _is_label_cell(_norm_label(sval)):
        return None
    return sval


def _first_nonblank_header_row(rows: list[list[Any]]) -> int | None:
    fallback = None
    for i, row in enumerate(rows[:20]):
        nonblank = [str(c).strip() for c in row if str(c or "").strip()]
        if len(nonblank) < 2:
            continue
        if fallback is None:
            fallback = i
        # A real header row carries >=2 LABEL-like cells (multi-char, not a bare
        # number). Skip column-group marker rows — a "A | F | H" banner of
        # single-letter section codes sitting one row above the true header
        # ("ID # | Material Description | OEM | ...") — so binding latches onto
        # the real column names, not "A:/H:".
        labelish = sum(
            1
            for c in nonblank
            if len(c) >= 3
            and not c.replace(",", "").replace(".", "").replace("$", "").lstrip("-").isdigit()
        )
        if labelish >= 2:
            return i
    return fallback


def _commercial_header_band(
    rows: list[list[Any]], money_cols: list[int]
) -> tuple[list[str], set[int], int]:
    """Find the header BAND for a commercial / rate sheet and compose ONE label
    per column.

    A simple catalog has a single header row; a rate card stacks several banner
    rows above the data (domain > technician level > min-charge tier) and keeps
    its dropdown SOURCE LISTS in the top rows, far from the matrix. Anchoring on
    "first non-blank row" then binds rate numbers to dropdown labels (the
    ``T&M: 0.85`` garbage). Instead: find the data block (the longest run of
    dense money rows), walk UP to collect the contiguous header rows directly
    above it, forward-fill the merged banner rows, and append the per-column
    header row beneath them.

    Returns ``(headers, header_row_indices, data_floor)`` where ``data_floor`` is
    the first row eligible to emit (rows above the header band — dropdown lists,
    base-rate scratch rows — are skipped, never mis-bound)."""
    def _isnum(s: str) -> bool:
        return bool(s) and s.replace(",", "").replace(".", "").replace("$", "").replace("%", "").lstrip("-").isdigit()

    def _simple() -> tuple[list[str], set[int], int]:
        hi = _first_nonblank_header_row(rows)
        hdr = (
            [("" if c is None else str(c).strip()) for c in rows[hi]]
            if hi is not None and 0 <= hi < len(rows) else []
        )
        return hdr, ({hi} if hi is not None else set()), (hi + 1 if hi is not None else 0)

    def _numeric_count(row: list[Any]) -> int:
        n = 0
        for v in row:
            if isinstance(v, bool):
                continue
            if isinstance(v, (int, float)):
                n += 1
            elif isinstance(v, str):
                s = v.strip().replace(",", "").replace("$", "").replace("%", "")
                if s and s.lstrip("-").replace(".", "", 1).isdigit():
                    n += 1
        return n

    # A data row in a rate matrix is dense with NUMBERS (a country's ~24 rates),
    # independent of which columns the money-column heuristic flagged — that
    # heuristic is fooled by the dropdown source lists on these sheets.
    dense = [_numeric_count(r) >= 5 for r in rows]
    runs, s = [], None
    for i, d in enumerate(dense):
        if d and s is None:
            s = i
        elif not d and s is not None:
            runs.append((s, i)); s = None
    if s is not None:
        runs.append((s, len(rows)))
    if not runs:
        return _simple()
    data_start = max(runs, key=lambda r: r[1] - r[0])[0]
    # Walk up from the data block, collecting contiguous header rows (>=2
    # multi-char labels), skipping lone category/banner cells, stopping at a
    # blank or a numeric/marker row.
    band: list[int] = []
    i = data_start - 1
    while i >= 0 and len(band) < 3:
        cells = [("" if c is None else str(c).strip()) for c in rows[i]]
        ne = [c for c in cells if c]
        if not ne:
            break
        if len(ne) == 1:
            i -= 1; continue
        if sum(1 for c in ne if len(c) >= 3 and not _isnum(c)) >= 2:
            band.append(i); i -= 1; continue
        break
    if not band:
        return _simple()
    band.sort()
    width = max((len(r) for r in rows), default=0)

    def _cells(bi: int) -> list[str]:
        return [("" if (ci >= len(rows[bi]) or rows[bi][ci] is None) else str(rows[bi][ci]).strip())
                for ci in range(width)]

    def _ff(cs: list[str]) -> list[str]:  # forward-fill merged banner cells
        out, last = [], ""
        for c in cs:
            if c:
                last = c
            out.append(last)
        return out

    banners = [_ff(_cells(bi)) for bi in band[:-1]]   # domain / level banners (merged)
    base = _cells(band[-1])                            # per-column header row, as-is
    headers: list[str] = []
    for ci in range(width):
        parts = [b[ci] for b in banners if b[ci]]
        if base[ci]:
            parts.append(base[ci])
        headers.append(" ".join(parts))
    # Emit from just below the header band — keeps any category divider that
    # sits between the header and the first data row (e.g. "CAT6…").
    return headers, set(band), band[-1] + 1


def _cell_to_text_op(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            iso = value.isoformat()
            return iso[:10] if "T00:00:00" in iso else iso  # date-only at midnight
        except Exception:
            pass
    return str(value).strip()

# Canonical column role -> normalized header tokens (lowercase, punctuation-stripped variants).
HEADER_ALIASES: dict[str, set[str]] = {
    "project": {"project", "job", "job number", "project number"},
    "site": {"site", "facility", "building", "school", "campus", "store", "venue"},
    "facility": {"facility name", "facility"},
    "building": {"building", "bldg", "structure"},
    "floor": {"floor", "level", "fl"},
    "wing": {"wing", "sector"},
    "area": {"area", "region"},
    "zone": {"zone", "district"},
    "room": {"room", "room #", "room number", "space"},
    "location": {"location", "position", "place", "mounting location", "area name"},
    "plate_id": {"plate id", "plate", "outlet id", "drop id", "jack id", "cable id", "plate #"},
    "outlet_id": {"outlet id", "outlet"},
    "drop_id": {"drop id", "drop"},
    "mdf": {"mdf", "mdf id", "main distribution frame"},
    "idf": {"idf", "idf id", "intermediate distribution frame"},
    "closet": {"closet", "telecom closet"},
    "rack": {"rack", "rack id"},
    "device": {"device", "asset", "equipment", "camera", "ap", "reader", "hostname"},
    "device_type": {"device type", "type", "equipment type"},
    "item": {"item", "line item", "line"},
    "description": {"description", "desc", "details", "summary"},
    "quantity": {"qty", "qty.", "quantity", "count", "#", "no", "units", "total qty", "# drops"},
    "count": {"count", "cnt"},
    "uom": {"uom", "unit", "units", "ea", "each"},
    "material": {"material", "wire", "cable"},
    "material_spec": {"material spec", "material / spec", "spec", "cable spec"},
    "cable_category": {"cable category", "cable cat"},
    "cable_type": {"cable type", "cable"},
    "shielding": {"shielding", "utp", "stp"},
    "jacket_rating": {"jacket", "plenum", "riser"},
    "connector": {"connector", "jack type"},
    "termination": {"termination", "terminations"},
    "patch_panel": {"patch panel", "panel"},
    "faceplate": {"faceplate", "wall plate", "plate type"},
    "scope": {"scope", "work type", "work package"},
    "included": {"included", "included?", "in scope?", "in scope", "base bid"},
    "excluded": {"excluded", "excluded?", "out of scope"},
    "access": {"access", "access window", "hours", "site access", "work window"},
    "access_window": {"access window", "hours"},
    "lift": {"lift", "lift required", "elevator"},
    "ceiling_access": {"ceiling access", "ceiling"},
    "after_hours": {"after hours", "after-hours", "nights", "weekends"},
    "escort": {"escort", "escort required"},
    "badge": {"badge", "badge access", "badge required", "mdf badge"},
    "customer_responsibility": {"customer responsibility", "owner responsibility", "customer provides"},
    "vendor_responsibility": {"vendor responsibility", "contractor provides"},
    "notes": {"notes", "comments", "clarifications", "assumptions", "remarks"},
    "open_question": {"open question", "question", "confirm"},
    "test_standard": {"test standard", "test standard required"},
    "certification": {"certification", "testing", "test standard", "certify", "tester export", "fluke"},
    "labeling": {"label", "labeling", "label standard"},
    "as_built": {"as built", "as-built", "redlines"},
    "status": {"status", "state"},
}

RowKind = Literal[
    "blank",
    "title",
    "header",
    "section_header",
    "line_item",
    "total",
    "subtotal",
    "grand_total",
    "note",
    "malformed",
]


@dataclass
class SheetParseModel:
    header_idx: int
    header_map: dict[str, int]
    wide_qty_columns: list[dict[str, Any]]
    header_mode: str
    diagnostics: list[str] = field(default_factory=list)


def _header_cell_tokens(cell: Any) -> set[str]:
    raw = str(cell or "").strip()
    if not raw:
        return set()
    lowered = raw.lower()
    tokens: set[str] = {
        normalize_text(raw).strip(".:?"),
        normalize_text(raw.replace("/", " ")).strip(".:?"),
        normalize_text(raw.replace("/", " ").replace("?", "")).strip(".:?"),
        re.sub(r"\s+", " ", lowered).strip(),
    }
    for part in re.split(r"[/|]", raw):
        t = normalize_text(part).strip(".:?! ")
        if t:
            tokens.add(t)
    return {t for t in tokens if t}


def _map_canonical_header(cell: Any) -> str | None:
    for key, aliases in HEADER_ALIASES.items():
        if _header_cell_tokens(cell) & aliases:
            return key
    return None


def _wide_quantity_header_meta(cell: Any) -> dict[str, Any] | None:
    """If this column header is a material/connector quantity column (wide schedule), return metadata."""
    blob = " ".join(sorted(_header_cell_tokens(cell))).lower()
    if not blob.strip():
        return None
    if re.search(r"\brj[-\s]?45\b", blob) or re.search(r"\bdata jack\b", blob) or blob == "jack count":
        return {
            "item": "RJ45",
            "normalized_item": "rj45",
            "item_kind": "termination",
            "material_family": "connector",
            "cable_category": None,
            "shielding": None,
            "entity_hint": ("connector", "rj45"),
        }
    if re.search(r"\bcat6a\b|\bcat\s*6a\b|category\s*6a", blob):
        return {
            "item": "Cat6A",
            "normalized_item": "cat6a",
            "item_kind": "cable_drop",
            "material_family": "cable",
            "cable_category": "cat6a",
            "shielding": None,
            "entity_hint": ("cable", "cat6a"),
        }
    if "fiber" in blob or "strand" in blob:
        return {
            "item": "Fiber",
            "normalized_item": "fiber",
            "item_kind": "cable_drop",
            "material_family": "fiber",
            "cable_category": None,
            "shielding": None,
            "entity_hint": ("cable", "fiber"),
        }
    if re.search(r"\bcat6\b", blob) and ("utp" in blob or "unshielded" in blob or "non.?shielded" in blob):
        return {
            "item": "Cat6 UTP",
            "normalized_item": "cat6_utp",
            "item_kind": "cable_drop",
            "material_family": "cable",
            "cable_category": "cat6",
            "shielding": "unshielded",
            "entity_hint": ("material", "cat6_utp"),
        }
    if re.search(r"\bcat6\b", blob) and ("stp" in blob or "shielded" in blob):
        return {
            "item": "Cat6 STP",
            "normalized_item": "cat6_stp",
            "item_kind": "cable_drop",
            "material_family": "cable",
            "cable_category": "cat6",
            "shielding": "shielded",
            "entity_hint": ("cable", "cat6_stp"),
        }
    if re.search(r"\bcat6\b", blob) and "utp" not in blob and "stp" not in blob and "shield" not in blob:
        return {
            "item": "Cat6",
            "normalized_item": "cat6",
            "item_kind": "cable_drop",
            "material_family": "cable",
            "cable_category": "cat6",
            "shielding": None,
            "entity_hint": ("cable", "cat6"),
        }
    return None


def _header_map_from_row(row: list[Any]) -> tuple[dict[str, int], list[dict[str, Any]]]:
    header_map: dict[str, int] = {}
    wide: list[dict[str, Any]] = []
    seen_wide_cols: set[int] = set()
    for col_idx, cell in enumerate(row):
        canon = _map_canonical_header(cell)
        if canon and canon not in header_map:
            header_map[canon] = col_idx
        meta = _wide_quantity_header_meta(cell)
        if meta and col_idx not in seen_wide_cols:
            seen_wide_cols.add(col_idx)
            wide.append({"col_idx": col_idx, "header_raw": str(cell or "").strip(), **meta})
    return header_map, wide


def _merge_header_cells(top: Any, bottom: Any) -> str:
    a = str(top or "").strip()
    b = str(bottom or "").strip()
    if a and b:
        return f"{a} {b}".strip()
    return a or b


def _header_map_from_two_rows(row_top: list[Any], row_bot: list[Any]) -> tuple[dict[str, int], list[dict[str, Any]]]:
    width = max(len(row_top), len(row_bot))
    merged: list[str] = []
    for col in range(width):
        top = row_top[col] if col < len(row_top) else None
        bot = row_bot[col] if col < len(row_bot) else None
        merged.append(_merge_header_cells(top, bot))
    return _header_map_from_row(merged)


def _sheet_qualifies(header_map: dict[str, int], wide_qty_columns: list[dict[str, Any]]) -> bool:
    if wide_qty_columns:
        return any(header_map.get(k) is not None for k in ("plate_id", "location", "room", "site", "description", "device"))
    entity_any = any(
        header_map.get(k) is not None
        for k in ("site", "device", "location", "room", "plate_id", "building", "floor", "project")
    )
    if header_map.get("quantity") is not None and entity_any:
        return True
    if entity_any and len(header_map) >= 2:
        return True
    return False


def _detect_header(rows: list[list[Any]], scan_limit: int = 45) -> SheetParseModel:
    diagnostics: list[str] = []
    best_idx: int | None = None
    best_map: dict[str, int] = {}
    best_wide: list[dict[str, Any]] = []
    best_score = -1.0
    best_mode = "single"

    limit = min(scan_limit, len(rows))
    for idx in range(limit):
        row = rows[idx]
        hm, wq = _header_map_from_row(row)
        score = len(hm) + 0.45 * len(wq)
        if _sheet_qualifies(hm, wq) and score > best_score:
            best_score = score
            best_idx = idx
            best_map = hm
            best_wide = list(wq)
            best_mode = "single"
        if idx + 1 < len(rows):
            hm2, wq2 = _header_map_from_two_rows(row, rows[idx + 1])
            score2 = len(hm2) + 0.45 * len(wq2) + 0.05
            if _sheet_qualifies(hm2, wq2) and score2 > best_score:
                best_score = score2
                best_idx = idx
                best_map = hm2
                best_wide = list(wq2)
                best_mode = "pair"

    if best_idx is None:
        diagnostics.append("no_header_found")
        return SheetParseModel(-1, {}, [], "none", diagnostics)

    diagnostics.append(f"header_row={best_idx + 1} mode={best_mode} keys={sorted(best_map.keys())} wide_qty_cols={len(best_wide)}")
    if not best_wide and not best_map.get("quantity"):
        diagnostics.append("no_quantity_bearing_column")
    elif best_wide:
        diagnostics.append(f"domain_wide_quantity_columns:{','.join(w['item'] for w in best_wide)}")
    return SheetParseModel(best_idx, best_map, best_wide, best_mode, diagnostics)


def _is_blank_row(row: list[Any]) -> bool:
    return all(str(c or "").strip() == "" for c in row)


# Sheet names that signal "this tab is human instructions, not data".
# Compared after lowercasing + stripping; substring match so workbooks
# titled e.g. ``Instructions (Read First)`` still hit. Universal across
# all xlsx workbooks; not domain-specific.
_INSTRUCTIONAL_SHEET_NAME_TOKENS: frozenset[str] = frozenset({
    "instruction",
    "instructions",
    "readme",
    "read me",
    "read_me",
    "cover",
    "cover sheet",
    "coversheet",
    "about",
    "overview",
    "guide",
    "how to",
    "how-to",
    "table of contents",
    "toc",
    "legend",
    "key",
    "intro",
    "introduction",
    "title page",
})


def _looks_instructional_sheet(sheet_name: str) -> bool:
    """True when the sheet name signals a non-data instructional tab."""
    name = (sheet_name or "").strip().lower()
    if not name:
        return False
    return any(token in name for token in _INSTRUCTIONAL_SHEET_NAME_TOKENS)


def _has_tabular_shape(rows: list[list[Any]]) -> bool:
    """Cheap structural check: does the sheet *look* like a real table?

    True when there are at least 2 non-blank rows AND the widest row has
    >= 2 non-empty cells. False for pure prose tabs (one row of
    instructions, blank rows, single-column dumps).
    """
    non_blank = [r for r in rows if not _is_blank_row(r)]
    if len(non_blank) < 2:
        return False
    widest = max(
        (sum(1 for c in r if str(c or "").strip()) for r in non_blank),
        default=0,
    )
    return widest >= 2


def _generic_table_from_rows(
    rows: list[list[Any]],
) -> tuple[list[str], list[dict[str, Any]]] | None:
    """Build a full table from a tabular sheet whose header the canonical
    detector could not recognise.

    The canonical ``_detect_header`` only fires for sheets whose columns
    match the device/quantity vocabulary it knows. Bill-of-materials, rate
    cards, financial P&Ls and other commercial tables use arbitrary
    headers, so they previously fell through to a 25-row truncated prose
    dump — silently losing most rows and all structure. This recovers
    them generically: pick the widest of the leading rows as the header,
    keep every non-blank data row (no truncation), and never invents
    vocabulary. Returns ``None`` when the sheet is not genuinely tabular.
    """
    if not _has_tabular_shape(rows):
        return None

    def _filled(row: list[Any]) -> int:
        return sum(1 for c in (row or []) if str(c or "").strip())

    # Width = the widest non-blank row's column span (drives column count).
    width = max((len(r or []) for r in rows if not _is_blank_row(r)), default=0)
    if width < 2:
        return None

    # Header = the first of the leading non-blank rows that is the most
    # populated (titles/notes above the real header have fewer cells).
    header_idx = -1
    best_filled = -1
    seen = 0
    for idx, row in enumerate(rows):
        if _is_blank_row(row):
            continue
        seen += 1
        f = _filled(row)
        if f > best_filled:
            best_filled = f
            header_idx = idx
        if seen >= 8:  # only scan the leading band for a header
            break
    if header_idx < 0:
        return None

    header_row = rows[header_idx] or []
    columns: list[str] = []
    used: dict[str, int] = {}
    for i in range(width):
        raw = str(header_row[i]).strip() if i < len(header_row) and header_row[i] is not None else ""
        name = raw or f"col_{i + 1}"
        if name in used:
            used[name] += 1
            name = f"{name}_{used[name]}"
        else:
            used[name] = 1
        columns.append(name)

    table_rows: list[dict[str, Any]] = []
    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx] or []
        if _is_blank_row(row):
            continue
        cells: dict[str, Any] = {}
        for col_idx, col_name in enumerate(columns):
            value = row[col_idx] if col_idx < len(row) else ""
            cells[col_name] = "" if value is None else str(value).strip()
        table_rows.append(cells)

    if not table_rows:
        return None
    return columns, table_rows


def _label_text_for_row(row: list[Any], label_col_indices: list[int]) -> str:
    parts: list[str] = []
    for idx in label_col_indices:
        if idx < len(row):
            parts.append(str(row[idx] or "").strip())
    return " ".join(p for p in parts if p).strip()


def _label_column_indices(header_map: dict[str, int]) -> list[int]:
    for key in ("plate_id", "location", "description", "item", "site", "room", "drop_id", "outlet_id"):
        if key in header_map:
            return [header_map[key]]
    return [0]


def _is_total_label(text: str) -> bool:
    t = normalize_text(text).strip()
    if not t:
        return False
    return bool(re.match(r"^(totals?|subtotal|grand\s*total)\b", t, re.I))


def _unlabeled_sum_row(rows: list[list[Any]], ri: int, lookback: int = 15) -> bool:
    """UNIVERSAL, arithmetic total-detection: a row that carries numbers but NO
    text label, whose value in each column equals the SUM of the contiguous
    numeric rows immediately above it, IS a totals row. The arithmetic proves it —
    no keyword, no guess, no invention — so a bare '14175 | 12375' summing the
    rows above is recognized as a Total on ANY sheet, not just one we hand-tagged."""
    if not (0 <= ri < len(rows)):
        return False
    row = rows[ri]
    # must have NO text label (a labeled total is handled by _is_total_label)
    if any(isinstance(v, str) and v.strip() for v in row):
        return False
    nums = {
        ci: float(v)
        for ci, v in enumerate(row)
        if isinstance(v, (int, float)) and not isinstance(v, bool) and abs(v) >= 1
    }
    if not nums:
        return False
    matched = 0
    for ci, val in nums.items():
        s, n = 0.0, 0
        for rj in range(ri - 1, max(-1, ri - 1 - lookback), -1):
            cell = rows[rj][ci] if ci < len(rows[rj]) else None
            if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                s += float(cell)
                n += 1
            elif _is_blank_row(rows[rj]):
                if n:
                    break
        # need >=2 summed rows and a tight match (allow float rounding)
        if n >= 2 and abs(s - val) <= max(0.5, abs(val) * 0.001):
            matched += 1
    return matched == len(nums)  # EVERY numeric cell must be a column sum


def _row_kind(
    row: list[Any],
    header_map: dict[str, int],
    label_indices: list[int],
) -> RowKind:
    if _is_blank_row(row):
        return "blank"
    label = _label_text_for_row(row, label_indices)
    low = label.lower()
    if _is_total_label(label):
        if low.startswith("grand"):
            return "grand_total"
        if low.startswith("subtotal") or low.startswith("sub total"):
            return "subtotal"
        return "total"
    if label.endswith(":") and len(label) < 56:
        return "section_header"
    if len(label) > 80 and sum(1 for c in row if str(c or "").strip()) <= 2:
        return "title"
    if label and not any(str(row[i] or "").strip() for i in range(len(row)) if i not in label_indices):
        return "note"
    return "line_item"


def _parse_schedule_quantity_cell(raw_val: Any) -> dict[str, Any]:
    raw = "" if raw_val is None else str(raw_val).strip()
    out: dict[str, Any] = {
        "quantity_raw": raw,
        "quantity": None,
        "quantity_status": "missing",
        "quantity_min": None,
        "quantity_max": None,
        "uom": None,
        "uncertain": True,
        "review_flags": [],
    }
    if raw == "":
        out["quantity_status"] = "missing"
        return out
    low = normalize_text(raw).replace(",", "")
    if low in {"n/a", "na", "tbd", "pending"}:
        out["quantity_status"] = "tbd" if "tbd" in low or "pending" in low else "not_applicable"
        return out
    if "allowance" in low or low == "lot" or "lot" in low:
        out["quantity_status"] = "allowance"
        out["uncertain"] = True
        return out
    if "included" in low and not re.search(r"\d", raw):
        out["quantity_status"] = "included"
        out["uncertain"] = False
        return out
    mr = re.match(r"^\s*(\d[\d,]*)\s*[-–]\s*(\d[\d,]*)\s*$", raw.replace(",", ""))
    if mr:
        lo = int(mr.group(1))
        hi = int(mr.group(2))
        out["quantity_min"] = lo
        out["quantity_max"] = hi
        out["quantity_status"] = "range"
        out["review_flags"].append("xlsx_parser:range_quantity")
        return out
    mq = re.match(r"^\s*(-?\d[\d,]*(?:\.\d+)?)\s*([a-z%]{0,8})?\s*$", low.replace(",", ""))
    if mq:
        q = float(mq.group(1))
        if q.is_integer():
            q = int(q)
        out["quantity"] = q
        out["uom"] = mq.group(2) or None
        out["quantity_status"] = "zero" if q == 0 else "known"
        out["uncertain"] = False
        return out
    legacy = parse_quantity(raw)
    out.update(
        {
            "quantity": legacy.get("quantity"),
            "uom": legacy.get("unit"),
            "uncertain": bool(legacy.get("uncertain", True)),
        }
    )
    out["quantity_status"] = "ambiguous" if legacy.get("uncertain") else ("known" if legacy.get("quantity") is not None else "ambiguous")
    if out["quantity_status"] == "ambiguous":
        out["review_flags"].append("xlsx_parser:ambiguous_quantity")
    return out


def _row_text_blob(row: list[Any], header_map: dict[str, int]) -> str:
    parts: list[str] = []
    for key in ("notes", "description", "scope", "access", "location", "item"):
        idx = header_map.get(key)
        if idx is not None and idx < len(row):
            parts.append(str(row[idx] or ""))
    parts.extend(str(c or "") for c in row)
    return normalize_text(" ".join(parts)).lower()


def _emit_scope_constraint_atoms(
    row_blob: str,
    site: str,
    device: str,
    floor: str,
    room: str,
    location: str,
    append_atom: Any,
    row_confidence: float,
    *,
    context_columns: dict[str, str],
) -> None:
    cols = context_columns

    # These are per-row KEYWORD sub-facts: a word in the row ("after-hours",
    # "badge", "access") mints a thin, context-less "X access constraint" atom
    # that just restates something already on the row. In one-atom-per-row mode
    # they fold onto the row (same rule as site_allocation / column-entities);
    # the brittle keyword typing they encode is the head's job, not the parser's.
    if os.environ.get("SOWSMITH_DROP_DERIVED_SUBATOMS") == "1":
        return

    def ap(
        atom_type: AtomType,
        raw_text: str,
        value: dict[str, Any],
        confidence: float,
        *,
        review_status: ReviewStatus = ReviewStatus.auto_accepted,
        review_flags: list[str] | None = None,
    ) -> None:
        append_atom(
            atom_type,
            raw_text,
            value,
            confidence,
            review_status=review_status,
            review_flags=review_flags,
            extra_columns=cols,
        )

    if re.search(r"\b(after[-\s]?hours|nights only|weekends only)\b", row_blob):
        ap(
            AtomType.constraint,
            "After-hours access constraint",
            {"constraint_type": "after_hours", "site": site, "device": device, "floor": floor, "room": room, "location": location},
            row_confidence * 0.9,
        )
    if re.search(r"\b(lift required|elevator|customer provides lift|customer provide lift)\b", row_blob):
        if "customer" in row_blob and "lift" in row_blob:
            ap(
                AtomType.action_item,
                "Customer lift responsibility",
                {"action": "customer_provides_lift", "site": site, "device": device, "location": location},
                row_confidence * 0.85,
            )
        ap(
            AtomType.constraint,
            "Lift access constraint",
            {"constraint_type": "lift", "site": site, "device": device, "location": location},
            row_confidence * 0.88,
        )
    if re.search(r"\b(badge|escort|ceiling access)\b", row_blob):
        ap(
            AtomType.constraint,
            "Site access constraint",
            {"constraint_type": "access", "detail": row_blob[:200], "site": site, "location": location},
            row_confidence * 0.85,
        )
    if re.search(r"\b(confirm|unknown|tbd)\b.*\b(badge|mdf|access)\b|\b(badge|mdf|access)\b.*\b(confirm|unknown|tbd)\b", row_blob):
        ap(
            AtomType.open_question,
            "Access confirmation open question",
            {"topic": "badge_or_access", "site": site, "location": location},
            row_confidence * 0.8,
            review_status=ReviewStatus.needs_review,
        )
    if re.search(r"\b(certification required|certify|test standard)\b", row_blob):
        ap(
            AtomType.constraint,
            "Certification requirement",
            {"constraint_type": "certification", "site": site, "location": location},
            row_confidence * 0.88,
        )
    if re.search(r"\blabel(ing)?\b.*\btbd\b|\btbd\b.*\blabel", row_blob):
        ap(
            AtomType.open_question,
            "Labeling standard TBD",
            {"topic": "labeling", "site": site, "location": location},
            row_confidence * 0.78,
            review_status=ReviewStatus.needs_review,
        )
    if re.search(r"\b(excluded|removed|deleted|not included|out of scope)\b", row_blob):
        ap(
            AtomType.exclusion,
            "Scope exclusion from schedule",
            {"exclusion_hint": row_blob[:240], "site": site, "location": location},
            row_confidence * 0.82,
        )


class XlsxParser(BaseParser):
    parser_name = parser_name
    parser_version = parser_version
    capability = ParserCapability(
        parser_name=parser_name,
        parser_version=parser_version,
        supported_extensions=[".xlsx", ".csv"],
        supported_artifact_types=[ArtifactType.xlsx, ArtifactType.csv],
        emitted_atom_types=[
            AtomType.entity,
            AtomType.quantity,
            AtomType.scope_item,
            AtomType.constraint,
            AtomType.exclusion,
            AtomType.open_question,
            AtomType.action_item,
        ],
        supported_domain_packs=["*"],
        requires_binary=False,
        supports_source_replay=True,
    )

    def match(self, path: Path, sample_text: str | None, domain_pack: DomainPack | None) -> ParserMatch:
        del sample_text, domain_pack
        suffix = path.suffix.lower()
        confidence = 0.0
        reasons: list[str] = []
        if suffix in {".xlsx", ".csv"}:
            from app.parsers.spreadsheet_route_signals import (
                path_roster_schedule_hint,
                sniff_operations_workbook_strength,
                sniff_xlsx_roster_schedule_strength,
            )

            confidence = 0.58
            reasons.append(f"spreadsheet_extension:{suffix}")

            # PR1 (post-v3 review) — multi-sheet operations workbook
            # detection. Wins decisively over QuoteParser when the
            # workbook has asset / site / port / circuit / risk /
            # cutover sheets (even if it ALSO has a BOM sheet).
            if suffix == ".xlsx":
                ops_score, ops_reasons = sniff_operations_workbook_strength(path)
                reasons.extend(ops_reasons)
                if ops_score >= 0.55:
                    return ParserMatch(
                        parser_name=self.parser_name,
                        confidence=0.97,
                        reasons=reasons + ["xlsx_match:operations_workbook"],
                        artifact_type=ArtifactType.xlsx,
                    )

            if path_roster_schedule_hint(path):
                confidence += 0.14
                reasons.append("xlsx_match:path_roster_schedule_token")
            try:
                xscore, _xr = sniff_xlsx_roster_schedule_strength(path)
                confidence = min(0.92, confidence + 0.22 * xscore)
                reasons.append(f"xlsx_match:schedule_strength={xscore:.2f}")
            except Exception:  # noqa: BLE001
                reasons.append("xlsx_match:schedule_sniff_failed")
        return ParserMatch(
            parser_name=self.parser_name,
            confidence=confidence,
            reasons=reasons,
            artifact_type=ArtifactType.xlsx if suffix == ".xlsx" else ArtifactType.csv,
        )

    def parse(self, artifact_path: Path) -> list[EvidenceAtom]:
        artifact_id = stable_id("art", str(artifact_path))
        return self.parse_artifact(
            project_id="unknown_project",
            artifact_id=artifact_id,
            path=artifact_path,
        )

    def segment_artifact(self, project_id: str, artifact_id: str, path: Path) -> list[ArtifactSegment]:
        return segment_xlsx(project_id=project_id, artifact_id=artifact_id, path=path, parser_version=self.parser_version)

    def parse_artifact(
        self,
        project_id: str,
        artifact_id: str,
        path: Path,
        domain_pack: DomainPack | None = None,
    ) -> list[EvidenceAtom]:
        """Back-compat wrapper that returns a flat list of atoms.

        Prefer :meth:`parse_artifact_full` — that surfaces the
        ``structured.json`` / ``structured.md`` derived files to the
        compiler so the OrbitBrief envelope can render this workbook
        with the same fidelity it gets for PDFs.
        """
        return self.parse_artifact_full(
            project_id=project_id,
            artifact_id=artifact_id,
            path=path,
            domain_pack=domain_pack,
        ).atoms

    def parse_artifact_full(
        self,
        project_id: str,
        artifact_id: str,
        path: Path,
        domain_pack: DomainPack | None = None,
    ) -> ParserOutput:
        del domain_pack
        suffix = path.suffix.lower()
        if suffix == ".csv":
            atoms, sheets, parse_error = self._parse_csv(
                project_id=project_id, artifact_id=artifact_id, path=path
            )
            schema = STRUCTURED_SCHEMA_CSV
            artifact_type = ArtifactType.csv
        else:
            atoms, sheets, parse_error = self._parse_xlsx(
                project_id=project_id, artifact_id=artifact_id, path=path
            )
            schema = STRUCTURED_SCHEMA_XLSX
            artifact_type = ArtifactType.xlsx
            # COVERAGE BACKSTOP (parse-coverage, no silent whole-file loss):
            # a VALID workbook whose every sheet routed to DROP/empty — e.g. an
            # unfilled RFP price/SLA response template (price column blank, so the
            # sheet_classifier drops it) — would otherwise vanish entirely, losing
            # the whole priced-scope catalog. If the read SUCCEEDED but produced
            # ZERO atoms, fall back to generic per-row extraction over any
            # data-bearing sheet so a populated file can never be silently dropped.
            # Purely additive: fires only when output would be empty, so it cannot
            # regress any deal that already extracts. Instructional/cover sheets
            # are still skipped by _emit_generic_rows' own guard.
            if not atoms and not parse_error:
                backstop: list[EvidenceAtom] = []
                for sh in sheets:
                    rws = sh.get("rows") or []
                    # Never undo an intentional router decision: a sheet the
                    # classifier sent to COMMERCIAL or DROP was processed on
                    # purpose (rate card / catalog / helper). Recovering it as
                    # generic scope_item atoms was the #010063 flood.
                    cl = classify_sheet(sh.get("name") or "", rws)
                    if cl.destination is not SheetDestination.SCOPE:
                        continue
                    nonempty = [r for r in rws if any(str(c or "").strip() for c in r)]
                    if len(nonempty) >= 3:
                        backstop.extend(self._emit_generic_rows(
                            project_id=project_id, artifact_id=artifact_id,
                            artifact_type=ArtifactType.xlsx, filename=path.name,
                            sheet_name=sh.get("name") or "", rows=rws,
                        ))
                if backstop:
                    atoms = backstop
                    self._coverage_backstop_note = (
                        f"INFO: coverage backstop recovered {len(backstop)} row(s) "
                        f"from {path.name} (every sheet had routed to drop/empty — "
                        f"e.g. an unfilled RFP price/SLA template)"
                    )

        structured_doc = self._build_structured_doc(
            schema=schema,
            artifact_type=artifact_type,
            filename=path.name,
            sheets=sheets,
            parse_error=parse_error,
        )
        stamp_section_and_block_ids(structured_doc, artifact_seed=artifact_id)
        warnings: list[str] = []
        if parse_error:
            warnings.append(f"{ARTIFACT_PARSE_ERROR_PREFIX}{artifact_id}:{parse_error}")
        _bs_note = getattr(self, "_coverage_backstop_note", "")
        if _bs_note:
            warnings.append(_bs_note)
            self._coverage_backstop_note = ""
        # Mark embedded charts / images / drawings / OLE objects in .xlsx so an
        # embedded diagram or logo-as-data can't silently vanish. (CSV has no
        # zip container, so this is a no-op there.)
        if suffix != ".csv":
            atoms = list(atoms) + emit_zip_binary_markers(
                path=path,
                project_id=project_id,
                artifact_id=artifact_id,
                filename=path.name,
                artifact_type=ArtifactType.xlsx,
                parser_version=self.parser_version,
            )
        return ParserOutput(
            atoms=atoms,
            derived_files=derived_files_for(artifact_path=path, structured_doc=structured_doc),
            warnings=warnings,
        )

    @staticmethod
    def _tabular_read_error_code(exc: BaseException, *, tabular: Literal["xlsx", "csv"]) -> str:
        name = type(exc).__name__
        msg = str(exc).lower()
        if tabular == "xlsx" and (name == "BadZipFile" or "zip" in msg or "central directory" in msg):
            return "corrupt_xlsx"
        if tabular == "csv":
            return "corrupt_csv"
        return f"{tabular}_read_error"

    @staticmethod
    def _hidden_dims(path: Path) -> dict[str, tuple[set[int], set[int]]]:
        """Map sheet title -> (hidden 0-based column indices, hidden 0-based row
        indices). Author-hidden columns/rows in a deal spreadsheet are helper /
        formula scaffolding — multipliers, lookup helpers, intermediate math —
        not deal content (e.g. a Gantt's "Country Multiplier", "Sell Helper",
        "Cost Helper"). They must not become atoms. Read from a non-read_only
        load because read_only worksheets don't expose column/row dimensions."""
        out: dict[str, tuple[set[int], set[int]]] = {}
        try:
            wb = load_workbook(path, read_only=False, data_only=True)
        except Exception:
            return out
        try:
            from openpyxl.utils import column_index_from_string

            for ws in wb.worksheets:
                hc = {
                    column_index_from_string(c) - 1
                    for c, d in ws.column_dimensions.items()
                    if d.hidden
                }
                hr = {i - 1 for i, d in ws.row_dimensions.items() if d.hidden}
                out[ws.title] = (hc, hr)
        except Exception:
            return out
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return out

    @staticmethod
    def _flag_hidden_source_atoms(
        atoms: list[EvidenceAtom], rows: list[list[Any]], hidden_rows: set[int]
    ) -> list[EvidenceAtom]:
        """Tag atoms whose row was AUTHOR-HIDDEN in the sheet.

        Hidden rows are still captured (no silent drops), but in the atom list a
        collapsed / 0-hour / scaffolding row looks identical to a live one — so a
        reviewer can't tell whether it parsed right. Match each atom's text to a
        hidden row's distinctive (longest) cell and stamp it with a visible
        '[hidden row in source sheet]' marker + the xlsx_parser:hidden_in_source
        flag. Content-match because the block detector discards row indices."""
        if not hidden_rows or not atoms:
            return atoms
        sigs: list[str] = []
        for i in hidden_rows:
            if 0 <= i < len(rows):
                texts = [str(c).strip() for c in rows[i] if c is not None and str(c).strip()]
                longest = max(texts, key=len) if texts else ""
                if len(longest) >= 8:                     # distinctive enough to match
                    sigs.append(longest.lower())
        if not sigs:
            return atoms
        out: list[EvidenceAtom] = []
        for a in atoms:
            rt = a.raw_text or ""
            if any(s in rt.lower() for s in sigs) and "[hidden row in source sheet]" not in rt:
                flags = list(a.review_flags or [])
                if "xlsx_parser:hidden_in_source" not in flags:
                    flags.append("xlsx_parser:hidden_in_source")
                try:
                    a = a.model_copy(update={
                        "raw_text": f"{rt}  [hidden row in source sheet]",
                        "review_flags": flags,
                    })
                except Exception:
                    pass
            out.append(a)
        return out

    @staticmethod
    def _sheet_styles(path: Path) -> dict[str, list[list[tuple[str | None, bool]]]]:
        """Map sheet title -> per-cell ``(fill_rgb_or_None, bold)`` grid, aligned
        row/col to the values grid. Cell STYLE is structure the author used to
        group and title regions, and a B2B deal kit leans on it heavily: a dark
        banner fill marks a section header; a pastel fill marks a highlighted
        box of related rows. The block detector reads this to title untitled
        boxes and keep same-fill rows together. Best-effort: any failure yields
        an empty map and the detector falls back to geometry alone."""
        out: dict[str, list[list[tuple[str | None, bool]]]] = {}
        try:
            wb = load_workbook(path, read_only=True)  # styles need no data_only
        except Exception:
            return out
        try:
            for ws in wb.worksheets:
                try:
                    ws.reset_dimensions()
                except Exception:
                    pass
                grid: list[list[tuple[str | None, bool]]] = []
                for row in ws.iter_rows():
                    rstyles: list[tuple[str | None, bool]] = []
                    for cell in row:
                        fill = None
                        bold = False
                        try:
                            f = cell.fill
                            if f is not None and f.patternType and f.fgColor is not None:
                                rgb = f.fgColor.rgb
                                if isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                                    fill = rgb
                        except Exception:
                            pass
                        try:
                            bold = bool(cell.font and cell.font.bold)
                        except Exception:
                            pass
                        rstyles.append((fill, bold))
                    grid.append(rstyles)
                out[ws.title] = grid
        except Exception:
            return out
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return out

    @staticmethod
    def _blank_hidden_cols(
        rows: list[list[Any]], hidden_cols: set[int]
    ) -> list[list[Any]]:
        """Blank author-hidden COLUMNS to None for the commercial/financial path.

        Scoped deliberately:
          * Columns only, never rows. A hidden column in a deal workbook is
            helper/formula scaffolding (a Gantt's "Country Multiplier", "Sell
            Helper", "Cost Helper"); a hidden ROW is usually outline-collapsed
            *real* content (LoE line items), so dropping rows loses data.
          * Commercial path only, never the SCOPE block atomizer. The commercial
            emitters bind by column index, so a blanked column simply doesn't
            bind. The block atomizer splits tables on empty columns, so blanking
            there would fragment a real table — hence we leave scope rows raw.
        Geometry-preserving (blank, not remove) so column-index locators hold."""
        if not hidden_cols:
            return rows
        return [
            [None if ci in hidden_cols else v for ci, v in enumerate(row)]
            for row in rows
        ]

    def _parse_xlsx(
        self, project_id: str, artifact_id: str, path: Path
    ) -> tuple[list[EvidenceAtom], list[dict[str, Any]], str | None]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            code = self._tabular_read_error_code(exc, tabular="xlsx")
            return [], [], f"{code}:{type(exc).__name__}:{exc}"
        hidden = self._hidden_dims(path)
        styles_by_sheet = self._sheet_styles(path)
        atoms: list[EvidenceAtom] = []
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            # read_only mode trusts the file's cached <dimension> tag; when that
            # is missing/stale (common when the source tool didn't update it, or
            # openpyxl strips a data-validation extension on read) iter_rows can
            # nondeterministically yield NOTHING. Force a real cell scan so the
            # parse is deterministic and never silently drops a whole sheet.
            try:
                sheet.reset_dimensions()
            except Exception:
                pass
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            hc, hr = hidden.get(sheet.title, (set(), set()))
            sheet_atoms = self._parse_sheet_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                filename=path.name,
                artifact_type=ArtifactType.xlsx,
                sheet_name=sheet.title,
                rows=rows,
                hidden_cols=hc,
                styles=styles_by_sheet.get(sheet.title),
            )
            # Single chokepoint (path-independent: block / legacy / commercial all
            # funnel here): mark atoms sourced from author-HIDDEN rows so a reviewer
            # can tell a collapsed/0-hour row from the live estimate. Captured, not
            # dropped (no silent loss) — just visibly tagged.
            sheet_atoms = self._flag_hidden_source_atoms(sheet_atoms, rows, hr)
            atoms.extend(sheet_atoms)
            sheets.append({"name": sheet.title, "rows": rows})
        # Dedup identical UNANSWERED questionnaire questions (e.g. a blank
        # "Phone Number" field repeated in Origin + Destination sections) —
        # they carry the same training signal; keep the first.
        seen_q: set[str] = set()
        deduped: list[EvidenceAtom] = []
        for a in atoms:
            if (a.value or {}).get("kind") == "qa_question":
                key = a.normalized_text or a.raw_text
                if key in seen_q:
                    continue
                seen_q.add(key)
            deduped.append(a)
        return deduped, sheets, None

    def _parse_csv(
        self, project_id: str, artifact_id: str, path: Path
    ) -> tuple[list[EvidenceAtom], list[dict[str, Any]], str | None]:
        try:
            with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                reader = csv.reader(handle)
                rows = [list(row) for row in reader]
        except Exception as exc:
            code = self._tabular_read_error_code(exc, tabular="csv")
            return [], [], f"{code}:{type(exc).__name__}:{exc}"
        sheet_atoms = self._parse_sheet_rows(
            project_id=project_id,
            artifact_id=artifact_id,
            filename=path.name,
            artifact_type=ArtifactType.csv,
            sheet_name="csv",
            rows=rows,
        )
        return sheet_atoms, [{"name": "csv", "rows": rows}], None

    def _build_structured_doc(
        self,
        *,
        schema: str,
        artifact_type: ArtifactType,
        filename: str,
        sheets: list[dict[str, Any]],
        parse_error: str | None = None,
    ) -> dict[str, Any]:
        """Render every sheet as a page with one section that contains the
        full table.  Empty / structureless sheets become a tiny note so
        downstream consumers can still see they exist.
        """
        pages: list[dict[str, Any]] = []
        for index, sheet in enumerate(sheets):
            sheet_name = sheet["name"]
            rows = sheet["rows"]
            section_blocks: list[dict[str, Any]] = []
            model = _detect_header(rows) if rows else None
            if model and model.header_idx >= 0:
                header_row_idx = model.header_idx
                header_row = rows[header_row_idx] or []
                if model.header_mode == "pair" and header_row_idx + 1 < len(rows):
                    bot = rows[header_row_idx + 1] or []
                    columns = [
                        " ".join(part for part in (str(a or "").strip(), str(b or "").strip()) if part).strip()
                        or f"col_{i + 1}"
                        for i, (a, b) in enumerate(
                            zip(header_row, bot + [None] * (len(header_row) - len(bot)))
                        )
                    ]
                    data_start = header_row_idx + 2
                else:
                    columns = [
                        (str(c).strip() if c is not None else "") or f"col_{i + 1}"
                        for i, c in enumerate(header_row)
                    ]
                    data_start = header_row_idx + 1
                table_rows: list[dict[str, Any]] = []
                for row_idx in range(data_start, len(rows)):
                    row = rows[row_idx] or []
                    if all(str(c or "").strip() == "" for c in row):
                        continue
                    cells: dict[str, Any] = {}
                    for col_idx, col_name in enumerate(columns):
                        if col_idx < len(row):
                            value = row[col_idx]
                        else:
                            value = ""
                        cells[col_name] = "" if value is None else str(value).strip()
                    table_rows.append(cells)
                if table_rows:
                    section_blocks.append(make_table(columns=columns, rows=table_rows))
            if not section_blocks:
                # Sheet has no detectable header; fall back to a raw textual
                # dump so the LLM still has something to look at.
                snippet_lines: list[str] = []
                for row in rows[:25]:
                    line = " | ".join(str(c).strip() for c in (row or []) if c is not None and str(c).strip())
                    if line:
                        snippet_lines.append(line)
                if snippet_lines:
                    section_blocks.append(
                        {"kind": "paragraph", "text": "\n".join(snippet_lines)}
                    )
            section = make_section(heading=sheet_name, level=2, blocks=section_blocks)
            pages.append(
                make_page(
                    page=index,
                    title=sheet_name,
                    sections=[section],
                )
            )
        metadata = [f"sheet: {s['name']}" for s in sheets]
        if parse_error:
            metadata.append(f"parse_error: {parse_error}")
        return make_structured_document(
            schema_version=schema,
            filename=filename,
            artifact_type=artifact_type.value,
            title=filename,
            metadata=metadata,
            pages=pages,
        )

    def _build_source_ref(
        self,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        columns: dict[str, str],
    ) -> SourceRef:
        return SourceRef(
            id=stable_id("src", artifact_id, sheet_name, row_number, stable_id("col", *sorted(columns.values()))),
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            locator={"sheet": sheet_name, "row": row_number, "columns": columns,
                     # Universal breadcrumb: every xlsx atom carries at least its
                     # sheet, so no atom renders under "(no section)".
                     "section_path": [sheet_name] if sheet_name else []},
            extraction_method="xlsx_table_mapping_v2_0",
            parser_version=self.parser_version,
        )

    def _maybe_emit_site_roster_atoms(
        self,
        *,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        rows: list[list[Any]],
    ) -> list[EvidenceAtom]:
        """Emit ``physical_site`` entity atoms when the sheet shape
        matches a site roster (header row with Site ID / Facility /
        Address / MDF / Escort + structured site-shaped IDs in data
        rows). Returns [] when this sheet is not a roster.
        """
        try:
            from app.parsers.site_roster_extractor import (
                extract_site_roster,
                looks_like_site_roster,
            )
        except Exception:  # pragma: no cover
            return []
        if not rows:
            return []
        # Skip blank leading rows
        first_data_row = 0
        for i, r in enumerate(rows):
            if any(str(c or "").strip() for c in (r or ())):
                first_data_row = i
                break
        body = rows[first_data_row:]
        if len(body) < 2:
            return []
        header_raw = [str(c or "").strip() for c in (body[0] or ())]
        data_rows: list[dict[str, Any]] = []
        for r in body[1:]:
            cells: dict[str, Any] = {}
            for i, v in enumerate(r or ()):
                col = header_raw[i] if i < len(header_raw) and header_raw[i] else f"col_{i}"
                cells[col] = "" if v is None else str(v)
            if any(v for v in cells.values()):
                data_rows.append(cells)
        if not data_rows:
            return []
        try:
            if not looks_like_site_roster(
                columns=header_raw, rows=data_rows, surrounding_text=sheet_name or ""
            ):
                return []
            # Stricter gate for XLSX (where site_id columns appear in
            # BOMs, decisions sheets, port maps, etc.): require at least
            # one additional roster-specific column beyond site_id.
            # Avoids treating a "Site ID | Decision | Approved By"
            # sheet as a roster.
            from app.parsers.site_roster_extractor import map_columns_to_fields
            field_map = map_columns_to_fields(header_raw)
            roster_specific = {
                "facility_name", "street_address", "mdf_idf",
                "access_window", "escort_owner", "city_state", "city", "state",
            }
            mapped = set(field_map.values())
            has_name = "facility_name" in mapped
            has_location = bool(mapped & {"street_address", "city", "state", "city_state"})
            if not (mapped & roster_specific) and not (has_name and has_location):
                return []
            roster_rows = extract_site_roster(
                columns=header_raw, rows=data_rows, surrounding_text=sheet_name or ""
            )
        except Exception:  # pragma: no cover
            return []
        if not roster_rows:
            return []
        out: list[EvidenceAtom] = []
        for site_row in roster_rows:
            sid = (site_row.site_id or "").strip()
            canon_id = sid or site_row.facility_name or ""
            if not canon_id:
                continue
            row_index = first_data_row + 1 + site_row.row_index
            atom_id = stable_id(
                "atm", artifact_id, sheet_name, row_index, "physical_site", canon_id
            )
            text_parts = []
            for label, val in [
                ("site_id", sid or site_row.site_id),
                ("facility", site_row.facility_name),
                ("address", site_row.street_address),
                ("mdf_idf", site_row.mdf_idf),
                ("access", site_row.access_window),
                ("escort", site_row.escort_owner),
                ("contact", site_row.contact),
                ("phone", site_row.phone),
                ("email", site_row.email),
            ]:
                if val:
                    text_parts.append(f"{label}: {val}")
            row_text = " | ".join(text_parts) or canon_id
            source_ref = SourceRef(
                id=stable_id("src", atom_id),
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                locator={
                    "sheet": sheet_name,
                    "row": row_index,
                    "section_path": [sheet_name] if sheet_name else [],
                    "extraction": "xlsx_site_roster_v1",
                },
                extraction_method="xlsx_site_roster_v1",
                parser_version=self.parser_version,
            )
            entity_keys: list[str] = []
            if sid:
                slug = re.sub(r"[^a-z0-9]+", "_", sid.lower()).strip("_")
                if slug:
                    entity_keys.append(f"site:{slug}")
            out.append(
                EvidenceAtom(
                    id=atom_id,
                    project_id=project_id,
                    artifact_id=artifact_id,
                    # v53.2 ROOT-CAUSE FIX: physical_site (not entity) so
                    # downstream gates can build the canonical site catalog.
                    atom_type=AtomType.physical_site,
                    raw_text=row_text,
                    normalized_text=row_text.lower(),
                    value={
                        "kind": "physical_site",
                        "id": sid or site_row.site_id,  # canonical id
                        "site_id": sid or site_row.site_id,
                        "name": site_row.facility_name,
                        "facility_name": site_row.facility_name,
                        "address": site_row.street_address,
                        "street_address": site_row.street_address,
                        "mdf_idf": site_row.mdf_idf,
                        "access_window": site_row.access_window,
                        "escort_owner": site_row.escort_owner,
                        "contact": site_row.contact,
                        "phone": site_row.phone,
                        "email": site_row.email,
                        "city": site_row.city,
                        "state": site_row.state,
                        "city_state": site_row.city_state,
                        "zip": site_row.zip,
                        "sqft": site_row.sqft,
                        "occupancy": site_row.occupancy,
                        "notes": site_row.notes,
                        "extras": dict(site_row.extra_fields),
                        # Bind the FULL raw row so EVERY column surfaces (Users,
                        # Rooms, Hardware/Services/Logistics budgets, Notes, ...),
                        # not just the canonical roster fields. _atom_bound_text
                        # renders these "Header: value | ..." at decide-time.
                        "cells": (
                            dict(data_rows[site_row.row_index])
                            if 0 <= site_row.row_index < len(data_rows)
                            else {}
                        ),
                    },
                    entity_keys=sorted(set(entity_keys)),
                    source_refs=[source_ref],
                    receipts=[],
                    authority_class=AuthorityClass.contractual_scope,
                    confidence=site_row.confidence,
                    confidence_raw=site_row.confidence,
                    calibrated_confidence=site_row.confidence,
                    review_status=ReviewStatus.auto_accepted,
                    review_flags=[],
                    parser_version=self.parser_version,
                )
            )
        return out

    # Sheets the classifier routes to COMMERCIAL (rate cards, master
    # catalogs, deal-financials) can be large pricebooks; cap the number
    # of pricing atoms per sheet so a 1000-row catalog can't flood the
    # envelope. Deal totals and rate cards are far smaller than this.
    _COMMERCIAL_ROW_CAP = 200
    # Bulk pricing sheets (rate cards / catalogs) fold every row into the
    # single rollup atom's value.rows rather than emitting per-row atoms.
    # This cap bounds envelope size for pathological sheets while staying
    # far above any realistic rate table.
    _COMMERCIAL_FOLD_CAP = 5000

    def _emit_financial_summary_rows(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        rows: list[list[Any]],
    ) -> list[EvidenceAtom]:
        """Structured extraction for a deal-financials / P&L summary sheet.

        Reads the 2-D label→value grid the way a PM does and emits:

        * one ``deal_metadata`` atom carrying the deal header (OPPTY #,
          customer, sales rep, billing type, duration, region, …);
        * one ``commercial_total`` atom per P&L category (Deal / Labor /
          PMO / Materials / Lift-Rental / Misc) with a structured
          ``value`` (revenue / cost / margin / margin_pct) so the
          OrbitBrief financial section renders without re-parsing text.

        This replaces the row-glue the generic commercial emitter produced
        for these sheets (``OPPTY # | 126 | Total Deal Revenue | 21560``).
        Deterministic, LLM-free, universal across estimating workbooks.
        """
        header: dict[str, Any] = {}
        header_locators: dict[str, dict[str, int]] = {}
        # canonical key -> RAW sheet label ("OPPTY #", "Project Duration (Months)")
        # so atoms DISPLAY the faithful label while still keying/merging by canon.
        header_labels: dict[str, str] = {}
        # category key -> {"display", "revenue", "cost", "margin",
        #                  "margin_pct", "row"}
        pl: dict[str, dict[str, Any]] = {}
        # per-column running category, so a margin-% row whose label carries no
        # category ("Net Margin (%)", or a garbled "Lift/Rental on Miscellaneous")
        # inherits the category of the P&L block it sits in.
        last_cat: dict[int, tuple[str, str]] = {}

        def _store_margin_pct(slot, rv, orig, ri, ci):
            """Store a margin-% value (or its #DIV/0! error) on a category slot,
            faithfully — a fraction becomes a percent, a failed formula is kept."""
            n = _coerce_pl_number(rv)
            if n is not None and slot.get("margin_pct") is None:
                slot["margin_pct"] = round(n * 100, 2) if abs(n) <= 1.5 else round(n, 2)
                slot["margin_pct_label"] = orig
                slot["margin_pct_pos"] = (ri + 1, ci + 1)
            elif n is None and slot.get("margin_pct") is None \
                    and slot.get("margin_pct_error") is None:
                err = _excel_error_str(rv)
                if err:
                    slot["margin_pct_error"] = err
                    slot["margin_pct_label"] = orig
                    slot["margin_pct_pos"] = (ri + 1, ci + 1)

        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row):
                if cell is None:
                    continue
                # Original-case, whitespace-collapsed text drives P&L
                # category *display*; the lowercased form drives matching.
                orig = re.sub(r"\s+", " ", str(cell).strip())
                label = orig.lower()
                if not label:
                    continue

                # ── deal header field (canonical key) ──
                key = _DEAL_HEADER_LABELS.get(label)
                if key and key not in header:
                    sval = _coerce_header_value(_right_neighbor_value(row, ci))
                    if sval is not None:
                        header[key] = sval
                        header_locators[key] = {"row": ri + 1, "col": ci + 1}
                        header_labels[key] = orig   # faithful raw label ("OPPTY #")
                    continue

                # ── P&L margin-percent line (regex OR semantic) ──
                is_mpct, cat_txt = _is_margin_pct_label(orig)
                if is_mpct:
                    if cat_txt:
                        ckey, disp = _norm_pl_category(cat_txt)
                    elif ci in last_cat:          # "Net Margin (%)" — inherit block
                        ckey, disp = last_cat[ci]
                    else:
                        ckey = disp = None
                    if ckey is not None:
                        rv = _right_neighbor_value(row, ci)
                        slot = pl.setdefault(ckey, {"display": disp, "row": ri + 1, "col": ci + 1})
                        last_cat[ci] = (ckey, disp)
                        _store_margin_pct(slot, rv, orig, ri, ci)
                        continue

                # ── P&L revenue / cost / margin line ──
                mm = _PL_METRIC_RE.match(orig)
                if mm and mm.group("metric").lower() in _PL_METRIC_WORDS:
                    ckey, disp = _norm_pl_category(mm.group("cat"))
                    metric = mm.group("metric").lower()
                    rv = _right_neighbor_value(row, ci)
                    n = _coerce_pl_number(rv)
                    slot = pl.setdefault(ckey, {"display": disp, "row": ri + 1, "col": ci + 1})
                    last_cat[ci] = (ckey, disp)
                    if n is not None and slot.get(metric) is None:
                        slot[metric] = round(n, 2)
                        slot[metric + "_label"] = orig   # faithful raw row label
                        slot[metric + "_pos"] = (ri + 1, ci + 1)
                    elif n is None and slot.get(metric) is None \
                            and slot.get(metric + "_error") is None:
                        err = _excel_error_str(rv)
                        if err:
                            slot[metric + "_error"] = err
                            slot[metric + "_label"] = orig
                            slot[metric + "_pos"] = (ri + 1, ci + 1)
                    continue

                # ── Structural backstop: a garbled label in the margin-% slot ──
                # An unclassified P&L-block row whose value reads as a percent or a
                # failed formula, sitting right after the category's Margin row
                # (which has a margin but no margin-% yet), IS that category's
                # margin %. Catches template typos ("Lift/Rental on Miscellaneous")
                # universally — by position, no customer-specific label needed.
                if ci in last_cat:
                    ckey, disp = last_cat[ci]
                    slot = pl.get(ckey)
                    if slot and slot.get("margin_pct") is None \
                            and slot.get("margin_pct_error") is None \
                            and (slot.get("margin") is not None or slot.get("margin_error") is not None):
                        rv = _right_neighbor_value(row, ci)
                        if _looks_like_pct_or_error(rv):
                            _store_margin_pct(slot, rv, orig, ri, ci)
                            continue

        # Confidence gate: only trust the structured P&L view when the grid
        # actually reads like a deal-kit financial summary — at least two
        # P&L categories carrying numbers, OR a multi-field deal header.
        # Otherwise (e.g. a country-multiplier matrix mislabelled as a
        # financial summary) fall back to the generic commercial emitter so
        # the sheet's real money rows aren't lost.
        pl_with_numbers = sum(
            1 for s in pl.values()
            if any(s.get(m) is not None for m in ("revenue", "cost", "margin", "margin_pct"))
        )
        if pl_with_numbers < 2 and len(header) < 3:
            return self._emit_commercial_sheet_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=rows,
                classification=classify_sheet(sheet_name, rows),
            )

        # ── generic header enrichment ──
        # The grid is now confirmed to be a deal-kit financial summary, so
        # capture ANY remaining label→value pair structurally (not just the
        # canonical vocabulary). This keeps non-standard header fields
        # (PO #, Account Manager, Site Contact, …) from being dropped. It
        # runs only past the confidence gate, so non-P&L grids — which fall
        # back above — are never affected. P&L metric/margin cells and
        # already-captured canonical fields are skipped.
        _GENERIC_HEADER_CAP = 40
        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row):
                if cell is None or len(header) >= _GENERIC_HEADER_CAP:
                    continue
                orig = re.sub(r"\s+", " ", str(cell).strip())
                label = orig.lower()
                if not label or label in _DEAL_HEADER_LABELS:
                    continue
                if not _looks_like_header_label(orig):
                    continue
                sval = _coerce_header_value(_right_neighbor_value(row, ci))
                if sval is None:
                    continue
                # A genuine header value is atomic — a name, code, number, or
                # date. A multi-word, non-numeric phrase ("Gross Margin Deal
                # Kit", "Overall Deal Kit Summary") is a section title that
                # happens to sit beside a caption, not a field value; skip it
                # so the sweep doesn't mint bogus heading fields.
                if len(sval.split()) >= 4 and not any(ch.isdigit() for ch in sval):
                    continue
                gkey = re.sub(r"[^a-z0-9]+", "_", label).strip("_")
                if not gkey or gkey in header:
                    continue
                header[gkey] = sval
                header_locators[gkey] = {"row": ri + 1, "col": ci + 1}
                header_labels[gkey] = orig   # faithful raw label

        atoms: list[EvidenceAtom] = []

        # 2-D section-title map. A titled financial block can sit ANYWHERE on the
        # sheet — left "Project Financials", right "Overall Deal Kit Summary",
        # stacked summaries — so a 1-D, column-A-only row scan mis-files a
        # right-side total under a left-side title. A title is a short banner cell
        # with an EMPTY right neighbour (not a label, which has a value to its
        # right; not a P&L metric / deal-header field). Each figure is then titled
        # by the nearest title ABOVE it IN ITS OWN COLUMN BAND — pure geometry, so
        # jumbled multi-block sheets attribute correctly and can't cross-file.
        _titles_2d: list[tuple[int, int, str]] = []   # (row1based, col0based, title)
        for _ri, _row in enumerate(rows):
            for _ci, _cell in enumerate(_row):
                if _cell is None or _is_money_number(_cell):
                    continue
                _at = re.sub(r"\s+", " ", str(_cell).strip())
                if not _at or not (1 <= len(_at.split()) <= 8):
                    continue
                _right = _row[_ci + 1] if _ci + 1 < len(_row) else None
                if _right is not None and str(_right).strip():
                    continue   # a value sits to the right -> this is a label, not a title
                _left = _row[_ci - 1] if _ci > 0 else None
                if _left is not None and str(_left).strip():
                    continue   # a label sits to the LEFT -> this is a VALUE (e.g. "PK"
                    #            next to "Sales Rep"), not a banner title
                _low = _at.lower()
                if _low in _DEAL_HEADER_LABELS or _PL_METRIC_RE.match(_at) or _PL_MARGIN_PCT_RE.match(_at):
                    continue
                _titles_2d.append((_ri + 1, _ci, _at))

        def _title_for_cell(r: Any, c: Any) -> str | None:
            if not isinstance(r, int) or r <= 0:
                return None
            cidx = (c - 1) if isinstance(c, int) and c > 0 else 0  # locators 1-based
            best_key = None
            best_title = None
            # nearest title ABOVE this cell, IN ITS COLUMN BAND (title at or up to
            # 3 cols left of the figure — covers merged banners).
            for tr, tc, tt in _titles_2d:
                if tr <= r and tc <= cidx and (cidx - tc) <= 3:
                    key = (tr, -(cidx - tc))   # closest row above, then closest col
                    if best_key is None or key > best_key:
                        best_key, best_title = key, tt
            if best_title is None:
                # no in-band title -> nearest title above by row, any column
                for tr, tc, tt in _titles_2d:
                    if tr <= r and (best_key is None or tr > best_key[0]):
                        best_key, best_title = (tr, 0), tt
            return best_title

        def _src(tag: str, locator: dict[str, Any]) -> SourceRef:
            loc = {"sheet": sheet_name, "extraction": "financial_summary", **locator}
            _title = loc.pop("section_title", None) or _title_for_cell(loc.get("row"), loc.get("col"))
            loc["section_path"] = [sheet_name] + ([_title] if _title and _title != sheet_name else [])
            return SourceRef(
                id=stable_id("src", artifact_id, sheet_name, tag),
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                locator=loc,
                extraction_method="financial_summary",
                parser_version=self.parser_version,
            )

        # ── deal header: ONE atom PER FIELD (uniform "row = atom": the deal
        #    header is a label->value grid, exactly like the P&L). Each field is
        #    individually typeable for the heads (Billing Type -> billing facet,
        #    Region -> geo, Customer -> party). The identity record is reassembled
        #    at render time — build_deal_header already MERGES deal_header atoms
        #    field-by-field, so each atom carries value.fields = {one field}. ──
        for k, v in header.items():
            loc = header_locators.get(k, {})
            label = header_labels.get(k) or k.replace("_", " ").title()  # RAW label
            text = f"{label}: {v}"[:4000]
            ekeys: list[str] = []
            if k == "opportunity_id" and v:
                ekeys.append(f"deal:{v}")
            elif k == "customer" and v:
                cust_slug = re.sub(r"[^a-z0-9]+", "_", str(v).lower()).strip("_")
                if cust_slug:
                    ekeys.append(f"customer:{cust_slug}")
            aid = stable_id("atm", artifact_id, "deal_metadata", sheet_name, k)
            atoms.append(
                EvidenceAtom(
                    id=aid,
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=AtomType.deal_metadata,
                    raw_text=text,
                    normalized_text=text.lower(),
                    value={"kind": "deal_header", "fields": {k: v}, "field": k,
                           "sheet_name": sheet_name},
                    entity_keys=ekeys,
                    source_refs=[_src(f"deal_{k}", {"row": loc.get("row", 1),
                                                    "col": loc.get("col", 1)})],
                    receipts=[],
                    authority_class=AuthorityClass.vendor_quote,
                    confidence=0.8,
                    confidence_raw=0.8,
                    calibrated_confidence=0.8,
                    review_status=ReviewStatus.needs_review,
                    review_flags=[],
                    parser_version=self.parser_version,
                )
            )

        # ── P&L line atoms — ONE atom PER ROW (uniform with PDF/docx tables): each
        # financial figure ("Total Labor Revenue: $47,150") is its own atom with
        # its FAITHFUL row label, titled by its own 2-D block. The PM brief regroups
        # them per category at render time (build_deal_financials). ──
        def _fmt(v: Any) -> str:
            return f"${int(round(v)):,}" if isinstance(v, (int, float)) else "n/a"

        def _pl_atom(ckey, slot, metric, label, val, text, flags=None, is_error=False):
            prow, pcol = slot.get(metric + "_pos", (slot.get("row", 0), slot.get("col", 1)))
            aid = stable_id("atm", artifact_id, "pl", sheet_name, ckey, metric)
            value = {"kind": "pl_metric", "category": slot["display"], "category_key": ckey,
                     "metric": metric, "value": val, "sheet_name": sheet_name}
            if is_error:
                value["formula_error"] = val
            return EvidenceAtom(
                id=aid, project_id=project_id, artifact_id=artifact_id,
                atom_type=AtomType.commercial_total, raw_text=text, normalized_text=text.lower(),
                value=value,
                entity_keys=([f"money:{int(round(val))}"]
                             if isinstance(val, (int, float)) and abs(val) >= _MIN_MONEY_VALUE
                             and metric != "margin_pct" else []),
                source_refs=[_src(f"pl_{ckey}_{metric}", {"row": prow, "col": pcol})],
                receipts=[], authority_class=AuthorityClass.vendor_quote,
                confidence=0.78, confidence_raw=0.78, calibrated_confidence=0.78,
                review_status=ReviewStatus.needs_review, review_flags=list(flags or []),
                parser_version=self.parser_version)

        for ckey, slot in sorted(pl.items(), key=lambda kv: (kv[1].get("col", 1), kv[1].get("row", 1_000_000))):
            disp = slot["display"]
            for metric in ("revenue", "cost", "margin"):
                val = slot.get(metric)
                if val is not None:
                    label = slot.get(metric + "_label") or f"{disp} {metric.title()}"
                    atoms.append(_pl_atom(ckey, slot, metric, label, val, f"{label}: {_fmt(val)}"))
                    continue
                # A formula that failed (#DIV/0!) is faithful, meaningful content —
                # emit it (flagged) so the metric isn't a silent gap in the P&L.
                err = slot.get(metric + "_error")
                if err:
                    label = slot.get(metric + "_label") or f"{disp} {metric.title()}"
                    atoms.append(_pl_atom(ckey, slot, metric, label, err, f"{label}: {err}",
                                          flags=["xlsx_parser:formula_error"], is_error=True))
            mpct = slot.get("margin_pct")
            if mpct is not None:
                label = slot.get("margin_pct_label") or f"Margin % on {disp}"
                atoms.append(_pl_atom(ckey, slot, "margin_pct", label, mpct, f"{label}: {mpct:g}%"))
            else:
                err = slot.get("margin_pct_error")
                if err:
                    label = slot.get("margin_pct_label") or f"Margin % on {disp}"
                    atoms.append(_pl_atom(ckey, slot, "margin_pct", label, err, f"{label}: {err}",
                                          flags=["xlsx_parser:formula_error"], is_error=True))

        # If the structured pass found nothing usable (an oddly-shaped
        # sheet), fall back to the generic commercial emitter so we never
        # regress to zero atoms.
        if not atoms:
            return self._emit_commercial_sheet_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=rows,
                classification=classify_sheet(sheet_name, rows),
            )
        # Right-side summary sub-tables ("Overall Deal Kit Summary", "Deal Kit
        # Excluding Expenses…", "Gross Margin Deal Kit") sit in their OWN titled
        # blocks the P&L category scan above never reaches. Pull them via the
        # block atomizer so each is captured under its title — deduped by dollar
        # amount so a block that merely restates the main P&L isn't doubled.
        atoms += self._emit_financial_summary_supplements(
            project_id, artifact_id, artifact_type, filename, sheet_name, rows, atoms
        )
        # Emit one SECTION at a time, column-major (left blocks top-to-bottom, then
        # right) so a block's atoms stay contiguous instead of interleaving — sort
        # by the cell's (col, row), the figure's true position on the sheet.
        def _pos(a: EvidenceAtom) -> tuple[int, int]:
            loc = a.source_refs[0].locator if a.source_refs else {}
            return (loc.get("col", 1) or 1, loc.get("row", 0) or 0)
        atoms.sort(key=_pos)
        return atoms

    def _emit_financial_summary_supplements(
        self, project_id, artifact_id, artifact_type, filename, sheet_name, rows, existing,
    ) -> list[EvidenceAtom]:
        """Capture titled financial summary BLOCKS the P&L scan missed (the
        stacked right-side summaries). Emits one ``commercial_total`` per titled
        keyval block carrying >=2 numbers, skipping the deal-header block and any
        block whose dollar figures are already represented by ``existing`` (so
        the main P&L isn't duplicated). section_path = [sheet, title]."""
        from app.parsers.xlsx_blocks import sheet_blocks
        try:
            blocks = sheet_blocks(rows)
        except Exception:
            return []
        seen = set()
        for a in existing:
            for m in re.findall(r"\d{4,}", (getattr(a, "raw_text", "") or "").replace(",", "")):
                seen.add(m)
        # position of each block's title cell (row,col) so supplement atoms sort
        # block-contiguous with the main P&L (column-major: left side, then right).
        def _title_pos(t: str) -> tuple[int, int]:
            for ri, row in enumerate(rows):
                for ci, c in enumerate(row):
                    if str(c or "").strip() == t:
                        return ri + 1, ci + 1
            return 1_000_000, 1_000_000

        out: list[EvidenceAtom] = []
        for bi, b in enumerate(blocks):
            title = b.get("title")
            if not title or b.get("kind") != "keyval":
                continue
            tl = title.lower()
            if "deal summary" in tl or "project financials" in tl:
                continue
            pairs = b["pairs"]
            trow, tcol = _title_pos(title)
            # Block-level dedup: skip a block whose big dollar figures are ALL
            # already represented (a pure restatement of the main P&L).
            blk_big = set()
            for _, v in pairs:
                vs = str(v).replace(",", "").replace("$", "").strip().rstrip("%")
                if re.match(r"^-?\d+(\.\d+)?$", vs) and abs(float(vs)) >= 1000:
                    blk_big.add(str(int(round(float(vs)))))
            if blk_big and blk_big <= seen:
                continue
            # ONE atom PER ROW (uniform with the main P&L + PDF/docx tables): each
            # label->value pair is its own faithful atom, titled by this block.
            emitted = False
            for pj, (k, v) in enumerate(pairs):
                kk = re.sub(r"\s+", " ", str(k).strip())
                if not kk or v in (None, ""):
                    continue
                vs = str(v).replace(",", "").replace("$", "").strip().rstrip("%")
                if not re.match(r"^-?\d+(\.\d+)?$", vs):
                    continue
                num = float(vs)
                mm = _PL_METRIC_RE.match(kk)
                mp = _PL_MARGIN_PCT_RE.match(kk)
                if mm and mm.group("metric").lower() in _PL_METRIC_WORDS:
                    ckey, _d = _norm_pl_category(mm.group("cat"))
                    metric = mm.group("metric").lower()
                elif mp:
                    ckey, _d = _norm_pl_category(mp.group("cat"))
                    metric = "margin_pct"
                else:
                    ckey, metric = "other", "value"
                # format uniformly with the main P&L: $ for money, % for margins
                if metric == "margin_pct":
                    pct = num * 100 if abs(num) <= 1.5 else num
                    disp_v, store_v = f"{round(pct, 2):g}%", round(pct, 2)
                elif metric in ("revenue", "cost", "margin"):
                    disp_v, store_v = f"${int(round(num)):,}", round(num, 2)
                else:
                    disp_v, store_v = str(v), num
                text = f"{kk}: {disp_v}"
                aid = stable_id("atm", artifact_id, "fin_block", sheet_name, bi, pj)
                out.append(EvidenceAtom(
                    id=aid, project_id=project_id, artifact_id=artifact_id,
                    atom_type=AtomType.commercial_total, raw_text=text[:4000],
                    normalized_text=text[:4000].lower(),
                    value={"kind": "pl_metric", "category": title, "category_key": ckey,
                           "metric": metric, "value": store_v, "block_title": title,
                           "sheet_name": sheet_name},
                    entity_keys=([f"money:{int(round(num))}"]
                                 if abs(num) >= _MIN_MONEY_VALUE and metric != "margin_pct" else []),
                    source_refs=[SourceRef(
                        id=stable_id("src", aid), artifact_id=artifact_id, artifact_type=artifact_type,
                        filename=filename,
                        locator={"sheet": sheet_name, "section_path": [sheet_name, title],
                                 "row": trow, "col": tcol, "extraction": "fin_summary_block_v1"},
                        extraction_method="fin_summary_block_v1", parser_version=self.parser_version)],
                    receipts=[], authority_class=AuthorityClass.contractual_scope,
                    confidence=0.80, confidence_raw=0.80, calibrated_confidence=0.80,
                    review_status=ReviewStatus.auto_accepted, review_flags=[],
                    parser_version=self.parser_version))
                emitted = True
            if emitted:
                for n in blk_big:
                    seen.add(n)
        return out

    def _emit_commercial_sheet_rows(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        rows: list[list[Any]],
        classification: Any,
    ) -> list[EvidenceAtom]:
        """Route a non-scope *pricing* sheet to typed commercial atoms.

        Rate cards, master catalogs and deal-financials tabs are not
        customer scope, but they carry pricing the PM needs. Rather than
        drop them, every money-bearing row becomes a typed commercial
        atom (``commercial_total`` for a deal-financials summary,
        ``pricing_assumption`` for rate-card / catalog rows), tagged
        ``vendor_quote`` authority and carrying ``money:`` entity keys so
        it feeds the OrbitBrief ``pricing_clarity`` dimension without ever
        landing in ``scope_truth``.
        """
        from app.parsers.sheet_classifier import SheetRole

        role = classification.role
        atom_type = (
            AtomType.commercial_total
            if role is SheetRole.FINANCIAL_SUMMARY
            else AtomType.pricing_assumption
        )

        # RATE_CARD / CATALOG are master reference pricing — NOT deal scope. They
        # fold into ONE rollup summary atom (``collapse_to_summary``) whose
        # ``value.rows`` carries the full matrix for drill-down / reconciliation.
        # Per-row ``pricing_assumption`` atoms from a 300-line catalog were the
        # #1 accuracy killer (inflated totals, drowned real deal pricing, tanked
        # grades). FINANCIAL_SUMMARY sheets emit per-row atoms — those ARE the
        # deal economics the PM reads.
        collapse_to_summary = role is not SheetRole.FINANCIAL_SUMMARY

        money_cols = _money_columns(rows)
        if role is SheetRole.RATE_CARD and not money_cols:
            money_cols = _rate_card_value_columns(rows)
        atoms: list[EvidenceAtom] = []
        all_values: list[float] = []
        folded_rows: list[dict[str, Any]] = []
        # Detect the column-header row so each row renders as "Header: value"
        # bound pairs instead of a bare pipe blob, and the hidden COMPUTED tail
        # columns (the formula cells beyond the visible header — the unlabeled
        # "1 | 82 | 63.5" Benjamin flagged) are dropped from the display. The
        # dollar figures in those columns still flow through ``money_keys`` /
        # ``value`` below, so the commercial pipeline is unchanged.
        # Compose the (possibly multi-row) header band and learn which rows are
        # header / above-data scaffolding so they are never emitted as priced
        # lines. ``_header_rows`` = the band; ``_data_floor`` = first emittable row.
        _headers, _header_rows, _data_floor = _commercial_header_band(rows, money_cols)
        # index of the first labelled column — a row only header-binds if it has
        # a value there (i.e. it belongs to THIS table). A side block living in
        # far-right columns (e.g. a travel calc next to the main pricing table)
        # has that column empty, so it falls back to its plain form instead of
        # being mis-mapped onto the main table's headers.
        _first_hdr_col = next((i for i, h in enumerate(_headers) if h), None)

        # Which rows carry money — used to tell a CATEGORY-DIVIDER row (a label
        # that HEADS a run of priced data rows, e.g. a "CAT6…" banner over the
        # cabling block) from a stray note. A divider is provable STRUCTURE, not
        # wording, so it generalizes to any category name without a keyword list.
        _money_row = [bool(_row_money_values(r, money_cols)) for r in rows]

        def _is_category_divider(cells: list[str], row_idx: int) -> bool:
            ne = [c for c in cells if c]
            distinct = {c for c in ne}
            if not (1 <= len(distinct) <= 2):
                return False
            if any(len(c) > 70 for c in distinct):
                return False
            if any(c.replace(",", "").replace(".", "").replace("$", "").lstrip("-").isdigit() for c in ne):
                return False
            # Must sit directly above a REAL data row (>=3 filled cells with
            # money) within the next few rows — that is what makes it a section
            # banner rather than a lone total label or footnote.
            for j in range(row_idx + 1, min(row_idx + 4, len(rows))):
                rj = [("" if c is None else str(c).strip()) for c in rows[j]]
                if not any(rj):
                    continue
                return _money_row[j] and sum(1 for c in rj if c) >= 3
            return False

        current_section: str | None = None
        for row_idx, row in enumerate(rows):
            cells = [("" if c is None else str(c).strip()) for c in row]
            if not any(cells):
                continue
            if row_idx in _header_rows or row_idx < _data_floor:
                # Header band rows AND everything above the data block (dropdown
                # source lists, base-rate scratch rows) are structure, not priced
                # lines — never emit them as atoms.
                continue
            values = _row_money_values(row, money_cols)
            if not values and _is_category_divider(cells, row_idx):
                # Becomes the running section breadcrumb for the rows beneath it
                # (e.g. Materials > CAT6…); not emitted as a priced atom itself.
                current_section = max((c for c in cells if c), key=len).strip(" .…")[:60]
                continue
            _aligned = (
                _headers
                and _first_hdr_col is not None
                and _first_hdr_col < len(cells)
                and cells[_first_hdr_col]
            )
            if not values:
                # Header / label rows with no dollar figure carry no pricing
                # signal — skip so the commercial view stays clean. But on a
                # per-line financial-summary sheet (not a collapsed catalog), a
                # row OUTSIDE the main header span — a side calc block beside the
                # table (e.g. a travel breakdown's "Team | 4") — is a real
                # label->value fact with no money of its own; keep it so it
                # isn't silently dropped. Catalogs stay strict (money only) so
                # their rollup counts don't drift.
                if collapse_to_summary or _aligned or not _is_side_label_value(cells):
                    continue
            all_values.extend(values)
            money_keys = sorted({f"money:{int(round(v))}" for v in values})
            if _aligned:
                _bound = [
                    f"{_headers[ci]}: {cells[ci]}"
                    for ci in range(len(_headers))
                    if ci < len(cells) and _headers[ci] and cells[ci]
                ]
                row_text = (" | ".join(_bound) or " | ".join(c for c in cells if c))[:4000]
            else:
                row_text = " | ".join(c for c in cells if c)[:4000]
            # A label-less row whose numbers SUM the rows above is a totals row —
            # give the orphan a "Total" label (arithmetic-proven, universal).
            if _unlabeled_sum_row(rows, row_idx):
                row_text = f"Total | {row_text}"
            label = " ".join(
                c for c in cells if c and not c.replace(",", "").replace(".", "").lstrip("-").isdigit()
            ).strip()[:300]

            if collapse_to_summary:
                # Fold into the rollup — summary-only emission for rate cards /
                # catalogs. No per-row atoms (the flood that broke #010063).
                folded_rows.append(
                    {
                        "row": row_idx + 1,
                        "label": label,
                        "text": row_text,
                        "money_keys": money_keys,
                        "cells": [c for c in cells if c],
                    }
                )
                continue

            atom_id = stable_id(
                "atm", artifact_id, atom_type.value, sheet_name, row_idx
            )
            src = SourceRef(
                id=stable_id("src", atom_id),
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                locator={
                    "sheet": sheet_name,
                    "row": row_idx + 1,
                    "section_path": (
                        [sheet_name] + ([current_section] if current_section else [])
                        if sheet_name else []
                    ),
                    "extraction": "commercial_sheet_routing",
                },
                extraction_method="commercial_sheet_routing",
                parser_version=self.parser_version,
            )
            atoms.append(
                EvidenceAtom(
                    id=atom_id,
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=row_text,
                    normalized_text=row_text.lower(),
                    value={
                        "label": label,
                        "money_keys": money_keys,
                        "sheet_role": role.value,
                        "sheet_name": sheet_name,
                        "cells": [c for c in cells if c],
                    },
                    entity_keys=money_keys,
                    source_refs=[src],
                    receipts=[],
                    authority_class=AuthorityClass.vendor_quote,
                    confidence=0.7,
                    confidence_raw=0.7,
                    calibrated_confidence=0.7,
                    review_status=ReviewStatus.needs_review,
                    review_flags=[],
                    parser_version=self.parser_version,
                )
            )

        # ``line_count`` reflects every money-bearing row found, whether it
        # became its own atom (financial summary) or was folded (rate card).
        line_count = len(atoms) if not collapse_to_summary else len(folded_rows)
        if line_count == 0:
            return []

        # Roll-up banner: only for COLLAPSED sheets (rate cards / catalogs),
        # where it IS the sheet's single atom — it carries the full row matrix
        # for drill-down and is the atom the pricing_rollup packet consumes.
        # On a per-row financial-summary sheet the granular rows ARE the atoms,
        # so a synthetic "N pricing lines, $lo-$hi" banner is redundant noise
        # that nothing downstream consumes (not packetized, not in
        # build_deal_financials) — the very "Gantt Financials: 12 pricing
        # lines, $-1,908-$26,450" atom the PM flagged as confusing. Drop it.
        if not collapse_to_summary:
            return atoms
        summary = self._commercial_summary_atom(
            project_id=project_id,
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            sheet_name=sheet_name,
            role=role,
            atom_type=atom_type,
            line_count=line_count,
            values=all_values,
            folded_rows=folded_rows,
        )
        # Summary-only: one atom per reference sheet, full matrix in value.rows.
        return [summary]

    def _commercial_summary_atom(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        role: Any,
        atom_type: AtomType,
        line_count: int,
        values: list[float],
        folded_rows: list[dict[str, Any]] | None = None,
    ) -> EvidenceAtom:
        lo = min(values) if values else 0.0
        hi = max(values) if values else 0.0
        total = sum(values)
        money_keys = sorted(
            {f"money:{int(round(v))}" for v in (lo, hi, total) if v >= _MIN_MONEY_VALUE}
        )
        # ASCII hyphen separator — the OrbitBrief renderer and several JSON
        # consumers round-tripped the en-dash through a non-UTF-8 stage and
        # surfaced it as U+FFFD ("$8�$5,390"). A plain "-" is unambiguous.
        label = (
            f"{sheet_name}: {line_count} pricing line"
            f"{'s' if line_count != 1 else ''}, "
            f"${int(round(lo)):,}-${int(round(hi)):,}"
        )
        atom_id = stable_id(
            "atm", artifact_id, atom_type.value, sheet_name, "summary"
        )
        src = SourceRef(
            id=stable_id("src", atom_id),
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            locator={
                "sheet": sheet_name,
                "section_path": [sheet_name] if sheet_name else [],
                "extraction": "commercial_sheet_routing",
                "rollup": True,
            },
            extraction_method="commercial_sheet_routing",
            parser_version=self.parser_version,
        )
        return EvidenceAtom(
            id=atom_id,
            project_id=project_id,
            artifact_id=artifact_id,
            atom_type=atom_type,
            raw_text=label,
            normalized_text=label.lower(),
            value={
                "is_summary": True,
                "label": label,
                "line_count": line_count,
                "money_min": round(lo, 2),
                "money_max": round(hi, 2),
                "money_sum": round(total, 2),
                "money_keys": money_keys,
                "sheet_role": role.value,
                "sheet_name": sheet_name,
                # Full row matrix for collapsed bulk sheets (rate cards /
                # catalogs). None for financial-summary sheets whose rows
                # are emitted as their own atoms.
                "rows": folded_rows if folded_rows else None,
            },
            entity_keys=money_keys,
            source_refs=[src],
            receipts=[],
            authority_class=AuthorityClass.vendor_quote,
            confidence=0.7,
            confidence_raw=0.7,
            calibrated_confidence=0.7,
            review_status=ReviewStatus.needs_review,
            review_flags=["pricing_rollup"],
            parser_version=self.parser_version,
        )

    def _dropped_sheet_marker(
        self,
        *,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        rows: list[list[Any]],
        reason: str,
        role: str = "",
    ) -> EvidenceAtom:
        """A retained marker for a whole sheet the role-router classified DROP.

        Carries the sheet's rows (capped) so a PM omission complaint can still
        recover the content, and is pre-stamped ``suppressed:sheet_router`` so
        the compiler diverts it into the suppressed sidecar — it never reaches
        scope_truth. This is the parse-time analog of the suppression ledger.
        """
        # Cap retained rows — a DROP sheet is noise; we keep enough to localize
        # an omission complaint without bloating the envelope.
        capped = [[("" if c is None else str(c)) for c in r] for r in rows[:500]]
        # Plain-English "why" so the dropped marker reads as a deliberate skip,
        # not a parser miss (a reviewer seeing a full grid on the left and one
        # line on the right otherwise assumes the parser broke).
        why = {
            "reference": "lookup / rate-backing sheet - feeds the pricing tabs by formula, "
            "not this deal's content",
            "empty": "empty sheet - no data",
            "instructions": "instructions / cover / terms - no deal data",
        }.get(role, "backing data, not deal content")
        marker_text = f"[skipped: {why}] '{sheet_name}', {len(rows)} rows"
        atom_id = stable_id("atm", artifact_id, "dropped_sheet", sheet_name)
        src = SourceRef(
            id=stable_id("src", atom_id),
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            locator={"sheet": sheet_name, "extraction": "sheet_role_router"},
            extraction_method="sheet_role_router",
            parser_version=self.parser_version,
        )
        return EvidenceAtom(
            id=atom_id,
            project_id=project_id,
            artifact_id=artifact_id,
            atom_type=AtomType.dropped_sheet,
            raw_text=marker_text,
            normalized_text=f"dropped sheet {sheet_name}".lower(),
            value={
                "sheet_name": sheet_name,
                "row_count": len(rows),
                "rows": capped,
                "_suppression": {"stage": "sheet_router", "reason": reason},
            },
            entity_keys=[],
            source_refs=[src],
            receipts=[],
            authority_class=AuthorityClass.machine_extractor,
            confidence=0.0,
            confidence_raw=0.0,
            calibrated_confidence=0.0,
            review_status=ReviewStatus.needs_review,
            review_flags=["suppressed:sheet_router"],
            parser_version=self.parser_version,
        )

    def _parse_sheet_rows(
        self,
        project_id: str,
        artifact_id: str,
        filename: str,
        artifact_type: ArtifactType,
        sheet_name: str,
        rows: list[list[Any]],
        hidden_cols: set[int] | None = None,
        styles: list[list[tuple[str | None, bool]]] | None = None,
    ) -> list[EvidenceAtom]:
        if not rows:
            return []

        # Sheet-role router: estimating workbooks carry rate-card backing
        # lists, master price catalogs, deal-financials tabs and empty
        # template sheets that are not customer scope. The classifier
        # routes each sheet so non-scope sheets can't bury real evidence:
        #   • DROP       → empty / cover / lookup-helper noise: emit nothing.
        #   • COMMERCIAL → rate card / catalog / deal financials: emit TYPED
        #     commercial atoms (pricing the PM needs) — never scope_item, so
        #     scope_truth stays clean while pricing still surfaces.
        #   • SCOPE      → fall through to normal row mining.
        classification = classify_sheet(sheet_name, rows)
        if classification.destination is SheetDestination.DROP:
            # Retained-suppression: instead of vanishing, a DROP-classified
            # sheet is emitted as ONE marker atom carrying its rows, stamped
            # suppressed:sheet_router. The compiler diverts pre-suppressed
            # parser atoms into CompileResult.suppressed_atoms, so the sheet
            # never reaches scope but stays auditable + complaint-localizable.
            return [
                self._dropped_sheet_marker(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    sheet_name=sheet_name,
                    rows=rows,
                    reason=getattr(classification, "reason", "") or "sheet_role=DROP",
                    role=getattr(getattr(classification, "role", None), "value", ""),
                )
            ]
        if classification.destination is SheetDestination.COMMERCIAL:
            # Blank author-hidden helper columns (Country Multiplier / Sell
            # Helper / Cost Helper …) so they can't leak as unlabeled atoms.
            # Scoped to the commercial path only — see _blank_hidden_cols.
            comm_rows = self._blank_hidden_cols(rows, hidden_cols or set())
            # Deal-financials / P&L sheets get a structured label→value
            # extractor (clean deal header + per-category P&L atoms);
            # rate cards / catalogs stay on the row/rollup path.
            if classification.role is SheetRole.FINANCIAL_SUMMARY:
                return self._emit_financial_summary_rows(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    sheet_name=sheet_name,
                    rows=comm_rows,
                )
            return self._emit_commercial_sheet_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=comm_rows,
                classification=classification,
            )

        # RF1 — explicit fast-path for known structured-row CSVs.
        # Files named asset_inventory / site_list / risk_register /
        # license_support_matrix / lifecycle route to the typed-row
        # profiler so they produce asset_record / site_roster / risk /
        # support_entitlement / lifecycle_status atoms (PR2). This must
        # run BEFORE the site-roster fast path: an asset_inventory CSV
        # (Asset ID / Serial / IP Address / MAC Address) would otherwise
        # be mis-read as a site roster (AST-001 matches the site-ID shape
        # and "IP Address" matches the street-address header) and emit
        # ghost physical_site atoms. Restricted to .csv to avoid
        # disturbing existing .xlsx fixtures (e.g. demo_project's
        # site_list.xlsx that legacy tests expect on the canonical path).
        _STRUCTURED_NAMES = (
            "asset_inventory", "site_list", "risk_register",
            "license_support", "support_matrix", "lifecycle",
        )
        stem = filename.lower().replace("-", "_").replace(" ", "_")
        if (
            artifact_type == ArtifactType.csv
            and any(s in stem for s in _STRUCTURED_NAMES)
        ):
            return self._emit_generic_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=rows,
            )

        # Site-roster fast path: when this sheet's first non-empty row
        # looks like site_roster headers (Site ID / Facility / Address /
        # MDF / Access / Escort), route the rows through the same
        # extractor the PDF parser uses so XLSX-shipped rosters produce
        # structured ``physical_site`` atoms (not just generic row atoms).
        roster_atoms = self._maybe_emit_site_roster_atoms(
            project_id=project_id,
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            sheet_name=sheet_name,
            rows=rows,
        )
        if roster_atoms:
            return roster_atoms

        # PR2 (post-v3 review) — operational-workbook sheet profiles.
        # If the sheet name matches one of the well-known ops sheets
        # (Asset Inventory / Site Survey / Port Map & VLANs / Circuit
        # Inventory / License Support / NOC Alert Matrix / Risk
        # Register / Cutover Validation / README / Dashboard /
        # Source Refs), dispatch every row to its typed AtomType.
        op_profile = _OPERATIONAL_SHEET_PROFILES.get(_norm_op_sheet(sheet_name))
        if op_profile is not None:
            ops_atoms = self._emit_operational_sheet_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=rows,
                profile=op_profile,
            )
            if ops_atoms:
                return ops_atoms
            # Fall through to the legacy path if the ops emitter
            # produced nothing (e.g. blank sheet).

        # Block-structured emission: a single sheet often holds MULTIPLE tables
        # under their own titles (a Deal Kit "Summary" tab = a "Detailed Level of
        # Effort" table + a side "Key Unit Metrics" table), plus key-value header
        # blocks. The legacy single-header-per-sheet model picks ONE header and
        # drops the rest (the LOE table vanished, titles were lost). The block
        # atomizer splits the sheet into title/header/rows blocks, so every atom
        # carries a real path (sheet > title) and each row is keyed to ITS block's
        # column headers — the organization signal the heads need.
        if os.environ.get("SOWSMITH_XLSX_BLOCKS", "1") != "0":
            block_atoms = self._emit_block_structured_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=rows,
                styles=styles,
            )
            if block_atoms:
                return block_atoms

        model = _detect_header(rows)
        if model.header_idx < 0:
            # PRODUCTION_GAPS P0.4: when the canonical-header detector can't
            # find a cabling/networking-style schedule (plate_id, site,
            # quantity columns), fall back to a generic row-as-atom emitter.
            # Real-world XLSX attachments — Q&A logs, fee schedules, vendor
            # response matrices, RFP cost tables — almost never use the
            # canonical column names this parser was originally tuned for.
            # The generic emitter still produces structured atoms so the
            # downstream entity_extraction stage can pull
            # device/vendor/quantity/site keys out of the row text.
            return self._emit_generic_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=rows,
            )
        data_start = model.header_idx + (2 if model.header_mode == "pair" else 1)
        atoms: list[EvidenceAtom] = []
        label_indices = _label_column_indices(model.header_map)

        # v49.2: capture raw header row once for raw_table_row emission.
        # The centralized _enrich_table_atoms() in entity_extraction
        # will classify per row using the column schema registry.
        _raw_headers: list[str] = []
        if 0 <= model.header_idx < len(rows):
            _raw_headers = [str(c or "").strip() for c in rows[model.header_idx]]

        for row_idx in range(data_start, len(rows)):
            row = rows[row_idx]
            if _is_blank_row(row):
                continue
            rk = _row_kind(row, model.header_map, label_indices)
            if rk in {"blank", "header", "title", "section_header"}:
                continue
            if rk == "note":
                continue

            extracted = self._extract_row_values(row, model.header_map)

            # v49.2: emit raw_table_row for centralized classification.
            if _raw_headers:
                _row_cells = [str(c or "").strip() for c in row]
                _row_text = " | ".join(c for c in _row_cells if c)[:4000]
                _rtr_id = stable_id("atm", artifact_id, "raw_table_row", sheet_name, row_idx)
                _rtr_columns = {
                    h: get_column_letter(i + 1)
                    for i, h in enumerate(_raw_headers)
                    if h
                }
                _rtr_src = SourceRef(
                    id=stable_id("src", _rtr_id),
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    locator={"sheet": sheet_name, "row": row_idx + 1, "columns": _rtr_columns, "extraction": "raw_table_row_v49_2", "section_path": [sheet_name] if sheet_name else []},
                    extraction_method="raw_table_row_v49_2",
                    parser_version=self.parser_version,
                )
                atoms.append(EvidenceAtom(
                    id=_rtr_id,
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=AtomType.raw_table_row,
                    raw_text=_row_text,
                    normalized_text=_row_text.lower(),
                    value={
                        "_columns": list(_raw_headers),
                        "_row": _row_cells,
                        "_table_idx": 0,
                        "_row_idx": row_idx,
                        "_filename": filename,
                        "_sheet": sheet_name,
                        "_artifact_type": "xlsx",
                    },
                    entity_keys=[],
                    source_refs=[_rtr_src],
                    receipts=[],
                    authority_class=AuthorityClass.contractual_scope,
                    confidence=0.80,
                    confidence_raw=0.80,
                    calibrated_confidence=0.80,
                    review_status=ReviewStatus.auto_accepted,
                    review_flags=[],
                    parser_version=self.parser_version,
                ))

            if rk in {"subtotal"}:
                atoms.extend(
                    self._emit_subtotal_row(
                        project_id=project_id,
                        artifact_id=artifact_id,
                        artifact_type=artifact_type,
                        filename=filename,
                        sheet_name=sheet_name,
                        row_number=row_idx + 1,
                        model=model,
                        extracted=extracted,
                        row=row,
                        label_indices=label_indices,
                    )
                )
                continue

            if rk in {"total", "grand_total"}:
                atoms.extend(
                    self._emit_total_row(
                        project_id=project_id,
                        artifact_id=artifact_id,
                        artifact_type=artifact_type,
                        filename=filename,
                        sheet_name=sheet_name,
                        row_number=row_idx + 1,
                        model=model,
                        row=row,
                        label_indices=label_indices,
                    )
                )
                continue

            if rk == "line_item":
                if _is_blank_row(list(extracted.values())) and not model.wide_qty_columns:
                    continue
                atoms.extend(
                    self._emit_line_item_row(
                        project_id=project_id,
                        artifact_id=artifact_id,
                        artifact_type=artifact_type,
                        filename=filename,
                        sheet_name=sheet_name,
                        row_number=row_idx + 1,
                        model=model,
                        extracted=extracted,
                        row=row,
                        label_indices=label_indices,
                    )
                )

        # Universal fallback: when the header-mapped path emitted
        # nothing (sheet had headers but they didn't match any
        # canonical extractor profile — Pricing sheets, summary
        # tables, etc.), fall back to the generic row emitter so the
        # rows aren't silently dropped.
        if not atoms:
            return self._emit_generic_rows(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                rows=rows,
            )
        return atoms

    @staticmethod
    def _canonicalize_header(value: str) -> str:
        """Map a raw column header into a stable canonical key.

        Used by :meth:`_generic_row_profile` to bucket a row into one
        of the structured AtomTypes (risk / asset_record / …) without
        requiring the upstream workbook to use a specific column name.
        """
        raw = normalize_text(str(value or "")).replace("-", " ").replace("_", " ")
        raw = re.sub(r"\s+", " ", raw).strip()

        aliases: dict[str, set[str]] = {
            "site_id": {"site id", "site code", "campus id", "location id"},
            "site_name": {"site", "site name", "campus", "location", "building"},
            "address": {
                "address",
                "street address",
                "service address",
                "site address",
            },
            "access_notes": {"access notes", "access window", "site access"},

            "risk_id": {"risk id", "raid id", "issue id"},
            "severity": {"severity", "priority", "risk rating"},
            "impact": {"impact", "business impact"},
            "likelihood": {"likelihood", "probability"},
            "mitigation": {"mitigation", "mitigation plan", "response plan"},
            "owner": {"owner", "assigned owner", "risk owner"},
            "status": {"status", "raid status"},

            "asset_id": {"asset id", "asset tag", "tag"},
            "serial": {"serial", "serial number", "s n", "sn"},
            "model": {"model", "model number", "part number"},
            "ip_address": {"ip", "ip address", "management ip", "mgmt ip"},
            "mac_address": {"mac", "mac address"},
            "lifecycle": {"lifecycle", "eol", "end of life", "refresh status"},

            "support_level": {"support level", "support tier", "coverage"},
            "contract_id": {
                "contract",
                "contract id",
                "support contract",
                "contract number",
            },
            "renewal_date": {
                "renewal",
                "renewal date",
                "expiration",
                "expiry",
                "expires",
            },
            "support_notes": {
                "support notes",
                "coverage notes",
                "entitlement notes",
            },

            "wan_provider": {
                "wan provider",
                "carrier",
                "isp",
                "circuit provider",
            },
            "circuit_id": {"circuit id", "circuit", "service id"},

            # PR2 (post-v3 review) — operational-workbook columns.
            "asset_type": {"asset type", "device type", "type"},
            "vendor": {"vendor", "manufacturer", "mfr", "make"},
            "manufacturer": {"manufacturer", "mfr", "make"},
            "hostname": {"hostname", "host name", "device name"},
            "mdf": {"mdf", "mdf id", "mdf area", "mdf/area"},
            "idf": {"idf", "idf id"},
            "in_service_date": {
                "in service",
                "in service date",
                "in-service date",
                "deployed",
                "deployment date",
            },
            "refresh_target": {
                "refresh target",
                "refresh date",
                "refresh planned",
            },
            "lifecycle_status": {
                "lifecycle status",
                "lifecycle state",
                "refresh status",
            },
            # NOC alert matrix
            "monitor_id": {"monitor id", "alert id"},
            "runbook_ref": {"runbook ref", "runbook", "playbook ref"},
            "alert_severity": {"alert severity", "alert level"},
            # Cutover validation
            "validation_id": {"validation id", "test id", "check id"},
            "customer_signoff": {"customer signoff", "customer sign off", "signoff"},
            "pass_flag": {"pass flag", "pass/fail", "result"},
            # Port / VLAN
            "switch_hostname": {"switch hostname", "switch name", "switch"},
            "port": {"port", "switch port", "interface"},
            "vlan_id": {"vlan id", "vlan"},
            "patch_panel_port": {"patch panel port", "panel port", "patch port"},
            # BOM detail
            "scope_bucket": {"scope bucket", "bucket", "scope phase"},
            "category": {"category", "line category"},
            "sku": {"sku", "part number", "part"},
            "unit_cost": {"unit cost", "unit price", "cost each"},
            "extended_cost": {"extended cost", "extended price", "ext cost"},
            "quote_status": {"quote status", "status quote", "po status"},
            "procurement_constraint": {
                "procurement constraint",
                "procurement note",
            },
            "distributor": {"distributor", "supplier", "wholesaler"},
            "quote_ref": {"quote ref", "quote number", "quote id"},
            "bom_line": {"bom line", "line number", "line"},
        }

        for canon, names in aliases.items():
            if raw in names:
                return canon
        return raw.replace(" ", "_")

    @staticmethod
    def _generic_row_profile(
        canon_cells: dict[str, str],
        sheet_name: str,
    ) -> tuple[AtomType, str, AuthorityClass, float]:
        """Inspect canonicalized cells + sheet name and pick the best
        structured AtomType / value.kind / authority class / confidence
        for this row. Defaults to ``scope_item`` / table_row / 0.84
        which preserves the legacy behavior."""
        keys = {k for k, v in canon_cells.items() if str(v).strip()}
        sheet = normalize_text(sheet_name)

        # Order matters — most specific first. Lifecycle is more
        # specific than the bare asset_record because it requires both
        # an asset identifier AND a lifecycle/status column.
        if {"risk_id", "severity"} & keys and (
            {"mitigation", "impact", "owner", "status", "likelihood"} & keys
        ):
            return (
                AtomType.risk,
                "risk_register_row",
                AuthorityClass.customer_current_authored,
                0.93,
            )
        if {"support_level", "renewal_date", "contract_id"} & keys:
            return (
                AtomType.support_entitlement,
                "support_entitlement_row",
                AuthorityClass.vendor_quote,
                0.92,
            )
        # PR2 (post-v3 review) — asset_record beats lifecycle_status
        # when the row has BOTH an asset identifier AND a vendor /
        # model / manufacturer / site signal. Pure lifecycle/EOL
        # records without a vendor/model still classify as
        # lifecycle_status.
        if (
            {"asset_id", "serial", "ip_address", "mac_address", "hostname"} & keys
            and (
                {"model", "manufacturer", "vendor", "asset_type", "site_name", "site"}
                & keys
            )
        ):
            return (
                AtomType.asset_record,
                "asset_inventory_row",
                AuthorityClass.approved_site_roster,
                0.94,
            )
        if {"lifecycle", "lifecycle_status", "refresh_target", "status"} & keys and (
            {"asset_id", "model", "serial"} & keys
        ):
            return (
                AtomType.lifecycle_status,
                "lifecycle_status_row",
                AuthorityClass.approved_site_roster,
                0.91,
            )
        if {"asset_id", "serial", "ip_address", "mac_address"} & keys:
            return (
                AtomType.asset_record,
                "asset_inventory_row",
                AuthorityClass.approved_site_roster,
                0.93,
            )
        if {"site_id", "address"} & keys or ("site" in sheet and "address" in keys):
            return (
                AtomType.site_roster,
                "site_roster_row",
                AuthorityClass.approved_site_roster,
                0.94,
            )
        return (
            AtomType.scope_item,
            "table_row",
            AuthorityClass.contractual_scope,
            0.84,
        )

    # ───── PR2 (post-v3 review) — operational sheet emitter ─────

    def _entity_keys_from_operational_row(
        self, cells: dict[str, str]
    ) -> list[str]:
        keys: list[str] = []
        for field in ("site_name", "site", "location"):
            if cells.get(field):
                key = normalize_entity_key("site", cells[field])
                if key:
                    keys.append(key)
        for field in ("mdf", "idf"):
            if cells.get(field):
                keys.append(normalize_entity_key(field, cells[field]))
        for field in ("asset_id", "hostname", "device_name"):
            if cells.get(field):
                keys.append(normalize_entity_key("device", cells[field]))
        for field in ("vendor", "manufacturer"):
            if cells.get(field):
                keys.append(normalize_entity_key("vendor", cells[field]))
        for field in ("part_number", "sku"):
            if cells.get(field):
                keys.append(normalize_entity_key("part_number", cells[field]))
        return sorted(set(keys))

    def _emit_operational_sheet_rows(
        self,
        *,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        rows: list[list[Any]],
        profile: _OperationalSheetProfile,
    ) -> list[EvidenceAtom]:
        header_idx = _first_nonblank_header_row(rows)
        if header_idx is None:
            return []

        headers = [
            (str(c).strip() if c is not None else "") or f"col_{i + 1}"
            for i, c in enumerate(rows[header_idx])
        ]
        canonical_headers = [self._canonicalize_header(h) for h in headers]

        atoms: list[EvidenceAtom] = []

        for row_idx in range(header_idx + 1, len(rows)):
            row = rows[row_idx] or []
            if _is_blank_row(row):
                continue

            cells: dict[str, str] = {}
            canonical_cells: dict[str, str] = {}
            cell_columns: dict[str, str] = {}

            for i, header in enumerate(headers):
                if i >= len(row):
                    continue
                value = row[i]
                text = _cell_to_text_op(value)
                if not text:
                    continue
                cells[header] = text
                canonical_cells[canonical_headers[i]] = text
                cell_columns[header] = get_column_letter(i + 1)

            if not cells:
                continue

            raw_text = " | ".join(f"{k}: {v}" for k, v in cells.items())

            value: dict[str, Any] = {
                "kind": profile.kind,
                "sheet": sheet_name,
                "row": row_idx + 1,
                "cells": cells,
                "canonical_cells": canonical_cells,
            }

            # PR2 — structured value payload per row kind. The brain
            # layer can read these without re-parsing.
            if profile.atom_type is AtomType.asset_record:
                value["asset"] = {
                    k: canonical_cells.get(k)
                    for k in (
                        "asset_id", "site_id", "site_name",
                        "asset_type", "vendor", "manufacturer",
                        "model", "quantity", "mdf", "serial",
                        "ip_address", "in_service_date",
                        "refresh_target", "status",
                    )
                    if canonical_cells.get(k)
                }
            elif profile.atom_type is AtomType.port_vlan_assignment:
                value["port"] = {
                    k: canonical_cells.get(k)
                    for k in (
                        "site_id", "switch_hostname", "port",
                        "vlan_id", "patch_panel_port",
                    )
                    if canonical_cells.get(k)
                }
            elif profile.atom_type is AtomType.circuit_inventory:
                value["circuit"] = {
                    k: canonical_cells.get(k)
                    for k in (
                        "circuit_id", "wan_provider", "site_id",
                        "site_name", "lead_time",
                    )
                    if canonical_cells.get(k)
                }
            elif profile.atom_type is AtomType.alert_route:
                value["alert"] = {
                    k: canonical_cells.get(k)
                    for k in (
                        "monitor_id", "runbook_ref", "alert_severity",
                        "owner", "site_id", "site_name",
                    )
                    if canonical_cells.get(k)
                }
            elif profile.atom_type is AtomType.cutover_validation:
                value["validation"] = {
                    k: canonical_cells.get(k)
                    for k in (
                        "validation_id", "customer_signoff", "pass_flag",
                        "owner", "status", "site_id",
                    )
                    if canonical_cells.get(k)
                }
            elif profile.atom_type is AtomType.support_entitlement:
                value["entitlement"] = {
                    k: canonical_cells.get(k)
                    for k in (
                        "support_level", "contract_id", "renewal_date",
                        "support_notes", "vendor",
                    )
                    if canonical_cells.get(k)
                }

            source_ref = self._build_source_ref(
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                row_number=row_idx + 1,
                columns=cell_columns,
            )
            atom_id = stable_id(
                "atm",
                project_id,
                artifact_id,
                sheet_name,
                str(row_idx + 1),
                profile.kind,
                normalize_text(raw_text)[:160],
            )
            atom = EvidenceAtom(
                id=atom_id,
                project_id=project_id,
                artifact_id=artifact_id,
                atom_type=profile.atom_type,
                raw_text=raw_text,
                normalized_text=normalize_text(raw_text),
                value=value,
                entity_keys=self._entity_keys_from_operational_row(canonical_cells),
                source_refs=[source_ref],
                receipts=[],
                authority_class=profile.authority_class,
                confidence=profile.confidence,
                review_status=ReviewStatus.auto_accepted,
                review_flags=[],
                parser_version=self.parser_version,
            )
            atoms.append(atom)

            atoms.extend(
                self._emit_field_aware_cell_fact_atoms(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    sheet_name=sheet_name,
                    row_number=row_idx + 1,
                    row_atom_id=atom_id,
                    cells=cells,
                    canonical_cells=canonical_cells,
                    cell_columns=cell_columns,
                )
            )

        return atoms

    def _emit_field_aware_cell_fact_atoms(
        self,
        *,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        row_atom_id: str,
        cells: dict[str, str],
        canonical_cells: dict[str, str],
        cell_columns: dict[str, str],
    ) -> list[EvidenceAtom]:
        """PR3 — field-aware cell-fact emitter. Same as
        :meth:`_emit_cell_fact_atoms` but knows that cells in
        support_level / coverage / support_entitlement columns are
        SUPPORT-TIER prose, not exclusions, and that conditional
        boundary phrasing ("8x5 vendor support unless noted") gets
        its own ``conditional_support_boundary`` atom type."""
        out: list[EvidenceAtom] = []
        for original_field, text in cells.items():
            text_str = str(text).strip()
            if not text_str:
                continue
            canonical_field = self._canonicalize_header(original_field)

            if canonical_field in _SUPPORT_LEVEL_FIELDS:
                # Support-tier prose is not a contractual exclusion;
                # only emit a conditional boundary atom if the cell
                # explicitly carries conditional language.
                if _CONDITIONAL_SUPPORT_RE.search(text_str):
                    atom_type = AtomType.conditional_support_boundary
                    confidence = 0.88
                    flags = ["conditional_support_boundary"]
                else:
                    continue
            elif _CONDITIONAL_SUPPORT_RE.search(text_str):
                atom_type = AtomType.conditional_support_boundary
                confidence = 0.88
                flags = ["conditional_support_boundary"]
            elif _EXCLUSION_CELL_RE.search(text_str):
                atom_type = AtomType.exclusion
                confidence = 0.90
                flags = []
            elif _RISK_CELL_RE.search(text_str):
                atom_type = AtomType.risk
                confidence = 0.88
                flags = []
            else:
                continue

            source_ref = SourceRef(
                id=stable_id(
                    "src", artifact_id, sheet_name, row_number,
                    original_field, "cell_fact",
                ),
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                locator={
                    "sheet": sheet_name,
                    "row": row_number,
                    "columns": {original_field: cell_columns.get(original_field)},
                    "parent_row_atom_id": row_atom_id,
                },
                extraction_method="xlsx_cell_fact_v2_field_aware",
                parser_version=self.parser_version,
            )
            out.append(
                EvidenceAtom(
                    id=stable_id(
                        "atm", project_id, artifact_id, sheet_name,
                        row_number, original_field, atom_type.value, text_str,
                    ),
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=f"{original_field}: {text_str}",
                    normalized_text=normalize_text(text_str),
                    value={
                        "kind": "cell_fact",
                        "field": original_field,
                        "canonical_field": canonical_field,
                        "parent_row_atom_id": row_atom_id,
                        "text": text_str,
                    },
                    entity_keys=[],
                    source_refs=[source_ref],
                    receipts=[],
                    authority_class=AuthorityClass.customer_current_authored,
                    confidence=confidence,
                    review_status=ReviewStatus.auto_accepted,
                    review_flags=flags,
                    parser_version=self.parser_version,
                )
            )
        return out

    def _emit_cell_fact_atoms(
        self,
        *,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        row_atom_id: str,
        cells: dict[str, str],
        cell_columns: dict[str, str],
    ) -> list[EvidenceAtom]:
        """Emit cell-level sub-atoms for buried facts inside a generic row.

        Specifically: if a cell mentions an explicit exclusion phrase
        ("not included", "excluded", "outside coverage", …) emit an
        ``exclusion`` atom anchored to that cell. Same for risk
        signals ("EOL", "single point of failure", "missing", …) →
        ``risk`` atom. Each sub-atom carries
        ``value.parent_row_atom_id`` so the packetizer can group them
        with the parent row.
        """
        out: list[EvidenceAtom] = []
        for field, text in cells.items():
            text_str = str(text).strip()
            if not text_str:
                continue
            if _EXCLUSION_CELL_RE.search(text_str):
                atom_type = AtomType.exclusion
                confidence = 0.90
            elif _RISK_CELL_RE.search(text_str):
                atom_type = AtomType.risk
                confidence = 0.88
            else:
                continue

            source_ref = SourceRef(
                id=stable_id(
                    "src", artifact_id, sheet_name, row_number, field, "cell_fact"
                ),
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                locator={
                    "sheet": sheet_name,
                    "row": row_number,
                    "columns": {field: cell_columns.get(field)},
                    "parent_row_atom_id": row_atom_id,
                    "section_path": [sheet_name] if sheet_name else [],
                },
                extraction_method="xlsx_cell_fact_v1",
                parser_version=self.parser_version,
            )
            out.append(
                EvidenceAtom(
                    id=stable_id(
                        "atm",
                        project_id,
                        artifact_id,
                        sheet_name,
                        row_number,
                        field,
                        atom_type.value,
                        text_str,
                    ),
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=f"{field}: {text_str}",
                    normalized_text=normalize_text(text_str),
                    value={
                        "kind": "cell_fact",
                        "field": field,
                        "parent_row_atom_id": row_atom_id,
                        "text": text_str,
                    },
                    entity_keys=[],
                    source_refs=[source_ref],
                    receipts=[],
                    authority_class=AuthorityClass.customer_current_authored,
                    confidence=confidence,
                    review_status=ReviewStatus.auto_accepted,
                    review_flags=[],
                    parser_version=self.parser_version,
                )
            )
        return out

    def _emit_block_structured_rows(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        rows: list[list[Any]],
        styles: list[list[tuple[str | None, bool]]] | None = None,
    ) -> list[EvidenceAtom]:
        """Emit atoms for a sheet split into title/header/rows BLOCKS.

        For each table block: one ``raw_table_row`` per data row (so the schema
        classifier types it AND binds its OWN block's column headers) PLUS a
        generic ``scope_item`` fallback (kept only for rows the classifier can't
        type — the provenance-aware dedup collapses the matched ones). Key-value
        blocks emit a single ``scope_item`` of ``label: value | …`` pairs. EVERY
        atom carries ``section_path=[sheet, title]`` so the head sees which table
        under which title the fact came from. A globally-unique per-sheet row
        counter keeps each block's cells at a distinct provenance cell."""
        from app.parsers.xlsx_blocks import sheet_blocks
        try:
            blocks = sheet_blocks(rows, styles=styles)
        except Exception:
            return []
        # Take over ONLY genuinely multi-block sheets (>=2 substantive table/
        # key-value blocks) — exactly the case the single-header-per-sheet model
        # provably cannot handle (it picks one header and drops the other tables,
        # e.g. Summary's "Detailed Level of Effort" + "Key Unit Metrics"). A
        # single-table sheet stays on the tuned legacy model so its quantity /
        # money / wide-column extraction is preserved.
        substantive = [b for b in blocks if b.get("kind") in ("table", "keyval")]
        if len(substantive) < 2:
            return []

        atoms: list[EvidenceAtom] = []
        seq = 0

        # ── Merge fragmented side-annotations + attach their neighbourhood ──
        # A loose annotation often lands split across cells — "Best Buy Quoted"
        # (label) in one cell, "920 hours" (value) two columns over — so it came
        # through as two meaningless fragments. Stitch a label-only annotation to
        # the value-only one that follows it (reading order), so the atom reads as
        # one fact ("Best Buy Quoted 920 hours"). Then compute the sheet's own
        # hour/total figures ONCE: a side note benchmarking our estimate is
        # meaningless without the estimate, so every annotation carries it as
        # context (this is what lets a reader — or the reconciliation head — see
        # "competitor quote vs our number").
        _ann_idx = [i for i, b in enumerate(blocks)
                    if b.get("kind") == "text" and not b.get("title")
                    and _is_substantive_annotation(str(b.get("text") or ""))]
        _consumed_ann: set[int] = set()
        for k, i in enumerate(_ann_idx):
            txt = str(blocks[i].get("text") or "").strip()
            if not _DIGIT_RE.search(txt) and k + 1 < len(_ann_idx):
                j = _ann_idx[k + 1]
                vtxt = str(blocks[j].get("text") or "").strip()
                if _DIGIT_RE.search(vtxt) and len(vtxt) <= 18:   # value-only fragment
                    blocks[i]["text"] = f"{txt} {vtxt}"
                    _consumed_ann.add(j)
        _sheet_ctx = _sheet_hours_context(rows)

        def _section_path(title: str | None) -> list[str]:
            return [sheet_name, title] if (sheet_name and title) else ([sheet_name] if sheet_name else [])

        def _src(rid: str, sp: list[str], extraction: str) -> SourceRef:
            return SourceRef(
                id=stable_id("src", rid),
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                locator={"sheet": sheet_name, "row": seq, "section_path": sp, "extraction": extraction},
                extraction_method=extraction,
                parser_version=self.parser_version,
            )

        for bi, b in enumerate(blocks):
            sp = _section_path(b.get("title"))
            if b["kind"] == "table":
                header = b["header"]
                for row_cells in b["rows"]:
                    seq += 1
                    # Summary / total rows ("Subtotal", "Recommended fixed fee
                    # hours", "Safer bid hours", "Grand Total") are NOT task rows —
                    # don't force the first column's header ("Task Category") onto
                    # them. Render the row label bare, then the remaining columns
                    # header-bound, so it reads "Subtotal | Labor Hours: 458.5".
                    first = (row_cells[0] if row_cells else "").strip()
                    is_total = bool(re.match(
                        r"^(sub-?\s*totals?|totals?|grand total|safer bid\b|.*\bfixed fee\b)",
                        first, re.I)) and (len([c for c in row_cells[1:] if c]) <= 3)
                    if is_total:
                        pairs = [first] + [
                            f"{header[j]}: {row_cells[j]}"
                            for j in range(1, min(len(header), len(row_cells)))
                            if row_cells[j] != ""
                        ]
                    else:
                        pairs = [f"{header[j]}: {row_cells[j]}" for j in range(min(len(header), len(row_cells))) if row_cells[j] != ""]
                    for j in range(len(header), len(row_cells)):
                        if row_cells[j] != "":
                            pairs.append(f"col{j+1}: {row_cells[j]}")
                    if not pairs:
                        continue
                    body = " | ".join(pairs)[:4000]
                    # raw_table_row -> schema classifier types it + binds headers
                    rtr_id = stable_id("atm", artifact_id, "xlsx_block_rtr", sheet_name, bi, seq)
                    atoms.append(EvidenceAtom(
                        id=rtr_id, project_id=project_id, artifact_id=artifact_id,
                        atom_type=AtomType.raw_table_row, raw_text=body, normalized_text=body.lower(),
                        value={"_columns": list(header), "_row": list(row_cells), "_table_idx": bi,
                               "_row_idx": seq, "_filename": filename, "_sheet": sheet_name,
                               "section_path": sp, "_artifact_type": "xlsx"},
                        entity_keys=[], source_refs=[_src(rtr_id, sp, "xlsx_block_raw_table_row")], receipts=[],
                        authority_class=AuthorityClass.contractual_scope,
                        confidence=0.80, confidence_raw=0.80, calibrated_confidence=0.80,
                        review_status=ReviewStatus.auto_accepted, review_flags=[],
                        parser_version=self.parser_version,
                    ))
                    # generic fallback (survives only if the classifier can't type the row)
                    si_id = stable_id("atm", artifact_id, "xlsx_block_row", sheet_name, bi, seq)
                    atoms.append(EvidenceAtom(
                        id=si_id, project_id=project_id, artifact_id=artifact_id,
                        atom_type=AtomType.scope_item, raw_text=body, normalized_text=body.lower(),
                        value={"kind": "table_row", "columns": list(header),
                               "cells": {header[j]: row_cells[j] for j in range(min(len(header), len(row_cells))) if row_cells[j] != ""}},
                        entity_keys=[], source_refs=[_src(si_id, sp, "xlsx_block_row_v1")], receipts=[],
                        authority_class=AuthorityClass.contractual_scope,
                        confidence=0.78, confidence_raw=0.78, calibrated_confidence=0.78,
                        review_status=ReviewStatus.auto_accepted, review_flags=[],
                        parser_version=self.parser_version,
                    ))
            elif b["kind"] == "keyval":
                # ONE atom per label:value pair — individual atoms (row=atom),
                # NOT a single glued "a: 1 | b: 2 | c: 3" blob. The pairs stay
                # grouped two ways: they share this block's section_path (the
                # box's title, real or synthesized) AND each carries
                # group_index/group_size so a head knows it is the i-th of N
                # rows in one box. ``box_label`` repeats the section title on
                # every atom so the grouping survives even where the renderer
                # flattens section_path.
                pairs = [(str(k).strip(), str(v).strip()) for k, v in b["pairs"]]
                pairs = [(k, v) for k, v in pairs if k or v]
                n = len(pairs)
                box_label = sp[-1] if len(sp) > 1 else None
                box_fill = b.get("fill")  # the highlight color the rows share
                for pi, (k, v) in enumerate(pairs):
                    seq += 1
                    line = (f"{k}: {v}" if v else k).strip()
                    if not line:
                        continue
                    # A numeric LABEL ('70: 0.3684') is not a field — it's two
                    # stray cells from an unlabeled totals row; and a label-less
                    # bare number ('1.15') is a loose cell, not a fact. Drop both
                    # (the real BOM rows in the same sheet are kept as table rows).
                    if (_is_bare_numeric(k) and _is_bare_numeric(v)) or (not v and _is_bare_numeric(k)):
                        continue
                    kv_id = stable_id("atm", artifact_id, "xlsx_block_kv", sheet_name, bi, seq)
                    atoms.append(EvidenceAtom(
                        id=kv_id, project_id=project_id, artifact_id=artifact_id,
                        atom_type=AtomType.scope_item, raw_text=line[:4000],
                        normalized_text=line[:4000].lower(),
                        value={"kind": "key_value", "label": k, "value": v,
                               "group_index": pi, "group_size": n,
                               "box_label": box_label, "box_fill": box_fill},
                        entity_keys=[], source_refs=[_src(kv_id, sp, "xlsx_block_keyval_v1")], receipts=[],
                        authority_class=AuthorityClass.contractual_scope,
                        confidence=0.78, confidence_raw=0.78, calibrated_confidence=0.78,
                        review_status=ReviewStatus.auto_accepted, review_flags=[],
                        parser_version=self.parser_version,
                    ))
            elif b["kind"] == "text":
                # A free-text paragraph block — emitted only when it carries a
                # title (a styled banner header above it, e.g. "Customer Facing
                # Quote Language" + its allowance paragraph). Without a title a
                # bare text block is a stray caption with no home, so it stays
                # dropped; WITH one it is real titled content and must surface as
                # a scope_item under that section, not vanish.
                txt = str(b.get("text") or "").strip()
                if txt and b.get("title"):
                    seq += 1
                    tx_id = stable_id("atm", artifact_id, "xlsx_block_text", sheet_name, bi, seq)
                    atoms.append(EvidenceAtom(
                        id=tx_id, project_id=project_id, artifact_id=artifact_id,
                        atom_type=AtomType.scope_item, raw_text=txt[:4000],
                        normalized_text=txt[:4000].lower(),
                        value={"kind": "section_text", "box_label": (sp[-1] if len(sp) > 1 else None)},
                        entity_keys=[], source_refs=[_src(tx_id, sp, "xlsx_block_text_v1")], receipts=[],
                        authority_class=AuthorityClass.contractual_scope,
                        confidence=0.78, confidence_raw=0.78, calibrated_confidence=0.78,
                        review_status=ReviewStatus.auto_accepted, review_flags=[],
                        parser_version=self.parser_version,
                    ))
                elif bi in _consumed_ann:
                    continue   # value fragment already stitched onto its label
                elif txt and _is_substantive_annotation(txt):
                    # Titleless free-text that still carries a FACT — a side note
                    # like "Best Buy quoted 950 hours" / "Best Buy Quoted 920 hours"
                    # beside an estimate table (competitive-quote intel). Dropping it
                    # is silent data loss; capture it as a low-confidence loose
                    # annotation, flagged for review. Attach the sheet's own hour
                    # figures as neighbour context so the note is legible on its own
                    # (their 920 hrs vs our estimate) and the reconciliation head
                    # has the number to compare against.
                    seq += 1
                    an_id = stable_id("atm", artifact_id, "xlsx_block_note", sheet_name, bi, seq)
                    disp = f"{txt}  [annotation on '{sheet_name}'; sheet hours: {_sheet_ctx}]" if _sheet_ctx else txt
                    atoms.append(EvidenceAtom(
                        id=an_id, project_id=project_id, artifact_id=artifact_id,
                        atom_type=AtomType.scope_item, raw_text=disp[:4000],
                        normalized_text=disp[:4000].lower(),
                        value={"kind": "sheet_annotation", "annotation": txt,
                               "neighbor_context": _sheet_ctx, "sheet_name": sheet_name,
                               "box_label": (sp[-1] if len(sp) > 1 else None)},
                        entity_keys=[], source_refs=[_src(an_id, sp, "xlsx_block_note_v1")], receipts=[],
                        authority_class=AuthorityClass.contractual_scope,
                        confidence=0.55, confidence_raw=0.55, calibrated_confidence=0.55,
                        review_status=ReviewStatus.needs_review,
                        review_flags=["xlsx_parser:loose_annotation"],
                        parser_version=self.parser_version,
                    ))
        return atoms

    def _emit_generic_rows(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        rows: list[list[Any]],
    ) -> list[EvidenceAtom]:
        """Generic fallback: emit one atom per non-empty row.

        Used when ``_detect_header`` can't find a canonical
        cabling/networking schedule.  Treats the first non-empty row
        with mostly text cells as the column-header row (so cell
        values get keyed by column name); falls back to ``col_N``
        labels otherwise.

        Each row becomes a ``scope_item`` atom with ``raw_text`` set
        to ``"col_a: value | col_b: value | ..."`` so the
        ``entity_extraction`` stage can pull entity keys out of the
        text just like for any other atom.

        See PRODUCTION_GAPS.md P0.4.
        """
        # Universal false-positive guard: workbook tabs whose only purpose
        # is to instruct the human filling in the form (e.g. "Instructions",
        # "Read Me", "Cover", "About") are not data. They reach this code
        # path because their prose can't satisfy the canonical header
        # detector. If the sheet name signals an instructional intent and
        # the body has no detectable schedule structure, emit zero atoms
        # rather than fabricating ``scope_item`` atoms from cover text.
        if _looks_instructional_sheet(sheet_name) and not _has_tabular_shape(rows):
            return []

        atoms: list[EvidenceAtom] = []

        # Step 1: pick a header row.  We scan the first 10 non-empty rows
        # and take the one with the most non-numeric, short string cells.
        # If no row qualifies we use ``col_N`` labels.
        header_idx = -1
        header_score = -1
        for idx, row in enumerate(rows[:10]):
            if _is_blank_row(row):
                continue
            non_empty = [c for c in row if str(c or "").strip()]
            if not non_empty:
                continue
            string_cells = sum(
                1 for c in non_empty
                if isinstance(c, str) and len(c) <= 60 and not c.strip().isdigit()
            )
            score = string_cells - max(0, len(non_empty) - string_cells)
            if score >= 2 and score > header_score:
                header_score = score
                header_idx = idx

        if header_idx >= 0:
            header_row = rows[header_idx]
            columns = [
                (str(c).strip() if c is not None else "") or f"col_{i + 1}"
                for i, c in enumerate(header_row)
            ]
            data_start = header_idx + 1
        else:
            # No detectable header — use col_1, col_2, ... and start
            # from row 0 so we don't lose any rows.
            width = max((len(r) for r in rows), default=0)
            columns = [f"col_{i + 1}" for i in range(width)]
            data_start = 0

        # Step 2: emit one atom per non-empty data row.
        for row_idx in range(data_start, len(rows)):
            row = rows[row_idx] or []
            if _is_blank_row(row):
                continue
            # Render the row as "col: value | col: value" — same shape
            # the canonical line-item emitter uses, so OrbitBrief and
            # entity_extraction get a familiar structure.
            bare: list[str] = []
            cell_columns: dict[str, str] = {}
            for col_idx, col_name in enumerate(columns):
                if col_idx >= len(row):
                    break
                value = row[col_idx]
                if value is None:
                    continue
                value_str = str(value).strip()
                if not value_str:
                    continue
                bare.append(value_str)
                cell_columns[col_name] = get_column_letter(col_idx + 1)
            # raw_text is the bare pipe-join of cell VALUES (not "col: value"),
            # matching the docx per-row blob and the schema row_text — so when a
            # raw_table_row routes this same row to a typed atom (deal_metadata,
            # etc.), cross_type_dedup collapses this scope twin into it. Column
            # meaning is preserved in value.cells and rendered back by
            # _atom_bound_text at decide-time, so nothing is lost for display.
            raw_text = " | ".join(bare).strip()
            if not raw_text:
                continue
            # Skip the header row itself so it doesn't reappear as data.
            if row_idx == header_idx:
                continue

            source_ref = self._build_source_ref(
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                row_number=row_idx + 1,
                columns=cell_columns,
            )
            atom_id = stable_id(
                "atm",
                artifact_id,
                sheet_name,
                str(row_idx + 1),
                normalize_text(raw_text)[:120],
            )
            cells_by_original_col = {
                col: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                for i, col in enumerate(columns)
                if i < len(row) and str(row[i] or "").strip()
            }
            # Profile the row: risk register, asset inventory, support
            # entitlement, site roster, lifecycle status, or generic
            # scope item.  Falls back to the legacy
            # (scope_item, table_row, contractual_scope, 0.84) tuple
            # when no profile fits.
            canonical_columns = [self._canonicalize_header(c) for c in columns]
            canon_cells = {
                canonical_columns[i]: str(row[i]).strip()
                for i in range(min(len(row), len(canonical_columns)))
                if row[i] is not None and str(row[i]).strip()
            }
            atom_type, row_kind, authority_class, confidence = (
                self._generic_row_profile(canon_cells, sheet_name)
            )
            value: dict[str, Any] = {
                "kind": row_kind,
                "sheet": sheet_name,
                "row": row_idx + 1,
                "cells": cells_by_original_col,
                "canonical_cells": canon_cells,
            }
            atoms.append(
                EvidenceAtom(
                    id=atom_id,
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=raw_text,
                    normalized_text=normalize_text(raw_text),
                    value=value,
                    entity_keys=[],  # populated by core.entity_extraction
                    source_refs=[source_ref],
                    receipts=[],
                    authority_class=authority_class,
                    confidence=confidence,
                    review_status=ReviewStatus.auto_accepted,
                    review_flags=[],
                    parser_version=self.parser_version,
                )
            )

            # v50: also emit a raw_table_row so the central schema registry can
            # TYPE this row (Field/Value -> deal_metadata, BOM -> bom_line, etc.)
            # — the same path the canonical-header and quote emitters already
            # use. Only when a real header row was found (col_N placeholders
            # can't match a schema). The scope_item above stays as the fail-open
            # fallback for rows that match no schema; when one DOES match,
            # cross_type_dedup collapses the scope twin into the typed atom
            # (shared (sheet,row) cell key + identical bare-pipe raw_text).
            if header_idx >= 0:
                _rtr_id = stable_id("atm", artifact_id, "raw_table_row", sheet_name, row_idx)
                _rtr_src = SourceRef(
                    id=stable_id("src", _rtr_id),
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    locator={
                        "sheet": sheet_name,
                        "row": row_idx + 1,
                        "columns": cell_columns,
                        "extraction": "raw_table_row_v49_2",
                        "section_path": [sheet_name] if sheet_name else [],
                    },
                    extraction_method="raw_table_row_v49_2",
                    parser_version=self.parser_version,
                )
                atoms.append(
                    EvidenceAtom(
                        id=_rtr_id,
                        project_id=project_id,
                        artifact_id=artifact_id,
                        atom_type=AtomType.raw_table_row,
                        raw_text=raw_text[:4000],
                        normalized_text=raw_text.lower()[:4000],
                        value={
                            "_columns": list(columns),
                            "_row": [str(c).strip() if c is not None else "" for c in row],
                            "_table_idx": 0,
                            # 1-based to match the scope_item's locator row
                            # (row_idx+1), so the typed atom shares the scope
                            # twin's (sheet,row) cell key and they collapse.
                            "_row_idx": row_idx + 1,
                            "_filename": filename,
                            "_sheet": sheet_name,
                            "_artifact_type": "xlsx",
                        },
                        entity_keys=[],
                        source_refs=[_rtr_src],
                        receipts=[],
                        authority_class=AuthorityClass.contractual_scope,
                        confidence=0.80,
                        confidence_raw=0.80,
                        calibrated_confidence=0.80,
                        review_status=ReviewStatus.auto_accepted,
                        review_flags=[],
                        parser_version=self.parser_version,
                    )
                )

            # Cell-level sub-atoms — emit exclusion / risk atoms for
            # any cells whose text matches the cell-fact patterns.
            atoms.extend(
                self._emit_cell_fact_atoms(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    sheet_name=sheet_name,
                    row_number=row_idx + 1,
                    row_atom_id=atom_id,
                    cells=cells_by_original_col,
                    cell_columns=cell_columns,
                )
            )

            # Week 6 P6.7: when this row is a Q&A pair (the columns
            # include both a question-column and a response-column),
            # also emit a *separate* atom for the question and the
            # response.  Keeps the row-level atom for context but lets
            # the packetizer surface the Q as ``open_question`` and the
            # A as ``customer_instruction`` independently — that's what
            # closes the XLSX_RARE atom_count gap (486 Q&A rows in
            # CalSAWS were producing 486 atoms instead of ~1500).
            for sub_atom in self._emit_qa_row_subatoms(
                project_id=project_id,
                artifact_id=artifact_id,
                artifact_type=artifact_type,
                filename=filename,
                sheet_name=sheet_name,
                columns=columns,
                row=row,
                row_idx=row_idx,
                cell_columns=cell_columns,
            ):
                atoms.append(sub_atom)
        return atoms

    # Names that mark a column as the *question* side of a Q&A row.
    _XLSX_QUESTION_COL_HINTS = (
        "question",
        "concern",
        "inquiry",
        "issue",
        "comment",
        "ask",
    )
    # Names that mark a column as the *response/answer* side.
    _XLSX_RESPONSE_COL_HINTS = (
        "response",
        "answer",
        "reply",
        "resolution",
        "clarification",
    )

    def _emit_qa_row_subatoms(
        self,
        *,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        columns: list[str],
        row: list[Any],
        row_idx: int,
        cell_columns: dict[str, str],
    ) -> list[EvidenceAtom]:
        """If this row has Q/A columns, emit one atom per side."""
        q_col_idx = a_col_idx = None
        for i, col in enumerate(columns):
            col_lower = col.lower()
            if q_col_idx is None and any(h in col_lower for h in self._XLSX_QUESTION_COL_HINTS):
                q_col_idx = i
            if a_col_idx is None and any(h in col_lower for h in self._XLSX_RESPONSE_COL_HINTS):
                a_col_idx = i
        # Need both sides to be a real Q&A row.
        if q_col_idx is None or a_col_idx is None or q_col_idx == a_col_idx:
            return []

        def _cell(idx: int | None) -> str:
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            if v is None:
                return ""
            if hasattr(v, "isoformat"):              # date/datetime -> date-only at midnight
                try:
                    s = v.isoformat()
                    return s[:10] if "00:00:00" in s else s
                except Exception:
                    pass
            return str(v).strip()

        q_val, a_val = _cell(q_col_idx), _cell(a_col_idx)
        if not q_val:
            return []
        # An ANSWERED questionnaire row is ONE field->value fact (the head types
        # it: deal_metadata / site_attribute / site_infrastructure / ...), kept
        # together so neither half is an orphan. An UNANSWERED row is a genuine
        # open question to put back to the customer. (Short answers like "NYC"
        # are real values — no length filter.)
        if a_val:
            sub_text = f"{q_val}: {a_val}"
            kind, atom_type, authority = "qa_pair", AtomType.customer_instruction, AuthorityClass.customer_current_authored
            anchor_col = columns[a_col_idx]
        else:
            sub_text = f"Question: {q_val}"
            kind, atom_type, authority = "qa_question", AtomType.open_question, AuthorityClass.contractual_scope
            anchor_col = columns[q_col_idx]
        sub_source_ref = self._build_source_ref(
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            sheet_name=sheet_name,
            row_number=row_idx + 1,
            columns={anchor_col: cell_columns.get(anchor_col, "")},
        )
        sub_atom_id = stable_id(
            "atm", artifact_id, sheet_name, str(row_idx + 1), kind, normalize_text(sub_text)[:120],
        )
        return [
            EvidenceAtom(
                id=sub_atom_id,
                project_id=project_id,
                artifact_id=artifact_id,
                atom_type=atom_type,
                raw_text=sub_text,
                normalized_text=normalize_text(sub_text),
                value={
                    "kind": kind,
                    "question": q_val,
                    "answer": a_val,
                    "sheet": sheet_name,
                    "row": row_idx + 1,
                },
                entity_keys=[],
                source_refs=[sub_source_ref],
                receipts=[],
                authority_class=authority,
                confidence=0.85,
                review_status=ReviewStatus.auto_accepted,
                review_flags=[],
                parser_version=self.parser_version,
            )
        ]

    def _extract_row_values(self, row: list[Any], header_map: dict[str, int]) -> dict[str, str]:
        extracted: dict[str, str] = {}
        for key, idx in header_map.items():
            value = row[idx] if idx < len(row) else ""
            extracted[key] = str(value).strip() if value is not None else ""
        return extracted

    def _entity_keys_from_extracted(self, extracted: dict[str, str]) -> list[str]:
        keys: list[str] = []
        mapping = [
            ("site", "site"),
            ("building", "building"),
            ("floor", "floor"),
            ("room", "room"),
            ("area", "area"),
            ("zone", "zone"),
            ("location", "location"),
            ("plate_id", "plate"),
            ("outlet_id", "location"),
            ("drop_id", "location"),
            ("mdf", "mdf"),
            ("idf", "idf"),
        ]
        for field, etype in mapping:
            v = extracted.get(field, "").strip()
            if v:
                key = normalize_entity_key(etype, v)
                if key:
                    keys.append(key)
        # Device vs service classification — BOM rows describing labor,
        # training, hypercare, project management, etc. are SERVICE line
        # items, not devices. Routing them to `service:` instead of
        # `device:` keeps the device namespace clean (only physical
        # hardware ends up under `device:`).
        device_value = extracted.get("device", "").strip()
        if device_value:
            etype = "service" if _looks_like_service_line(device_value) else "device"
            key = normalize_entity_key(etype, device_value)
            if key:
                keys.append(key)
        return keys

    def _emit_subtotal_row(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        model: SheetParseModel,
        extracted: dict[str, str],
        row: list[Any],
        label_indices: list[int],
    ) -> list[EvidenceAtom]:
        del extracted, row, label_indices
        # Subtotal: no entity/quantity atoms (avoid double-count with grand total).
        _ = model.diagnostics
        return []

    def _emit_total_row(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        model: SheetParseModel,
        row: list[Any],
        label_indices: list[int],
    ) -> list[EvidenceAtom]:
        atoms: list[EvidenceAtom] = []
        label_col = label_indices[0] if label_indices else 0
        label_letter = get_column_letter(label_col + 1)

        if model.wide_qty_columns:
            for wcol in model.wide_qty_columns:
                idx = wcol["col_idx"]
                if idx >= len(row):
                    continue
                parsed = _parse_schedule_quantity_cell(row[idx])
                if parsed.get("quantity") is None and parsed.get("quantity_status") not in {"zero"}:
                    continue
                qty = parsed.get("quantity")
                if qty is None and parsed.get("quantity_status") == "zero":
                    qty = 0
                if qty is None:
                    continue
                qcol = get_column_letter(idx + 1)
                columns = {"total_label": label_letter, "quantity": qcol}
                source_ref = self._build_source_ref(artifact_id, artifact_type, filename, sheet_name, row_number, columns)
                entity_keys: list[str] = []
                hint = wcol.get("entity_hint")
                if hint:
                    entity_keys.append(normalize_entity_key(hint[0], hint[1]))
                value = {
                    **parsed,
                    "item": wcol["item"],
                    "normalized_item": wcol["normalized_item"],
                    "item_kind": wcol.get("item_kind"),
                    "material_family": wcol.get("material_family"),
                    "cable_category": wcol.get("cable_category"),
                    "shielding": wcol.get("shielding"),
                    "source_row_type": "total",
                    "aggregate": True,
                }
                value = merge_parser_value_identity(value, raw_text=f"Total {wcol['item']} {qty}")
                atoms.append(
                    EvidenceAtom(
                        id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, "qty_total", wcol["normalized_item"], str(qty)),
                        project_id=project_id,
                        artifact_id=artifact_id,
                        atom_type=AtomType.quantity,
                        raw_text=f"Total {wcol['item']} {qty}",
                        normalized_text=normalize_text(f"total {wcol['item']} {qty}"),
                        value=value,
                        entity_keys=entity_keys,
                        source_refs=[source_ref],
                        authority_class=AuthorityClass.approved_site_roster,
                        confidence=0.94,
                        review_status=ReviewStatus.auto_accepted,
                        review_flags=["xlsx_parser:aggregate_total"],
                        parser_version=self.parser_version,
                    )
                )
            return atoms

        # Traditional single quantity column total row
        qidx = model.header_map.get("quantity")
        if qidx is None or qidx >= len(row):
            return atoms
        parsed = _parse_schedule_quantity_cell(row[qidx])
        qty = parsed.get("quantity")
        if qty is None and parsed.get("quantity_status") != "zero":
            return atoms
        if qty is None and parsed.get("quantity_status") == "zero":
            qty = 0
        columns = {"total_label": label_letter, "quantity": get_column_letter(qidx + 1)}
        source_ref = self._build_source_ref(artifact_id, artifact_type, filename, sheet_name, row_number, columns)
        value = {**parsed, "source_row_type": "total", "aggregate": True, "item": "total", "normalized_item": "total"}
        value = merge_parser_value_identity(value, raw_text=f"Total quantity {qty}")
        atoms.append(
            EvidenceAtom(
                id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, "qty_total_single", str(qty)),
                project_id=project_id,
                artifact_id=artifact_id,
                atom_type=AtomType.quantity,
                raw_text=f"Total quantity {qty}",
                normalized_text=normalize_text(f"total quantity {qty}"),
                value=value,
                entity_keys=[],
                source_refs=[source_ref],
                authority_class=AuthorityClass.approved_site_roster,
                confidence=0.92,
                review_status=ReviewStatus.auto_accepted,
                review_flags=["xlsx_parser:aggregate_total"],
                parser_version=self.parser_version,
            )
        )
        return atoms

    def _emit_line_item_row(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        model: SheetParseModel,
        extracted: dict[str, str],
        row: list[Any],
        label_indices: list[int],
    ) -> list[EvidenceAtom]:
        atoms: list[EvidenceAtom] = []
        site = extracted.get("site", "").strip()
        device = extracted.get("device", "").strip()
        floor = extracted.get("floor", "").strip()
        room = extracted.get("room", "").strip()
        location = extracted.get("location", "").strip() or extracted.get("description", "").strip()
        scope = extracted.get("scope", "").strip()
        access = extracted.get("access", "").strip() or extracted.get("access_window", "").strip()
        plate = extracted.get("plate_id", "").strip()
        notes_blob = _row_text_blob(row, model.header_map)

        # Do not treat "total" in notes as skipping — row_kind already line_item.
        label_text = _label_text_for_row(row, label_indices)
        if _is_total_label(label_text):
            return []

        entity_keys = self._entity_keys_from_extracted(extracted)
        major = sum(1 for k in ("site", "device", "location", "plate_id", "room") if extracted.get(k))
        row_confidence = 0.92 if major >= 2 else (0.88 if major == 1 else 0.78)

        def append_atom(
            atom_type: AtomType,
            raw_text: str,
            value: dict[str, Any],
            confidence: float,
            *,
            entity_keys_out: list[str] | None = None,
            review_status: ReviewStatus = ReviewStatus.auto_accepted,
            review_flags: list[str] | None = None,
            extra_columns: dict[str, str] | None = None,
        ) -> None:
            cols = dict(extra_columns or {})
            ek = entity_keys_out if entity_keys_out is not None else entity_keys
            atoms.append(
                EvidenceAtom(
                    id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, atom_type.value, raw_text),
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=raw_text,
                    normalized_text=normalize_text(raw_text),
                    value=value,
                    entity_keys=ek,
                    source_refs=[
                        self._build_source_ref(artifact_id, artifact_type, filename, sheet_name, row_number, cols)
                    ],
                    authority_class=AuthorityClass.approved_site_roster,
                    confidence=confidence,
                    review_status=review_status,
                    review_flags=review_flags or [],
                    parser_version=self.parser_version,
                )
            )

        # Entities: precise types, not everything as site. These are per-row
        # SUB-FIELD extractions (the Site/Floor/Room column already lives on the
        # row atom) used for entity resolution — emit them only when not in
        # one-atom-per-row mode, same rule as site_allocation / dependency.
        for field, etype, label in () if os.environ.get(
            "SOWSMITH_DROP_DERIVED_SUBATOMS"
        ) == "1" else (
            ("site", "site", "Site"),
            ("building", "building", "Building"),
            ("floor", "floor", "Floor"),
            ("room", "room", "Room"),
            ("area", "area", "Area"),
            ("zone", "zone", "Zone"),
            ("location", "location", "Location"),
            ("plate_id", "plate", "Plate"),
            ("mdf", "mdf", "MDF"),
            ("idf", "idf", "IDF"),
            ("device", "device", "Device"),
        ):
            val = extracted.get(field, "").strip()
            if not val:
                continue
            idx = model.header_map.get(field)
            letter = get_column_letter((idx or 0) + 1) if idx is not None else "A"
            single_key = [normalize_entity_key(etype, val)]
            append_atom(
                AtomType.entity,
                f"{label} {val}",
                {"entity_type": etype, "name": val},
                row_confidence,
                entity_keys_out=single_key,
                extra_columns={field: letter},
            )

        if model.wide_qty_columns:
            ctx_cols: dict[str, str] = {}
            for key in ("plate_id", "location", "room", "site", "description"):
                idx = model.header_map.get(key)
                if idx is not None and idx < len(row):
                    ctx_cols[key] = get_column_letter(idx + 1)
            for wcol in model.wide_qty_columns:
                idx = wcol["col_idx"]
                if idx >= len(row):
                    continue
                parsed = _parse_schedule_quantity_cell(row[idx])
                if parsed.get("quantity") is None and parsed.get("quantity_status") not in {"zero", "included", "allowance", "range"}:
                    if str(row[idx] or "").strip() == "":
                        continue
                qty_val = parsed.get("quantity")
                qcol = get_column_letter(idx + 1)
                cols = {**ctx_cols, "quantity": qcol}
                hint = wcol.get("entity_hint")
                hint_keys = [normalize_entity_key(hint[0], hint[1])] if hint else []
                q_entity_keys = list(dict.fromkeys(entity_keys + hint_keys))
                value = {
                    **parsed,
                    "item": wcol["item"],
                    "normalized_item": wcol["normalized_item"],
                    "item_kind": wcol.get("item_kind"),
                    "material_family": wcol.get("material_family"),
                    "cable_category": wcol.get("cable_category"),
                    "shielding": wcol.get("shielding"),
                    "source_row_type": "line_item",
                    "aggregate": False,
                    "plate_id": plate or None,
                    "location": location or None,
                }
                value = merge_parser_value_identity(
                    value,
                    raw_text=f"{wcol['item']} {scope} {location} {row[idx]}".strip(),
                )
                rev = ReviewStatus.needs_review if parsed.get("review_flags") else ReviewStatus.auto_accepted
                atoms.append(
                    EvidenceAtom(
                        id=stable_id(
                            "atm",
                            project_id,
                            artifact_id,
                            sheet_name,
                            row_number,
                            "qty_wide",
                            wcol["normalized_item"],
                            str(qty_val or parsed.get("quantity_status")),
                        ),
                        project_id=project_id,
                        artifact_id=artifact_id,
                        atom_type=AtomType.quantity,
                        raw_text=f"Quantity {wcol['item']} {row[idx]}",
                        normalized_text=normalize_text(f"quantity {wcol['item']} {row[idx]}"),
                        value=value,
                        entity_keys=q_entity_keys,
                        source_refs=[self._build_source_ref(artifact_id, artifact_type, filename, sheet_name, row_number, cols)],
                        authority_class=AuthorityClass.approved_site_roster,
                        confidence=row_confidence * 0.95,
                        review_status=rev,
                        review_flags=parsed.get("review_flags") or [],
                        parser_version=self.parser_version,
                    )
                )
        else:
            quantity_raw = extracted.get("quantity", "").strip()
            if quantity_raw:
                parsed = _parse_schedule_quantity_cell(quantity_raw)
                qidx = model.header_map.get("quantity")
                qcol = get_column_letter((qidx or 0) + 1) if qidx is not None else "D"
                cols = {k: get_column_letter(model.header_map[k] + 1) for k in ("site", "device", "floor", "room", "quantity") if k in model.header_map}
                if "quantity" not in cols:
                    cols["quantity"] = qcol
                rev = ReviewStatus.needs_review if parsed.get("uncertain") or parsed.get("review_flags") else ReviewStatus.auto_accepted
                flags = list(parsed.get("review_flags") or [])
                if parsed.get("uncertain"):
                    flags.append("quantity_uncertain")
                value = {**parsed, "source_row_type": "line_item", "aggregate": False}
                value = merge_parser_value_identity(value, raw_text=f"Quantity {quantity_raw} {scope}")
                atoms.append(
                    EvidenceAtom(
                        id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, "qty", quantity_raw),
                        project_id=project_id,
                        artifact_id=artifact_id,
                        atom_type=AtomType.quantity,
                        raw_text=f"Quantity {quantity_raw}",
                        normalized_text=normalize_text(f"quantity {quantity_raw}"),
                        value=value,
                        entity_keys=entity_keys,
                        source_refs=[self._build_source_ref(artifact_id, artifact_type, filename, sheet_name, row_number, cols)],
                        authority_class=AuthorityClass.approved_site_roster,
                        confidence=row_confidence,
                        review_status=rev,
                        review_flags=flags,
                        parser_version=self.parser_version,
                    )
                )

        if scope or (site and device):
            work_scope = scope if scope else "work_item"
            cols = {k: get_column_letter(model.header_map[k] + 1) for k in ("scope", "site", "device", "floor", "room") if k in model.header_map}
            append_atom(
                AtomType.scope_item,
                f"Scope {work_scope}",
                {"scope": work_scope, "site": site, "device": device, "floor": floor, "room": room, "location": location},
                row_confidence * 0.9,
                extra_columns=cols or {"scope": "E"},
            )

        if access:
            aidx = model.header_map.get("access")
            if aidx is None:
                aidx = model.header_map.get("access_window")
            acol = get_column_letter((aidx or 0) + 1) if aidx is not None else "E"
            append_atom(
                AtomType.constraint,
                f"Access {access}",
                {"access_window": access, "site": site, "device": device, "location": location},
                row_confidence * 0.88,
                extra_columns={"access": acol},
            )

        scope_context_cols: dict[str, str] = {}
        if label_indices:
            scope_context_cols["label"] = get_column_letter(label_indices[0] + 1)
        for nk in ("notes", "description", "scope"):
            nix = model.header_map.get(nk)
            if nix is not None:
                scope_context_cols[nk] = get_column_letter(nix + 1)
        if not scope_context_cols:
            scope_context_cols["context"] = "A"

        def scope_append(
            atom_type: AtomType,
            raw_text: str,
            value: dict[str, Any],
            confidence: float,
            *,
            review_status: ReviewStatus = ReviewStatus.auto_accepted,
            review_flags: list[str] | None = None,
            extra_columns: dict[str, str] | None = None,
        ) -> None:
            append_atom(
                atom_type,
                raw_text,
                value,
                confidence,
                review_status=review_status,
                review_flags=review_flags,
                extra_columns=extra_columns or scope_context_cols,
            )

        _emit_scope_constraint_atoms(
            notes_blob,
            site,
            device,
            floor,
            room,
            location,
            scope_append,
            row_confidence,
            context_columns=scope_context_cols,
        )

        return atoms
