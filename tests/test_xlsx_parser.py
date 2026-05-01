from __future__ import annotations

from openpyxl import Workbook

from app.core.schemas import AtomType
from app.parsers.xlsx_parser import XlsxParser


def test_xlsx_parser_emits_atoms_with_provenance(tmp_path) -> None:
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "site_roster"
    ws.append(["Site", "Floor", "Device", "Quantity", "Access Window", "Scope"])
    ws.append(["Main Campus", "1", "IP Camera", "50", "Weekdays 8am-5pm", "Install"])
    ws.append(["West Wing", "2", "IP Camera", "41", "Escort required", "Install"])
    ws.append(["TOTAL", "", "", "91", "", ""])
    wb.save(file_path)

    atoms = XlsxParser().parse_artifact(
        project_id="proj_1",
        artifact_id="art_1",
        path=file_path,
    )
    assert atoms
    assert all(atom.source_refs for atom in atoms)

    quantity_atoms = [atom for atom in atoms if atom.atom_type == AtomType.quantity]
    quantities = [atom.value.get("quantity") for atom in quantity_atoms]
    line_qty = [a.value.get("quantity") for a in quantity_atoms if not a.value.get("aggregate")]
    agg_qty = [a.value.get("quantity") for a in quantity_atoms if a.value.get("aggregate")]
    assert 50 in line_qty
    assert 41 in line_qty
    assert 91 in agg_qty
    assert 91 not in line_qty

    all_keys = {key for atom in atoms for key in atom.entity_keys}
    assert "site:west_wing" in all_keys
    assert "device:ip_camera" in all_keys

    constraint_atoms = [atom for atom in atoms if atom.atom_type == AtomType.constraint]
    assert constraint_atoms
    assert any("escort required" in atom.normalized_text for atom in constraint_atoms)

    first_ref = atoms[0].source_refs[0]
    assert "sheet" in first_ref.locator
    assert "row" in first_ref.locator
    assert first_ref.locator["sheet"] == "site_roster"
    assert first_ref.parser_version == "xlsx_parser_v2_0"


def test_copper_drop_schedule_wide_totals(tmp_path) -> None:
    """COPPER_001-style governing addendum drop schedule (wide quantity columns + TOTALS)."""
    path = tmp_path / "drop_schedule.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "drops"
    ws.append(["Plate ID", "Location", "RJ45", "Cat6 UTP", "Cat6 STP"])
    ws.append(["P-01", "Sound Booth", 2, 2, 0])
    ws.append(["P-02", "Light Booth", 1, 1, 0])
    ws.append(["TOTALS", "", 72, 66, 6])
    wb.save(path)

    atoms = XlsxParser().parse_artifact("proj", "art", path)
    qty = [a for a in atoms if a.atom_type == AtomType.quantity]
    agg = [a for a in qty if a.value.get("aggregate")]
    line = [a for a in qty if not a.value.get("aggregate")]
    by_norm = {a.value.get("normalized_item"): a for a in agg}
    assert by_norm["rj45"].value.get("quantity") == 72
    assert by_norm["cat6_utp"].value.get("quantity") == 66
    assert by_norm["cat6_stp"].value.get("quantity") == 6
    assert all(a.atom_type == AtomType.quantity for a in agg)
    assert line, "expected line-item quantity atoms"
    assert not any(a.atom_type == AtomType.entity and "totals" in a.normalized_text for a in atoms)
    for a in agg:
        cols = a.source_refs[0].locator.get("columns", {})
        assert "quantity" in cols
        assert cols["quantity"] in {"C", "D", "E"}
    utp = next(a for a in agg if a.value.get("normalized_item") == "cat6_utp")
    stp = next(a for a in agg if a.value.get("normalized_item") == "cat6_stp")
    assert utp.value.get("shielding") == "unshielded"
    assert stp.value.get("shielding") == "shielded"


def test_wide_quantity_columns_distinct_items(tmp_path) -> None:
    path = tmp_path / "wide.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "cable_counts"
    ws.append(["Plate", "Room", "RJ45", "Cat6", "Cat6A", "Cat6 UTP", "Cat6 STP", "Fiber Strand"])
    ws.append(["A1", "101", 1, 1, 0, 1, 0, 12])
    wb.save(path)
    atoms = XlsxParser().parse_artifact("p", "a", path)
    qty = [a for a in atoms if a.atom_type == AtomType.quantity]
    norms = {a.value.get("normalized_item") for a in qty}
    assert "rj45" in norms
    assert "cat6" in norms
    assert "cat6a" in norms
    assert "cat6_utp" in norms
    assert "cat6_stp" in norms
    assert "fiber" in norms


def test_header_detection_shifted_title_two_row_slash(tmp_path) -> None:
    path = tmp_path / "headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["Project X", "", "", "", ""])
    ws.append(["", "", "", "", ""])
    ws.append(["", "Room / Area", "Material / Spec", "Qty / Count", "Included?"])
    ws.append(["", "101", "Cat6 UTP", "5", "Yes"])
    ws2 = wb.create_sheet("second")
    ws2.append(["Site", "Device", "Quantity"])
    ws2.append(["East", "Switch", "3"])
    wb.save(path)

    atoms = XlsxParser().parse_artifact("p", "a", path)
    assert atoms
    assert any(a.source_refs[0].locator.get("sheet") == "second" for a in atoms)


def test_row_classification_notes_with_total_word(tmp_path) -> None:
    path = tmp_path / "notes_total.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "s"
    ws.append(["Site", "Notes", "Quantity"])
    ws.append(["Main", "Grand total of work TBD", "4"])
    wb.save(path)
    atoms = XlsxParser().parse_artifact("p", "a", path)
    qty = [a for a in atoms if a.atom_type == AtomType.quantity and not a.value.get("aggregate")]
    assert any(a.value.get("quantity") == 4 for a in qty)


def test_section_header_no_entity_atoms(tmp_path) -> None:
    path = tmp_path / "section.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "s"
    ws.append(["Location", "Qty"])
    ws.append(["Section A:", ""])
    ws.append(["Room 1", "2"])
    wb.save(path)
    atoms = XlsxParser().parse_artifact("p", "a", path)
    entities = [a for a in atoms if a.atom_type == AtomType.entity]
    assert not any("section a" in e.normalized_text for e in entities)


def test_entity_keys_distinct_locations_and_mdf_idf(tmp_path) -> None:
    path = tmp_path / "entities.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "s"
    ws.append(["Plate ID", "Location", "MDF", "IDF", "RJ45"])
    ws.append(["P1", "Electric #1", "MDF-A", "IDF-1", 1])
    ws.append(["P2", "Electric #10", "MDF-A", "IDF-2", 1])
    ws.append(["P3", "Catwalk House Left", "", "", 1])
    ws.append(["P4", "Catwalk House Center", "", "", 1])
    wb.save(path)
    atoms = XlsxParser().parse_artifact("p", "a", path)
    keys = {k for a in atoms for k in a.entity_keys}
    assert "location:electric_1" in keys
    assert "location:electric_10" in keys
    assert "location:catwalk_house_left" in keys
    assert "location:catwalk_house_center" in keys
    assert "mdf:mdf_a" in keys
    assert "idf:idf_1" in keys
    assert "idf:idf_2" in keys


def test_constraint_and_open_question_from_notes(tmp_path) -> None:
    path = tmp_path / "constraints.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "s"
    ws.append(["Site", "Notes", "Quantity"])
    ws.append(["X", "after-hours only; customer provides lift", "1"])
    ws2 = wb.create_sheet("s2")
    ws2.append(["Site", "Notes", "Qty"])
    ws2.append(["Y", "confirm MDF badge access unknown", "2"])
    ws3 = wb.create_sheet("s3")
    ws3.append(["Site", "Notes", "Qty"])
    ws3.append(["Z", "certification required per TIA", "1"])
    ws4 = wb.create_sheet("s4")
    ws4.append(["Site", "Notes", "Qty"])
    ws4.append(["W", "labeling standard TBD", "1"])
    wb.save(path)
    atoms = XlsxParser().parse_artifact("p", "a", path)
    types = {(a.atom_type, a.value.get("constraint_type") or a.value.get("topic") or a.value.get("action")) for a in atoms}
    assert any(t[0] == AtomType.constraint and t[1] == "after_hours" for t in types)
    assert any(t[0] == AtomType.action_item for t in types)
    assert any(t[0] == AtomType.open_question and t[1] == "badge_or_access" for t in types)
    assert any(t[0] == AtomType.constraint and t[1] == "certification" for t in types)
    assert any(t[0] == AtomType.open_question and t[1] == "labeling" for t in types)


def test_false_positive_instruction_sheet(tmp_path) -> None:
    path = tmp_path / "instr.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws.append(["Read this sheet before entering data."])
    ws.append(["", "", ""])
    wb.save(path)
    assert XlsxParser().parse_artifact("p", "a", path) == []


def test_blank_workbook_no_crash(tmp_path) -> None:
    path = tmp_path / "blank.xlsx"
    wb = Workbook()
    wb.save(path)
    assert XlsxParser().parse_artifact("p", "a", path) == []


def test_every_atom_has_source_ref_and_parser_version(tmp_path) -> None:
    path = tmp_path / "sr.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Plate ID", "Location", "RJ45"])
    ws.append(["A", "Here", 3])
    wb.save(path)
    atoms = XlsxParser().parse_artifact("p", "a", path)
    for a in atoms:
        assert a.source_refs
        for ref in a.source_refs:
            assert ref.parser_version == "xlsx_parser_v2_0"
            assert ref.filename == "sr.xlsx"
            assert "columns" in ref.locator
            assert "row" in ref.locator
            assert "sheet" in ref.locator
