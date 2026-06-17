"""Commercial-sheet routing: pricing is captured, not dropped, and
never masquerades as scope.

Estimating workbooks carry rate cards, master price catalogs and
deal-financials tabs. Those are not customer scope, but they hold
pricing the PM needs. The sheet-role router sends them to typed
commercial atoms (``commercial_total`` / ``pricing_assumption``) with
``money:`` entity keys — out of ``scope_truth``, into the OrbitBrief
``pricing_clarity`` surface. Pure backing data (helper / dropdown /
empty / cover) is still dropped.
"""

from __future__ import annotations

from openpyxl import Workbook

from app.core.schemas import AtomType, AuthorityClass
from app.parsers.sheet_classifier import (
    SheetDestination,
    SheetRole,
    classify_sheet,
)
from app.parsers.xlsx_parser import XlsxParser


def _atoms(path):
    out = XlsxParser().parse_artifact("proj", "art", path)
    return out if isinstance(out, list) else out.atoms


# ── classifier: role → destination mapping ──────────────────────────


def test_role_destination_mapping() -> None:
    from app.parsers.sheet_classifier import SheetClassification

    cases = {
        SheetRole.SCOPE: SheetDestination.SCOPE,
        SheetRole.RATE_CARD: SheetDestination.COMMERCIAL,
        SheetRole.CATALOG: SheetDestination.COMMERCIAL,
        SheetRole.FINANCIAL_SUMMARY: SheetDestination.COMMERCIAL,
        SheetRole.EMPTY: SheetDestination.DROP,
        SheetRole.INSTRUCTIONS: SheetDestination.DROP,
        SheetRole.REFERENCE: SheetDestination.DROP,
    }
    for role, expected in cases.items():
        c = SheetClassification(role=role, suppress=role is not SheetRole.SCOPE,
                                reason="t", confidence=1.0)
        assert c.destination is expected, role


# ── deal-financials summary (label → value pairs) ───────────────────


def test_financial_summary_emits_commercial_totals(tmp_path) -> None:
    # A deal-kit financial summary is a 2-D label→value grid, not a row
    # table. The structured extractor recovers a clean deal header +
    # per-category P&L instead of gluing unrelated cells together — so the
    # economics surface as structured ``pl_line`` totals, not money-keyed
    # row glue.
    path = tmp_path / "Deal_Kit.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Deal Kit"
    ws.append(["OPPTY #", 126, "Total Deal Revenue", 21560])
    ws.append(["Sales Rep", "Dan", "Total Deal Cost", 15660])
    ws.append(["Customer", "DCW", "Total Deal Margin", 5900])
    ws.append(["Billing Type", "T&M", "Margin % on Total Deal", 0.27])
    wb.save(path)

    atoms = _atoms(path)
    # Never scope.
    assert not [a for a in atoms if a.atom_type == AtomType.scope_item]

    # ── deal header recovered cleanly — ONE atom per field (uniform row=atom),
    #    merged here into the identity record (build_deal_header does this for real) ──
    headers = [a for a in atoms if a.atom_type == AtomType.deal_metadata]
    f: dict = {}
    ekeys: set = set()
    for a in headers:
        for k, v in (a.value.get("fields") or {}).items():
            f.setdefault(k, v)
        ekeys.update(a.entity_keys or [])
    assert f["opportunity_id"] == "126"
    assert f["sales_rep"] == "Dan"
    assert f["customer"] == "DCW"
    assert f["billing_type"] == "T&M"
    assert "deal:126" in ekeys

    # ── P&L economics as per-row commercial_total ``pl_metric`` atoms (one per
    #    sheet row, faithful label; the PM brief regroups them at render time) ──
    totals = [a for a in atoms if a.atom_type == AtomType.commercial_total]
    assert totals
    deal = {
        a.value.get("metric"): a.value.get("value")
        for a in totals
        if a.value.get("kind") == "pl_metric" and a.value.get("category_key") == "deal"
    }
    assert deal["revenue"] == 21560
    assert deal["cost"] == 15660
    assert deal["margin"] == 5900
    # 0.27 fraction normalized to a 27% margin (a ratio, never a money key).
    assert deal["margin_pct"] == 27.0
    keys = {k for a in atoms for k in (a.entity_keys or [])}
    assert "money:0" not in keys
    # Tagged as vendor/internal pricing.
    assert all(a.authority_class == AuthorityClass.vendor_quote for a in totals)


# ── master catalog (money-keyword columns) ──────────────────────────


def test_catalog_emits_pricing_assumptions(tmp_path) -> None:
    path = tmp_path / "Deal_Kit.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Price List"
    ws.append(["ID #", "Material Description", "OEM", "Order QTY", "USA Cost $"])
    ws.append([1, "CAT6 Plenum cable 1000ft", "CommScope", None, 661.48])
    ws.append([2, "CAT6 Non-plenum 1000ft", "CommScope", None, 338.71])
    wb.save(path)

    atoms = _atoms(path)
    assert not [a for a in atoms if a.atom_type == AtomType.scope_item]
    pricing = [a for a in atoms if a.atom_type == AtomType.pricing_assumption]
    assert pricing
    # Bulk pricing sheets (catalogs / rate cards) collapse to a SINGLE
    # rollup atom — the granular rows are folded into value.rows, not
    # emitted as per-row atoms that bloat the envelope and never packetize.
    assert len(atoms) == 1
    summary = atoms[0]
    assert summary.atom_type == AtomType.pricing_assumption
    assert summary.value.get("is_summary") is True
    assert "pricing_rollup" in (summary.review_flags or [])
    assert summary.value["line_count"] == 2
    # Rows preserved losslessly for drill-down, with their money keys.
    folded = summary.value["rows"]
    assert len(folded) == 2
    folded_keys = {k for r in folded for k in r["money_keys"]}
    assert "money:661" in folded_keys and "money:339" in folded_keys
    # The rollup's own aggregate money keys cover the $-range (lo/hi).
    assert "money:661" in (summary.entity_keys or [])
    assert "money:339" in (summary.entity_keys or [])


# ── pure backing data is still dropped ──────────────────────────────


def test_helper_sheet_drops_to_nothing(tmp_path) -> None:
    path = tmp_path / "Deal_Kit.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Helper - Do not Edit"
    ws.append(["L0"])
    ws.append(["L1"])
    ws.append(["L2"])
    wb.save(path)

    assert _atoms(path) == []
