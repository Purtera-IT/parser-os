from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.core.schemas import AtomType, AuthorityClass, ReviewStatus
from app.core.source_replay import replay_atom_receipts
from app.parsers.quote_parser import (
    QuoteParser,
    _material_heuristics,
    normalize_inclusion,
    parse_money_cell,
    parse_quote_quantity,
)
from scripts.make_demo_fixtures import create_demo_project


def test_quote_parser_vendor_quote_atoms(tmp_path: Path) -> None:
    root = tmp_path / "repo"
    root.mkdir(parents=True, exist_ok=True)
    project_dir = create_demo_project(root)
    quote_path = project_dir / "vendor_quote.xlsx"

    atoms = QuoteParser().parse_artifact(
        project_id="proj_1",
        artifact_id="art_quote_1",
        path=quote_path,
    )

    assert atoms
    assert all(atom.source_refs for atom in atoms)
    assert all(atom.authority_class == AuthorityClass.vendor_quote for atom in atoms)

    line_items = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    assert line_items

    quantity_atoms = [a for a in atoms if a.atom_type == AtomType.quantity]
    assert any(a.value.get("quantity") == 72 for a in quantity_atoms)

    all_keys = {key for atom in atoms for key in atom.entity_keys}
    assert "device:ip_camera" in all_keys

    constraints = [a for a in atoms if a.atom_type == AtomType.constraint]
    assert any(a.value.get("lead_time") == "2 weeks" for a in constraints)

    locator = atoms[0].source_refs[0].locator
    assert "row" in locator
    assert "columns" in locator


def test_quote_parser_bom_style_line_item_sheet(tmp_path: Path) -> None:
    """COPPER_001-style short vendor quote: Line Item + Quoted Qty + material + Included (no unit price)."""
    path = tmp_path / "synthetic_vendor_quote_short.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.append(
        [
            "Line Item",
            "Quoted Qty",
            "UOM",
            "Quoted Material / Spec",
            "Included?",
            "Notes",
        ]
    )
    ws.append(["Cat6 UTP cable drops", 60, "EA", "Blue Cat6 UTP riser cable", "Yes", ""])
    ws.append(["Cat6 STP cable drops", 8, "EA", "Shielded Cat6 STP", "Yes", ""])
    ws.append(["RJ45 terminations", 68, "EA", "RJ45 terminations", "Yes", "Short against addendum total 72"])
    ws.append(["48-port patch panel", 1, "EA", "Cat6 48-port patch panel", "Yes", ""])
    ws.append(["Raceway/conduit allowance", 0, "Allowance", "Existing only", "No", ""])
    ws.append(["Cable certification report exports", 0, "Report", "Tester export", "No", ""])
    ws.append(["20 amp power locations", 4, "EA", "Dedicated circuit", "Yes", ""])
    wb.save(path)

    atoms = QuoteParser().parse_artifact(project_id="proj_copper", artifact_id="art_quote_bom", path=path)
    assert atoms
    assert all(atom.source_refs for atom in atoms)
    assert all(atom.authority_class == AuthorityClass.vendor_quote for atom in atoms)

    vli = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    qty_atoms = [a for a in atoms if a.atom_type == AtomType.quantity]
    assert len(vli) == 7
    assert len(qty_atoms) == 7

    by_desc = {a.value.get("description", ""): a for a in vli}
    assert by_desc["RJ45 terminations"].value.get("quantity") == "68"
    assert by_desc["Cat6 UTP cable drops"].value.get("material_spec") == "Blue Cat6 UTP riser cable"
    assert by_desc["Cat6 STP cable drops"].value.get("material_spec") == "Shielded Cat6 STP"
    assert by_desc["Raceway/conduit allowance"].value.get("included") is False
    assert by_desc["Cable certification report exports"].value.get("included") is False
    assert by_desc["20 amp power locations"].value.get("quantity") == "4"

    assert any(q.value.get("quantity") == 68 for q in qty_atoms)
    assert any(q.value.get("quantity") == 60 for q in qty_atoms)
    assert any(q.value.get("quantity") == 8 for q in qty_atoms)
    assert any(q.value.get("quantity") == 0 for q in qty_atoms)
    assert any(q.value.get("quantity") == 4 for q in qty_atoms)

    ref = vli[0].source_refs[0]
    assert ref.locator.get("sheet") == "Quote"
    assert "columns" in ref.locator and "description" in ref.locator["columns"]
    assert ref.parser_version == "quote_parser_v1_4_1"

    utp = by_desc["Cat6 UTP cable drops"].value
    stp = by_desc["Cat6 STP cable drops"].value
    assert utp.get("cable_category") == "cat6" and utp.get("shielding") == "unshielded"
    assert stp.get("cable_category") == "cat6" and stp.get("shielding") == "shielded"
    assert by_desc["48-port patch panel"].value.get("port_count") == 48
    assert by_desc["48-port patch panel"].value.get("item_kind") == "patch_panel"
    assert by_desc["Cable certification report exports"].value.get("inclusion_status") == "excluded"
    assert by_desc["Cable certification report exports"].value.get("item_kind") == "certification"
    assert by_desc["Raceway/conduit allowance"].value.get("inclusion_status") == "excluded"
    pow_row = by_desc["20 amp power locations"].value
    assert pow_row.get("item_kind") == "power"
    assert pow_row.get("is_scope_pollution_candidate") is True

    auto = [a for a in vli if a.review_status == ReviewStatus.auto_accepted]
    review = [a for a in vli if a.review_status == ReviewStatus.needs_review]
    assert len(auto) == 5 and len(review) == 2
    assert {a.value.get("description") for a in review} == {
        "Raceway/conduit allowance",
        "Cable certification report exports",
    }

    expected_keys = {
        "Cat6 UTP cable drops": (
            "cabling:cat6:utp:drop",
            "material",
            "in_scope_candidate",
        ),
        "Cat6 STP cable drops": (
            "cabling:cat6:stp:drop",
            "material",
            "in_scope_candidate",
        ),
        "RJ45 terminations": (
            "cabling:rj45:termination",
            "material",
            "in_scope_candidate",
        ),
        "48-port patch panel": (
            "cabling:patch_panel:48_port",
            "material",
            "in_scope_candidate",
        ),
        "Raceway/conduit allowance": (
            "pathway:raceway_conduit",
            "allowance",
            "excluded_candidate",
        ),
        "Cable certification report exports": (
            "testing:certification_export",
            "testing",
            "excluded_candidate",
        ),
        "20 amp power locations": (
            "electrical:power_location",
            "electrical",
            "scope_pollution_candidate",
        ),
    }
    for desc, (ck, cr, sr) in expected_keys.items():
        v = by_desc[desc].value
        row = by_desc[desc].source_refs[0].locator["row"]
        assert v.get("comparison_key") == ck
        assert v.get("commercial_role") == cr
        assert v.get("scope_relevance") == sr
        assert v.get("source_row_key") == f"synthetic_vendor_quote_short.xlsx:Quote:row_{row}"
        assert isinstance(v.get("confidence_dimensions"), dict)
        assert set(v["confidence_dimensions"].keys()) == {
            "header_mapping",
            "quantity_parse",
            "item_normalization",
            "inclusion_parse",
            "source_ref",
        }
        expl = v.get("parser_explanation")
        assert isinstance(expl, list) and len(expl) == 6
        assert any("comparison_key:" in x for x in expl)

    vlis_by_row = sorted(vli, key=lambda a: a.source_refs[0].locator["row"])
    qtys_by_row = sorted(qty_atoms, key=lambda a: a.source_refs[0].locator["row"])
    for v_atom, q_atom in zip(vlis_by_row, qtys_by_row):
        assert v_atom.value.get("comparison_key") == q_atom.value.get("comparison_key")
        assert v_atom.value.get("source_row_key") == q_atom.value.get("source_row_key")

    for q in qty_atoms:
        assert q.value.get("comparison_key")
        assert q.value.get("comparison_basis") == "vendor_proposed_quantity"
        assert q.value.get("authority_boundary") == "vendor_quote_can_conflict_but_not_define_scope"

    pow_qty = next(q for q in qty_atoms if q.value.get("comparison_key") == "electrical:power_location")
    assert pow_qty.value.get("scope_relevance") == "scope_pollution_candidate"
    rac_qty = next(q for q in qty_atoms if q.value.get("comparison_key") == "pathway:raceway_conduit")
    cert_qty = next(q for q in qty_atoms if q.value.get("comparison_key") == "testing:certification_export")
    assert rac_qty.value.get("scope_relevance") == "excluded_candidate"
    assert cert_qty.value.get("scope_relevance") == "excluded_candidate"


def test_quote_parser_does_not_skip_row_when_notes_mention_total(tmp_path: Path) -> None:
    """Notes column may contain 'total' (e.g. validation hint); line item must still parse."""
    path = tmp_path / "quote_with_total_in_notes.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.append(["Line Item", "Quoted Qty", "UOM", "Quoted Material / Spec", "Included?", "Notes"])
    ws.append(["RJ45 terminations", 68, "EA", "RJ45 jacks", "Yes", "Short against addendum total 72"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "art", path)
    vli = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    assert len(vli) == 1
    assert vli[0].value.get("description") == "RJ45 terminations"
    assert vli[0].value.get("quantity") == "68"


def test_quote_parser_header_on_row_five_with_title_rows(tmp_path: Path) -> None:
    path = tmp_path / "late_header.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Vendor Quote — Confidential"])
    ws.append([])
    ws.append(["Do not distribute"])
    ws.append([])
    ws.append(["Line Item", "Qty", "Unit Price"])
    ws.append(["Widget A", "10", "$5.00"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "art", path)
    vli = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    assert len(vli) == 1
    assert vli[0].source_refs[0].locator.get("row") == 6


def test_quote_parser_two_row_merged_headers(tmp_path: Path) -> None:
    path = tmp_path / "two_row_hdr.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Spring Lake Auditorium"])
    ws.append(["Quoted", "", "Line", "Included?"])
    ws.append(["Qty", "Material / Spec", "Item", ""])
    ws.append(["5", "Cat6 UTP", "Cable drops", "Yes"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "art", path)
    vli = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    assert len(vli) == 1
    assert vli[0].value.get("description") == "Cable drops"
    assert vli[0].value.get("quantity_parsed", {}).get("quantity") == 5


def test_quote_parser_slash_header_single_row(tmp_path: Path) -> None:
    path = tmp_path / "slash_hdr.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Line Item", "Quoted Qty", "Material / Spec", "Included?"])
    ws.append(["Patch panel", "2", "Cat6 48-port", "Yes"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "art", path)
    vli = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    assert len(vli) == 1
    assert vli[0].value.get("material_spec") == "Cat6 48-port"


def test_quote_parser_csv_pipe_and_txt_tab(tmp_path: Path) -> None:
    pipe = tmp_path / "q_pipe.csv"
    pipe.write_text("Description|Qty|Unit Price\nPanel|3|$100\n", encoding="utf-8")
    tab_path = tmp_path / "q_tab.txt"
    tab_path.write_text("Description\tQty\nBracket\t12\n", encoding="utf-8")
    p_atoms = QuoteParser().parse_artifact("p", "pipe", pipe)
    t_atoms = QuoteParser().parse_artifact("p", "tab", tab_path)
    assert any(a.value.get("description") == "Panel" for a in p_atoms if a.atom_type == AtomType.vendor_line_item)
    assert any(a.value.get("description") == "Bracket" for a in t_atoms if a.atom_type == AtomType.vendor_line_item)


def test_quote_parser_skips_subtotal_grand_total_tax(tmp_path: Path) -> None:
    path = tmp_path / "totals.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Description", "Qty", "Unit Price"])
    ws.append(["Part A", "2", "$1"])
    ws.append(["Subtotal", "", ""])
    ws.append(["Tax", "1", "$0.10"])
    ws.append(["Grand Total", "", "$9.99"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "art", path)
    vli = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    assert len(vli) == 1
    assert vli[0].value.get("description") == "Part A"


def test_quote_quantity_normalization_matrix() -> None:
    assert parse_quote_quantity("", "68 EA", "", "")["quantity"] == 68
    assert parse_quote_quantity("", "1,200", "", "")["quantity"] == 1200
    assert parse_quote_quantity("", "1 lot", "", "")["quantity"] == 1
    assert parse_quote_quantity("", "included", "", "")["quantity_status"] == "included_no_qty"
    assert parse_quote_quantity("", "TBD", "", "")["quantity_status"] == "tbd"
    assert parse_quote_quantity("", "N/A", "", "")["quantity_status"] == "not_applicable"
    r = parse_quote_quantity("", "60-68", "", "")
    assert r["quantity_status"] == "range" and r["quantity_min"] == 60 and r["quantity_max"] == 68
    assert parse_quote_quantity("", "approx. 68", "", "")["quantity"] == 68
    emb = parse_quote_quantity("(68) RJ45 terminations", "", "", "")
    assert emb["quantity"] == 68
    emb2 = parse_quote_quantity("RJ45 terminations - qty 68", "", "", "")
    assert emb2["quantity"] == 68


def test_quote_inclusion_normalization() -> None:
    assert normalize_inclusion("Yes", "")["inclusion_status"] == "included"
    assert normalize_inclusion("No", "")["inclusion_status"] == "excluded"
    assert normalize_inclusion("", "By Others")["inclusion_status"] == "excluded"
    assert normalize_inclusion("", "NIC scope")["inclusion_status"] == "excluded"
    assert normalize_inclusion("", "Not Included per owner")["inclusion_status"] == "excluded"
    assert normalize_inclusion("Yes", "Alternate pricing")["inclusion_status"] == "optional"
    assert normalize_inclusion("", "Allowance for spares")["inclusion_status"] == "allowance"
    assert normalize_inclusion("No", "Allowance raceway")["inclusion_status"] == "excluded"


def test_quote_money_parsing() -> None:
    u = parse_money_cell("$1,234.56", side="unit")
    assert u["unit_price_amount"] == 1234.56 and u["price_status"] == "known"
    assert u["extended_price_amount"] is None
    neg = parse_money_cell("($50.00)", side="unit")
    assert neg["unit_price_amount"] == -50.0
    nc = parse_money_cell("N/C", side="unit")
    assert nc["price_status"] == "no_charge"
    inc = parse_money_cell("included", side="extended")
    assert inc["price_status"] == "included"
    blank = parse_money_cell("", side="unit")
    assert blank["price_status"] == "missing"
    bad = parse_money_cell("not-a-price", side="unit")
    assert bad["price_status"] == "malformed"


def test_quote_parser_false_positive_no_line_items(tmp_path: Path) -> None:
    instr = tmp_path / "instructions_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["INSTRUCTIONS TO BIDDERS"])
    ws.append(["Submit pricing in the attached Excel template."])
    wb.save(instr)
    assert QuoteParser().parse_artifact("p", "i", instr) == []

    terms = tmp_path / "terms_only.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Terms and Conditions"])
    ws2.append(["Payment net 30."])
    wb2.save(terms)
    assert QuoteParser().parse_artifact("p", "t", terms) == []

    cover = tmp_path / "cover_only.xlsx"
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(["Acme Low Voltage Inc."])
    ws3.append(["Quote for Auditorium Renovation"])
    wb3.save(cover)
    assert QuoteParser().parse_artifact("p", "c", cover) == []


def test_quote_parser_source_replay_all_verified(tmp_path: Path) -> None:
    path = tmp_path / "replay_quote.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Line Item", "Quoted Qty", "Quoted Material / Spec", "Included?"])
    ws.append(["Item A", 3, "Spec", "Yes"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("proj", "art_replay", path)
    assert atoms
    art_map = {"art_replay": path}
    line_qty = [a for a in atoms if a.atom_type in (AtomType.vendor_line_item, AtomType.quantity)]
    assert line_qty
    for atom in line_qty:
        receipts = replay_atom_receipts(atom, art_map)
        assert receipts and all(r.replay_status == "verified" for r in receipts)
    for atom in atoms:
        for ref in atom.source_refs:
            assert ref.filename == path.name
            assert isinstance(ref.locator.get("row"), int)
            assert ref.parser_version
            assert ref.extraction_method
            if ref.locator.get("columns") is not None:
                assert ref.locator.get("columns")
        receipts = replay_atom_receipts(atom, art_map)
        assert receipts
        assert all(
            r.replay_status in ("verified", "unsupported") for r in receipts
        ), f"{atom.atom_type} {[(r.replay_status, r.reason) for r in receipts]}"


def test_quote_parser_polish_cat6a_never_classified_as_cat6() -> None:
    assert _material_heuristics("Cat6A UTP horizontal cable", "", "")["cable_category"] == "cat6a"
    assert _material_heuristics("Cat 6A STP runs", "", "")["cable_category"] == "cat6a"
    assert _material_heuristics("Category 6A UTP", "", "")["cable_category"] == "cat6a"
    assert _material_heuristics("Cat6 UTP riser", "", "")["cable_category"] == "cat6"
    assert _material_heuristics("Category 6 plenum", "", "")["cable_category"] == "cat6"


def test_quote_parser_polish_jack_keystone_faceplate_kinds() -> None:
    assert _material_heuristics("RJ45 jacks bulk pack", "", "")["item_kind"] == "jack"
    assert _material_heuristics("Keystone jack white", "", "")["item_kind"] == "keystone_jack"
    assert _material_heuristics("2-gang faceplate blank", "", "")["item_kind"] == "faceplate"


def test_quote_parser_price_math_mismatch_flag(tmp_path: Path) -> None:
    path = tmp_path / "price_math.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Description", "Qty", "Unit Price", "Extended Price"])
    ws.append(["Bad math line", 10, 5.0, 60.0])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "price", path)
    vli = next(a for a in atoms if a.atom_type == AtomType.vendor_line_item)
    assert "quote_parser:price_math_mismatch" in vli.review_flags
    assert vli.review_status == ReviewStatus.needs_review


def test_quote_parser_needs_review_excluded_optional_allowance(tmp_path: Path) -> None:
    path = tmp_path / "risk_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Description", "Qty", "Included?", "Notes"])
    ws.append(["Opt add", 1, "Yes", "Alternate pricing"])
    ws.append(["Allow line", 0, "Yes", "Allowance for scrap"])
    ws.append(["No line", 1, "No", ""])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "risk", path)
    vli = [a for a in atoms if a.atom_type == AtomType.vendor_line_item]
    assert all(a.review_status == ReviewStatus.needs_review for a in vli)


def test_quote_parser_ordinary_included_row_auto_accepted(tmp_path: Path) -> None:
    path = tmp_path / "clean.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Description", "Qty", "Included?"])
    ws.append(["Cable pull", 12, "Yes"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "clean", path)
    vli = next(a for a in atoms if a.atom_type == AtomType.vendor_line_item)
    qty = next(a for a in atoms if a.atom_type == AtomType.quantity)
    assert vli.review_status == ReviewStatus.auto_accepted
    assert qty.review_status == ReviewStatus.auto_accepted


def test_quote_parser_universals_cat6a_and_jacks(tmp_path: Path) -> None:
    path = tmp_path / "uni.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Line Item", "Quoted Qty", "Quoted Material / Spec", "Included?", "Notes"])
    ws.append(["Cat6A UTP drops", 10, "Cat6A UTP", "Yes", ""])
    ws.append(["RJ45 jacks", 5, "Modular", "Yes", ""])
    ws.append(["Keystone jack", 2, "White", "Yes", ""])
    ws.append(["Wall plate", 1, "2-gang", "Yes", ""])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "uni", path)
    by = {a.value.get("description"): a.value for a in atoms if a.atom_type == AtomType.vendor_line_item}
    assert by["Cat6A UTP drops"]["comparison_key"] == "cabling:cat6a:utp:drop"
    assert by["RJ45 jacks"]["comparison_key"] == "cabling:jack"
    assert by["Keystone jack"]["comparison_key"] == "cabling:keystone_jack"
    assert by["Wall plate"]["comparison_key"] == "cabling:faceplate"


def test_quote_parser_universals_owner_furnished_and_phrases(tmp_path: Path) -> None:
    path = tmp_path / "ofe.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Line Item", "Qty", "Included?", "Notes"])
    ws.append(["Owner pull wire", 1, "Yes", "OFE — owner furnished"])
    ws.append(["By others demo", 1, "Yes", "by others scope"])
    ws.append(["NIC line", 1, "Yes", "NIC"])
    wb.save(path)
    atoms = QuoteParser().parse_artifact("p", "ofe", path)
    vli = {a.value.get("description"): a.value for a in atoms if a.atom_type == AtomType.vendor_line_item}
    ofe = vli["Owner pull wire"]
    assert ofe.get("included") is False
    assert ofe.get("commercial_role") == "owner_furnished"
    assert ofe.get("scope_relevance") == "excluded_candidate"
    assert vli["By others demo"].get("inclusion_status") == "excluded"
    assert vli["NIC line"].get("inclusion_status") == "excluded"
