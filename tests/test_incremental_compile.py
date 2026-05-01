from __future__ import annotations

import hashlib
import shutil
from pathlib import Path

import yaml
from openpyxl import load_workbook

from app.core.compiler import compile_project


def _copy_demo_project(demo_project: Path, target: Path) -> Path:
    shutil.copytree(demo_project, target)
    return target


def _artifact_count(result) -> int:
    assert result.manifest is not None
    return len(result.manifest.artifact_fingerprints)


def _artifact_id_for_filename(result, filename: str) -> str:
    assert result.manifest is not None
    for row in result.manifest.artifact_fingerprints:
        if row.filename == filename:
            return row.artifact_id
    raise AssertionError(f"Missing artifact fingerprint for {filename}")


def _project_id(prefix: str, tmp_path: Path) -> str:
    digest = hashlib.sha256(str(tmp_path).encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{digest}"


def test_compile_demo_once_has_cache_misses(demo_project: Path, tmp_path: Path) -> None:
    project = _copy_demo_project(demo_project, tmp_path / "project")
    result = compile_project(
        project,
        project_id=_project_id("inc_case_once", tmp_path),
        allow_errors=True,
        allow_unverified_receipts=True,
    )
    assert result.manifest is not None
    assert result.manifest.cache_hits == 0
    assert result.manifest.cache_misses == _artifact_count(result)
    assert not result.manifest.reused_artifact_ids


def test_compile_demo_twice_has_cache_hits(demo_project: Path, tmp_path: Path) -> None:
    project = _copy_demo_project(demo_project, tmp_path / "project")
    project_id = _project_id("inc_case_twice", tmp_path)
    first = compile_project(project, project_id=project_id, allow_errors=True, allow_unverified_receipts=True)
    second = compile_project(project, project_id=project_id, allow_errors=True, allow_unverified_receipts=True)
    assert first.manifest is not None and second.manifest is not None
    assert second.manifest.cache_hits == _artifact_count(second)
    assert second.manifest.cache_misses == 0
    assert sorted(second.manifest.reused_artifact_ids) == sorted(
        [row.artifact_id for row in second.manifest.artifact_fingerprints]
    )


def test_modify_vendor_quote_only_quote_cache_miss(demo_project: Path, tmp_path: Path) -> None:
    project = _copy_demo_project(demo_project, tmp_path / "project")
    project_id = _project_id("inc_case_delta", tmp_path)
    _ = compile_project(project, project_id=project_id, allow_errors=True, allow_unverified_receipts=True)

    quote_path = project / "vendor_quote.xlsx"
    workbook = load_workbook(quote_path)
    sheet = workbook.active
    sheet["C2"] = "99"
    workbook.save(quote_path)
    workbook.close()

    updated = compile_project(project, project_id=project_id, allow_errors=True, allow_unverified_receipts=True)
    assert updated.manifest is not None
    quote_artifact_id = _artifact_id_for_filename(updated, "vendor_quote.xlsx")
    assert updated.manifest.cache_misses == 1
    assert quote_artifact_id not in set(updated.manifest.reused_artifact_ids)
    assert updated.manifest.cache_hits == _artifact_count(updated) - 1


def test_packetizer_output_updates_after_modified_artifact(demo_project: Path, tmp_path: Path) -> None:
    project = _copy_demo_project(demo_project, tmp_path / "project")
    project_id = _project_id("inc_case_packets", tmp_path)
    baseline = compile_project(project, project_id=project_id, allow_errors=True, allow_unverified_receipts=True)

    quote_path = project / "vendor_quote.xlsx"
    workbook = load_workbook(quote_path)
    sheet = workbook.active
    sheet["C2"] = "120"
    workbook.save(quote_path)
    workbook.close()

    updated = compile_project(project, project_id=project_id, allow_errors=True, allow_unverified_receipts=True)
    assert baseline.manifest is not None and updated.manifest is not None
    assert baseline.manifest.output_signature != updated.manifest.output_signature
    assert any(packet.family.value == "quantity_conflict" for packet in updated.packets)


def test_no_cache_reparses_all(demo_project: Path, tmp_path: Path) -> None:
    project = _copy_demo_project(demo_project, tmp_path / "project")
    project_id = _project_id("inc_case_nocache", tmp_path)
    _ = compile_project(project, project_id=project_id, allow_errors=True, allow_unverified_receipts=True)
    result = compile_project(
        project,
        project_id=project_id,
        allow_errors=True,
        allow_unverified_receipts=True,
        use_cache=False,
    )
    assert result.manifest is not None
    assert result.manifest.cache_hits == 0
    assert result.manifest.cache_misses == _artifact_count(result)
    assert not result.manifest.reused_artifact_ids


def test_cache_key_includes_domain_pack_version(demo_project: Path, tmp_path: Path) -> None:
    project = _copy_demo_project(demo_project, tmp_path / "project")
    project_id = _project_id("inc_case_packver", tmp_path)
    pack_path = tmp_path / "pack.yaml"
    pack_payload = yaml.safe_load(Path("app/domain/security_camera_pack.yaml").read_text(encoding="utf-8"))
    assert isinstance(pack_payload, dict)
    pack_path.write_text(yaml.safe_dump(pack_payload, sort_keys=False), encoding="utf-8")

    _ = compile_project(
        project,
        project_id=project_id,
        allow_errors=True,
        allow_unverified_receipts=True,
        domain_pack=pack_path,
    )

    pack_payload["version"] = "9.9.9"
    pack_path.write_text(yaml.safe_dump(pack_payload, sort_keys=False), encoding="utf-8")
    updated = compile_project(
        project,
        project_id=project_id,
        allow_errors=True,
        allow_unverified_receipts=True,
        domain_pack=pack_path,
    )
    assert updated.manifest is not None
    assert updated.manifest.cache_hits == 0
    assert updated.manifest.cache_misses == _artifact_count(updated)
