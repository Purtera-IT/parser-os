"""The router must decide on evidence, and on evidence in the right order.

Every case here was a real misroute observed on real deal packs. The packs
themselves live outside the repository and are untracked, so the shapes are
reconstructed as fixtures: these run anywhere, and a revert of any of the
fixes below fails here rather than silently on somebody's deal.

The ordering the tests pin down, strongest evidence first:

    container / magic bytes   what the file IS
    content structure         RFC-5322 headers, quote header rows, timestamps
    extension                 a convention, checkable but forgeable
    filename                  a PRIOR -- may raise a claim, may never create one
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.parsers.registry import choose_parser


def _parser_name(path: Path) -> str:
    parser, _match, _all = choose_parser(path)
    return type(parser).__name__ if parser is not None else "NONE"


# ── a business document is not a transcript ──────────────────────────────


def test_rfp_text_is_not_a_transcript(tmp_path: Path) -> None:
    """One "Label:" line used to claim the whole document at 0.82.

    ``detect_speaker`` matches any "Label: value", which is what business
    documents are built from, and the rule accepted a single hit in the first
    forty lines. "School District Contact:" on page one of a Request for
    Proposals was enough. Reading an RFP as a transcript cost every quantity
    and every constraint atom in it.
    """
    rfp = tmp_path / "rfp_original.txt"
    rfp.write_text(
        "Request for Proposals\n"
        "Structured Cabling\n"
        "School District Contact:\n"
        "David Miller, david@example.com\n"
        "Prepared by: Consulting Group\n"
        + "The contractor shall install Cat6 UTP drops at each location.\n" * 30,
        encoding="utf-8",
    )
    assert _parser_name(rfp) != "TranscriptParser"


def test_transcript_with_timestamps_still_routes_to_transcript(tmp_path: Path) -> None:
    """A timestamp is unambiguous transcript evidence and qualifies on sight."""
    t = tmp_path / "kickoff.txt"
    t.write_text(
        "\n".join(
            f"[00:0{i % 10}:12] Speaker {i % 3}: we need forty sites by Q3."
            for i in range(30)
        ),
        encoding="utf-8",
    )
    assert _parser_name(t) == "TranscriptParser"


def test_transcript_without_timestamps_routes_on_speaker_density(tmp_path: Path) -> None:
    """Teams "Save as .txt" has speaker turns and no timestamps.

    Speaker labels qualify when they are how the document is BUILT -- which is
    what a transcript is -- not when one appears.
    """
    t = tmp_path / "teams_export.txt"
    t.write_text(
        "\n".join(
            f"{'Cliff Creech' if i % 2 else 'Purtera PM'}: line {i} of the discussion."
            for i in range(30)
        ),
        encoding="utf-8",
    )
    assert _parser_name(t) == "TranscriptParser"


# ── a filename is a prior, not a verdict ─────────────────────────────────


def test_pm_note_named_like_a_quote_is_read_as_prose(tmp_path: Path) -> None:
    """Named "pricing", contains no quote table.

    This was claimed by QuoteParser at 0.95 on its name alone and produced
    ZERO atoms -- the narrative was lost outright.
    """
    note = tmp_path / "pm_note_pricing_schedule_not_scope.txt"
    note.write_text(
        "PM Note - Structured Cabling Services\n\n"
        "Appendix I is a pricing schedule/catalog, not a project-specific scope "
        "schedule.\n"
        "The work order is the project scope source.\n"
        "Vendor quote should be checked against the work order quantities.\n",
        encoding="utf-8",
    )
    assert _parser_name(note) != "QuoteParser"


def test_quote_is_recognised_without_a_helpful_filename(tmp_path: Path) -> None:
    """The same table under a neutral name must still be a quote.

    ``looks_like_quote_artifact`` used to open with
    ``if path_quote_filename_hint(path): return True`` -- a content predicate
    answering from the filename. Content has to stand on its own.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Part Number", "Description", "Qty", "Unit Price", "Extended"])
    ws.append(["CAM-IP-001", "IP Camera", 192, 338.71, 65032.32])
    ws.append(["BRK-88-B117", "Mounting bracket", 64, 12.50, 800.00])
    neutral = tmp_path / "attachment_b.xlsx"
    wb.save(neutral)
    assert _parser_name(neutral) == "QuoteParser"


def _quote_sheet(target: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Part Number", "Description", "Qty", "Unit Price", "Extended"])
    ws.append(["CAM-IP-001", "IP Camera, 4MP dome", 192, 338.71, 65032.32])
    ws.append(["BRK-88-B117", "Mounting bracket", 64, 12.50, 800.00])
    wb.save(target)
    return target


def _roster_sheet(target: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Site", "Building", "Room", "Drops", "Plate Type", "Access Window"])
    for i in range(12):
        ws.append([f"ATL-{i:02d}", "Building C", f"IDF {i}", 24, "RJ45 quad",
                   "after hours, escort required"])
    wb.save(target)
    return target


def test_a_quote_under_a_roster_name_is_still_a_quote(tmp_path: Path) -> None:
    """Identical tables, two names, and they must route the same way.

    ``QuoteParser.match`` returned a hard 0.00 on a filename substring before
    reading anything, so this file never entered the candidate list -- while
    its twin named ``attachment_b.xlsx`` was claimed at 0.86 on the very same
    header rows. Two more layers had the same defect: the quote/xlsx
    tie-break computed ``quote_ok`` and then returned on the name two lines
    later, and ``looks_like_quote_artifact`` once opened by returning True on
    the name.
    """
    assert _parser_name(_quote_sheet(tmp_path / "site_list_schedule.xlsx")) == "QuoteParser"
    assert _parser_name(_quote_sheet(tmp_path / "attachment_b.xlsx")) == "QuoteParser"


@pytest.mark.parametrize(
    "filename",
    ["site_list.xlsx", "asset_inventory.xlsx", "risk_register.xlsx",
     "license_support_matrix.xlsx"],
)
def test_the_cede_still_holds_for_a_real_roster(tmp_path: Path, filename: str) -> None:
    """The other half, and the reason the cede exists at all.

    These filenames carry richer AtomTypes -- ``site_roster``,
    ``asset_record``, ``risk``, ``support_entitlement`` -- and letting the
    quote parser claim them collapses those structured fields into
    ``vendor_line_item`` plus ``quantity``. Making the cede consult content
    must not cost that: a roster with roster headers still cedes.
    """
    from app.parsers.quote_parser import QuoteParser

    match = QuoteParser().match(_roster_sheet(tmp_path / filename), None, None)
    assert match.confidence == 0.0
    assert "ceded_to_xlsx_typed_row_profiler" in match.reasons


def test_quote_filename_alone_cannot_claim_a_file(tmp_path: Path) -> None:
    """A quote-ish name over content that is plainly not a quote."""
    fake = tmp_path / "vendor_quote_notes.txt"
    fake.write_text(
        "The customer confirmed that pricing is NOT part of this scope.\n"
        "Escort access is required at the Atlanta dock before 2pm.\n"
        "Contractor shall install forty cameras across five sites.\n" * 4,
        encoding="utf-8",
    )
    assert _parser_name(fake) != "QuoteParser"


# ── one keyword is not evidence of correspondence ────────────────────────


def test_table_dump_with_a_scope_phrase_is_not_an_email(tmp_path: Path) -> None:
    """A pipe-delimited site table containing "escort required" thirty times.

    EmailParser claimed it at 0.55 on a single keyword hit -- a fallback that
    existed so keyword-bearing text "isn't silently dropped", from when
    nothing claimed .txt at all.
    """
    table = tmp_path / "roster_dump.txt"
    table.write_text(
        "\n".join(
            ["Site | Qty | Notes"]
            + [f"ATL-{i:02d} | {i} | escort required" for i in range(30)]
        ),
        encoding="utf-8",
    )
    assert _parser_name(table) != "EmailParser"


def test_real_email_still_claims_on_its_headers(tmp_path: Path) -> None:
    """Narrowing the keyword fallback must not touch actual correspondence."""
    mail = tmp_path / "customer_email.txt"
    mail.write_text(
        "From: jane.customer@example.com\n"
        "Sent: 2026-01-15 09:00\n"
        "Subject: Scope update\n\n"
        "Please remove West Wing from scope.\n",
        encoding="utf-8",
    )
    assert _parser_name(mail) == "EmailParser"


# ── the coverage contract for text that has nothing in it ────────────────


def test_contentless_text_still_reports_no_parser(tmp_path: Path) -> None:
    """The prose floor must not swallow the "I did not read this" signal.

    ``test_random_txt_produces_warning_no_crash`` requires an unrecognised
    .txt to warn. That warning IS the coverage record: saying nothing was read
    is more honest than manufacturing atoms out of filler. Every real deal
    document measured is at least 4 non-empty lines and 258 characters; this
    is one line and 44.
    """
    filler = tmp_path / "random.txt"
    filler.write_text("just filler words with no structured signals", encoding="utf-8")
    assert _parser_name(filler) == "NONE"


def test_prose_text_with_a_body_is_claimed(tmp_path: Path) -> None:
    """The other half of the same contract: a real document is read."""
    doc = tmp_path / "scope_notes.txt"
    doc.write_text(
        "Scope of Work\n\n"
        "Contractor shall install forty cameras at the Atlanta facility.\n"
        "Escort is required at all times inside the yard.\n"
        "Mid-turn jumpers are excluded from this bill of materials.\n",
        encoding="utf-8",
    )
    assert _parser_name(doc) != "NONE"


# ── content beats a lying extension ──────────────────────────────────────


@pytest.mark.parametrize(
    "renamed, builder",
    [
        ("contract.pdf", "docx"),
        ("message", "eml"),
    ],
)
def test_content_signature_overrules_the_name(tmp_path: Path, renamed: str, builder: str) -> None:
    """Magic bytes and format declarations outrank whatever the file is called."""
    target = tmp_path / renamed
    if builder == "docx":
        from docx import Document

        doc = Document()
        doc.add_paragraph("Contractor shall install conduit.")
        doc.save(target)
        assert _parser_name(target) == "DocxParser"
    else:
        target.write_bytes(
            b"From: cliff@example.com\r\n"
            b"To: pm@example.com\r\n"
            b"Subject: Rollout\r\n\r\n"
            b"Forty sites by Q3.\r\n"
        )
        assert _parser_name(target) == "EmailParser"


# ── the Otter / Rev / Zoom dialect ───────────────────────────────────────


def _otter(root: Path, turns: int = 10) -> Path:
    """Speaker on its own line, body on the next -- two people alternating."""
    lines: list[str] = []
    for i in range(turns):
        lines += ["Cliff Creech", f"we need forty cameras at site {i}, escort required."]
        lines += ["Dana Whitfield", f"mid-turn jumpers are excluded at site {i}."]
    target = root / "otter_export.txt"
    target.write_text("\n".join(lines), encoding="utf-8")
    return target


def test_own_line_speakers_route_to_transcript(tmp_path: Path) -> None:
    """Colon density is 0% here, and it is still a transcript.

    Otter, Rev and Zoom put the speaker on its own line. The density rule
    reads that as prose -- a safe landing, but every utterance arrives
    unattributed, which costs speaker_role and the whole meeting_decision
    family.
    """
    assert _parser_name(_otter(tmp_path)) == "TranscriptParser"


def test_folded_utterances_keep_their_true_line_numbers(tmp_path: Path) -> None:
    """The subtle half, and the one that actually bit.

    Folding blanks the speaker line rather than deleting it, so line numbers
    survive. That guarantee was defeated once already: the fold ran inside
    ``normalize_transcript_text``, which strips, and which is called twice on
    the way in -- the second strip removed the leading blank and shifted every
    locator by one. Receipt replay would have failed on precisely the files
    the fold was written to support.
    """
    target = _otter(tmp_path)
    source = target.read_text(encoding="utf-8").splitlines()
    parser, _match, _all = choose_parser(target)
    output = parser.parse_artifact(project_id="p", artifact_id="a", path=target)
    atoms = output.atoms if hasattr(output, "atoms") else output
    assert atoms, "the fold must not cost coverage"

    attributed = 0
    for atom in atoms:
        locator = (atom.source_refs[0].locator if atom.source_refs else {}) or {}
        if locator.get("speaker"):
            attributed += 1
        line = locator.get("line_start")
        if line is None:
            continue
        assert atom.raw_text.strip()[:30].lower() in source[line - 1].lower(), (
            f"locator line {line} holds {source[line - 1]!r}, "
            f"not {atom.raw_text[:40]!r}"
        )
    assert attributed == len(atoms), "every folded utterance must carry its speaker"


@pytest.mark.parametrize(
    "name, body",
    [
        # Names all the way down: every line disqualifies the one above it.
        ("attendee_roster", "\n".join(
            ["Attendees", "Cliff Creech", "Dana Whitfield", "Marcus Lee",
             "Priya Raman", "Tom Alvarez", "Nina Osei"])),
        # Title-case headings are short and capitalised too -- but each
        # appears once, and a conversation is defined by recurrence.
        ("spec_headings", "\n".join(
            f"Section Heading {c}\nContractor shall install conduit and certify each drop."
            for c in "ABCDEFGHIJKLMNOP")),
    ],
)
def test_transcript_lookalikes_do_not_fold(name: str, body: str) -> None:
    """The fold has to be narrow or it becomes the misroute it replaced."""
    from app.core.normalizers import fold_standalone_speaker_lines

    folded, stats = fold_standalone_speaker_lines(body)
    assert not stats["qualifies"], f"{name} folded: {stats}"
    assert folded == body, f"{name} was rewritten despite not qualifying"
