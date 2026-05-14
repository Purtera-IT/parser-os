"""PR1 — operational-workbook routing.

A multi-sheet ops workbook (Asset Inventory + Site Survey + Port/VLAN
+ Risk Register + BOM Detail + …) must NOT get swallowed by the
quote parser. The xlsx parser should win decisively so every sheet
gets its own typed-row profile downstream.
"""
from __future__ import annotations

from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook  # noqa: E402

from app.parsers.registry import choose_parser  # noqa: E402


def test_expanded_operations_workbook_routes_to_xlsx(tmp_path: Path):
    p = tmp_path / "STRESS_MULTI_CAM_expanded_operations_workbook.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["Workbook", "Expanded operational workbook"])
    ws.append(
        ["Contents", "Asset inventory, raw survey data, port/VLAN map, circuits, BOM"]
    )

    sheets = {
        "Asset Inventory": [
            "Asset ID", "Site ID", "Site", "Hostname", "Model", "Serial", "IP",
        ],
        "Port Map & VLANs": [
            "Site ID", "Switch Hostname", "Port", "VLAN ID", "Patch Panel Port",
        ],
        "Risk Register": ["Risk ID", "Severity", "Impact", "Mitigation", "Owner"],
        "BOM Detail": [
            "BOM Line", "Scope Bucket", "Category", "Manufacturer", "SKU",
            "Description", "Qty", "Unit Cost", "Quote Status",
        ],
    }
    for title, headers in sheets.items():
        ws = wb.create_sheet(title)
        ws.append(headers)
        ws.append(["x"] * len(headers))

    wb.save(p)

    parser, match, _ = choose_parser(p)

    assert parser is not None
    assert parser.capability.parser_name == "xlsx"
    assert match.confidence >= 0.95
    assert "xlsx_match:operations_workbook" in match.reasons


def test_pure_quote_xlsx_still_routes_to_quote(tmp_path: Path):
    """Sanity: a workbook that's just a BOM/quote (no asset/site/port
    sheets) still routes to the quote parser."""
    p = tmp_path / "vendor_quote.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.append(["Part Number", "Description", "Qty", "Unit Price", "Lead Time"])
    ws.append(["SKU-001", "Cisco Switch", 2, 1200, "8 weeks"])
    wb.save(p)

    parser, match, _ = choose_parser(p)
    assert parser is not None
    assert parser.capability.parser_name == "quote"
